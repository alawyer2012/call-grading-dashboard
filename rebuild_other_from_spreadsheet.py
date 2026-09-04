#!/usr/bin/env python3
"""
Other-calls rebuild: extract data from '20 Call Other Comparison' spreadsheet
and patch index.html — generate otherRuns / otherAnswerData /
otherRunAnswerData / otherMatrixQuestions, wire leadConfig.other, and remove
'other' from PENDING_TYPES.

Other scorecard (from AI_ QA 2026 (22).xlsx tab 'AI Other'):
  Q1  weight 50  correct_submission  "Based on the transcript did the agent
                                       submit the call correctly"
  Q2  weight 50  probing_qs          "Did the agent use probing questions to
                                       determine the caller's needs?"

No DQs. Score = (Q1_yes*50 + Q2_yes*50) / 100 × 100%.
"""

import openpyxl
import re
import sys
from datetime import date

XLSX_PATH = "/Users/alawyer/Downloads/20 Call Other Comparison.xlsx"
HTML_PATH = "/Users/alawyer/Entrata PM/Dashboard/call-grading/index.html"

# One tab per AI run; extend as new runs land (mirrors resident script).
AI_TABS = [
    {
        "tab": "AI 1.0",
        "id": 1,
        "label": "Run 1.0",
        "date": "September 4, 2026",
        "description": "Baseline — Other-calls scorecard (2 scored questions, 100 pts). 20-call benchmark.",
        "changes": "Initial Other benchmark run (20 matching call IDs).",
    },
]

# New Manual: Col1=Score, Col2=CallId, Col3=SubmittedAs, Col4=CorrectSubmission,
# Col5=Q1 (correct_submission), Col6=Q2 (probing_qs), Col7=notes
MANUAL_COL_MAP = {
    5: "correct_submission",
    6: "probing_qs",
}
MANUAL_META_COLS = {
    3: "submitted_as",
    4: "correct_category",
    7: "notes",
}

# AI 1.0 tab: Col1=call_id, Col2=Q1, Col3=Q2
AI_COL_MAP = {
    2: "correct_submission",
    3: "probing_qs",
}

WEIGHTS = {
    "correct_submission": 50,
    "probing_qs": 50,
}
TOTAL_POINTS = 100
DQ_KEYS = []           # no DQs on Other
SCORED_KEYS = list(WEIGHTS.keys())
ALL_KEYS = SCORED_KEYS + DQ_KEYS

Q_LABELS = {
    "correct_submission": "Submitted correctly",
    "probing_qs": "Probing questions",
}

Q_FULL_LABELS = {
    "correct_submission": "Based on the transcript did the agent submit the call correctly",
    "probing_qs": "Did the agent use probing questions to determine the caller's needs?",
}

STRICT_REASONS = {
    "correct_submission": "Human credited the agent's category call. AI marked it wrong.",
    "probing_qs": "Human heard probing / clarifying questions. AI missed them.",
}
LENIENT_REASONS = {
    "correct_submission": "AI accepted the agent's category at face value; human found it should have been a different category (usually Lead or Resident, not Solicitor).",
    "probing_qs": "AI credited probing on a light call where the agent asked no questions to determine who the caller was.",
}


def parse_yes_no(val):
    if val is None or str(val).strip() == "":
        return None
    return 1 if str(val).strip().lower().startswith("y") else 0


def calc_score(answers):
    earned = sum(WEIGHTS[q] for q in SCORED_KEYS if answers.get(q) == 1)
    return round((earned / TOTAL_POINTS) * 100, 2)


def js_escape(s):
    return (
        s.replace("\\", "\\\\")
        .replace("`", "\\`")
        .replace("$", "\\$")
    )


def js_str_escape(s):
    # Single-quoted string. Escape backslashes then single quotes.
    return s.replace("\\", "\\\\").replace("'", "\\'").replace("\n", "\\n")


print("Reading spreadsheet...")
wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)

# --- Human tab (New Manual) ---
ws_manual = wb["New Manual"]
human_data = {}
for row in range(2, ws_manual.max_row + 1):
    call_id_raw = ws_manual.cell(row, 2).value
    if call_id_raw is None:
        continue
    call_id = str(int(float(call_id_raw)))
    answers = {q: parse_yes_no(ws_manual.cell(row, col).value) for col, q in MANUAL_COL_MAP.items()}
    meta = {name: ws_manual.cell(row, col).value for col, name in MANUAL_META_COLS.items()}
    sheet_score = ws_manual.cell(row, 1).value
    sheet_norm = float(sheet_score) if sheet_score is not None else 0.0
    if sheet_norm > 1.5:
        sheet_norm = sheet_norm / 100.0
    if call_id in human_data:
        # Keep first row; log duplicates.
        print(f"  Duplicate human row for {call_id} — kept first ({human_data[call_id]['sheet_norm']:.2f}, drop {sheet_norm:.2f})")
        continue
    human_data[call_id] = {
        "answers": answers,
        "meta": meta,
        "sheet_score": sheet_score,
        "sheet_norm": sheet_norm,
    }
print(f"Extracted {len(human_data)} unique calls from New Manual tab")

# --- AI tabs ---
ai_runs = {}
for run_info in AI_TABS:
    tab_name = run_info["tab"]
    if tab_name not in wb.sheetnames:
        print(f"⚠️  Tab '{tab_name}' not found — skipping")
        continue
    ws_ai = wb[tab_name]
    ai_data = {}
    for row in range(2, ws_ai.max_row + 1):
        call_id_raw = ws_ai.cell(row, 1).value
        if call_id_raw is None:
            continue
        call_id = str(int(float(call_id_raw)))
        answers = {q: parse_yes_no(ws_ai.cell(row, col).value) for col, q in AI_COL_MAP.items()}
        if call_id in ai_data:
            continue
        ai_data[call_id] = {"answers": answers}
    ai_runs[tab_name] = ai_data
    print(f"Extracted {len(ai_data)} unique calls from '{tab_name}' tab")


def matching_ids(tab_name):
    return sorted(set(human_data.keys()) & set(ai_runs[tab_name].keys()))


benchmark_call_ids = matching_ids(AI_TABS[0]["tab"])
print(f"\nRun 1.0 matching set: {len(benchmark_call_ids)} calls")
print("  " + ", ".join(benchmark_call_ids))

# --- Score validation vs sheet ---
print("\n" + "=" * 70)
print("SCORE VALIDATION — New Manual (benchmark calls)")
print("=" * 70)
mismatches = 0
for call_id in benchmark_call_ids:
    computed = calc_score(human_data[call_id]["answers"])
    sheet_pct = round(human_data[call_id]["sheet_norm"] * 100, 2)
    delta = abs(computed - sheet_pct)
    mark = "OK" if delta <= 2.0 else "MISMATCH"
    if delta > 2.0:
        mismatches += 1
    print(f"  [{mark}] {call_id}: computed={computed}% sheet={sheet_pct}% (delta {delta:.2f})")
print(f"\nScore validation: {mismatches} mismatches out of {len(benchmark_call_ids)} (tolerance +/-2pts)")


def build_run_data(human_data, ai_data, benchmark_ids, run_info=None):
    answer_data = {}
    for call_id in benchmark_ids:
        h = dict(human_data[call_id]["answers"])
        a = dict(ai_data.get(call_id, {}).get("answers", {}))
        answer_data[call_id] = {q: [h.get(q), a.get(q)] for q in ALL_KEYS}

    calls_data = []
    total_agree = 0
    total_disagree = 0
    total_strict = 0
    total_lenient = 0
    total_comparisons = 0
    score_deltas = []

    for call_id in sorted(answer_data.keys()):
        cd = answer_data[call_id]
        h_score = calc_score({q: cd[q][0] for q in ALL_KEYS})
        ai_answers = {q: cd[q][1] for q in ALL_KEYS}
        has_ai = any(v is not None for v in ai_answers.values())
        a_score = calc_score(ai_answers) if has_ai else None

        strict = 0
        lenient = 0
        details = []
        meta = human_data[call_id]["meta"]

        for q in ALL_KEYS:
            h_val, a_val = cd[q]
            if a_val is None or h_val is None:
                continue
            total_comparisons += 1
            if h_val == a_val:
                total_agree += 1
            else:
                total_disagree += 1
                w = WEIGHTS.get(q, 0)
                if h_val == 1 and a_val == 0:
                    strict += 1
                    total_strict += 1
                    reason = STRICT_REASONS.get(q, "AI missed this behavior.")
                    details.append({"q": Q_LABELS[q], "w": w, "ai": "No", "h": "Yes", "reason": reason, "key": q})
                else:
                    lenient += 1
                    total_lenient += 1
                    reason = LENIENT_REASONS.get(q, "AI over-credited this behavior.")
                    if q == "correct_submission":
                        submitted = str(meta.get("submitted_as") or "").strip()
                        correct = str(meta.get("correct_category") or "").strip()
                        if submitted or correct:
                            reason += f" Agent submitted as '{submitted or 'unspecified'}'; correct category '{correct or 'unspecified'}'."
                    details.append({"q": Q_LABELS[q], "w": w, "ai": "Yes", "h": "No", "reason": reason, "key": q})

        weight_lost = sum(d["w"] for d in details)
        if a_score is not None:
            score_deltas.append(abs(h_score - a_score))

        total_disagree_call = strict + lenient
        if total_disagree_call == 0:
            note = "Perfect agreement — AI and human matched on every question."
            rec = "No action needed — perfect agreement."
        else:
            if strict > 0 and lenient > 0:
                note = f"{total_disagree_call} disagreements: {strict} strict, {lenient} lenient."
            elif strict > 0:
                note = f"{strict} disagreement{'s' if strict > 1 else ''} — AI too strict."
            else:
                note = f"{lenient} disagreement{'s' if lenient > 1 else ''} — AI too lenient."
            rec_parts = []
            strict_qs = sorted([d for d in details if d["ai"] == "No"], key=lambda x: -x["w"])
            lenient_qs = sorted([d for d in details if d["ai"] == "Yes"], key=lambda x: -x["w"])
            if strict_qs:
                rec_parts.append("Strict: " + ", ".join(f"{d['q']} ({d['w']}pts)" for d in strict_qs))
            if lenient_qs:
                rec_parts.append("Lenient: " + ", ".join(f"{d['q']} ({d['w']}pts)" for d in lenient_qs))
            rec = ". ".join(rec_parts) + "."

        calls_data.append({
            "id": call_id,
            "human": h_score,
            "ai": a_score if a_score is not None else 0,
            "strict": strict,
            "lenient": lenient,
            "weightLost": weight_lost,
            "disqualifier": False,
            "note": note,
            "details": details,
            "recommendation": rec,
            "meta": meta,
        })

    questions_data = []
    for q_key in ALL_KEYS:
        agree = disagree = strict = lenient = total = 0
        for cd in answer_data.values():
            h, a = cd[q_key]
            if a is None or h is None:
                continue
            total += 1
            if h == a:
                agree += 1
            elif h == 1 and a == 0:
                disagree += 1
                strict += 1
            else:
                disagree += 1
                lenient += 1
        if disagree == 0:
            direction = "agree"
        elif strict > lenient:
            direction = "strict"
        elif lenient > strict:
            direction = "lenient"
        else:
            direction = "mixed"
        questions_data.append({
            "short": q_key,
            "label": Q_FULL_LABELS.get(q_key, Q_LABELS[q_key]),
            "weight": WEIGHTS.get(q_key, 0),
            "agree": agree,
            "disagree": disagree,
            "total": total,
            "strict": strict,
            "lenient": lenient,
            "dir": direction,
        })
    questions_data.sort(key=lambda q: -q["disagree"])

    agreement_pct = round((total_agree / total_comparisons) * 100, 1) if total_comparisons else 0
    avg_delta = round(sum(score_deltas) / len(score_deltas), 1) if score_deltas else 0

    return {
        "answer_data": answer_data,
        "calls": calls_data,
        "questions": questions_data,
        "meta": {
            "agreement": agreement_pct,
            "totalDisagreements": total_disagree,
            "avgDelta": avg_delta,
            "strictErrors": total_strict,
            "lenientErrors": total_lenient,
            "target": 90,
        },
        "total_comparisons": total_comparisons,
    }


all_runs = []
for run_info in AI_TABS:
    tab_name = run_info["tab"]
    if tab_name not in ai_runs:
        continue
    print(f"\n{'=' * 70}")
    print(f"BUILDING {run_info['label']} (tab: {tab_name})")
    print(f"{'=' * 70}")
    run_ids = matching_ids(tab_name)
    rd = build_run_data(human_data, ai_runs[tab_name], run_ids, run_info)
    rd["info"] = run_info
    all_runs.append(rd)
    m = rd["meta"]
    print(f"  Agreement: {m['agreement']}% ({round(m['agreement']*rd['total_comparisons']/100)}/{rd['total_comparisons']})")
    print(f"  Total disagreements: {m['totalDisagreements']} ({m['strictErrors']} strict / {m['lenientErrors']} lenient)")
    print(f"  Avg score delta: {m['avgDelta']}%")
    perfects = sum(1 for c in rd["calls"] if c["strict"] == 0 and c["lenient"] == 0)
    print(f"  Perfect agreement: {perfects} / {len(rd['calls'])} calls")


# --- Root cause + recommendations (hand-authored for Run 1.0 baseline) ---
def author_run_1_narrative(rd):
    m = rd["meta"]
    q1 = next((q for q in rd["questions"] if q["short"] == "correct_submission"), None)
    q2 = next((q for q in rd["questions"] if q["short"] == "probing_qs"), None)
    perfects = sum(1 for c in rd["calls"] if c["strict"] == 0 and c["lenient"] == 0)

    key_findings = (
        f"<p>Baseline Other-calls run against the 2-question scorecard "
        f"(<em>{Q_FULL_LABELS['correct_submission']}</em> and "
        f"<em>{Q_FULL_LABELS['probing_qs']}</em>). "
        f"Agreement is <strong>{m['agreement']}%</strong> "
        f"({round(m['agreement']*rd['total_comparisons']/100)}/{rd['total_comparisons']}) "
        f"with <strong>{m['totalDisagreements']} disagreements</strong> "
        f"(<strong>{m['strictErrors']} strict / {m['lenientErrors']} lenient</strong>). "
        f"Perfect agreement: <strong>{perfects} of {len(rd['calls'])}</strong>. "
        f"Avg score delta <strong>{m['avgDelta']}%</strong>.</p>"
        f"<p>The bias is one-sided: AI is <strong>lenient on every disagreement</strong>. "
        f"There are no strict errors. On the submission question "
        f"({q1['agree']}/{q1['total']} = {round(q1['agree']/q1['total']*100)}%), AI accepts the agent's "
        f"'Solicitor' label at face value on calls where the human graders re-categorize the caller as "
        f"a <em>Lead</em> or, more commonly, a <em>Resident</em> — usually because the transcript reveals "
        f"the caller is a current or past resident (rental verification, package pickup, safety complaint) "
        f"and the agent never asked enough to figure that out. That miscategorization also breaks the "
        f"downstream Resident routing and skews the resident-call denominator.</p>"
        f"<p>On the probing question ({q2['agree']}/{q2['total']} = {round(q2['agree']/q2['total']*100)}%), "
        f"the 3 lenient errors are all cases where the human wrote a note like "
        f"'No Questions were asked to determine who the caller was' — the agent handled the call "
        f"transactionally without asking a single clarifying question, and AI still credited probing.</p>"
        f"<p>→ Open the <strong>Recommendations</strong> tab for protocol tightening on both questions.</p>"
    )

    # Per-question root cause
    root_cause = {
        "correct_submission": (
            "AI trusts the agent's category selection instead of grading the transcript. "
            "All 6 lenient errors are Solicitor submissions where the correct answer was Lead (2) "
            "or Resident (4). The AI needs to independently reason about who the caller is based on "
            "transcript evidence (mentions of a current lease, rent, unit issues, callbacks, "
            "verifications, package pickups from residents, deliveries)."
        ),
        "probing_qs": (
            "AI credits probing whenever the agent asks any question (name, phone, callback preference), "
            "even when the agent never tried to determine WHO the caller was or WHY they were calling. "
            "The 3 lenient errors are calls where the agent moved straight to logistics without any "
            "clarifying question about the caller's relationship to the property."
        ),
    }

    recommendations = [
        {
            "num": 1,
            "title": "Grade submission from the transcript, not the agent's category",
            "severity": "critical",
            "severityLabel": f"{q1['lenient']} lenient errors, all Solicitor→Lead/Resident",
            "owner": "Monica",
            "ownerClass": "warning",
            "problem_html": (
                "<p><strong>All 6 Q1 lenient errors are the same failure mode:</strong> the agent "
                "picked Solicitor, and AI accepted it. The transcript said otherwise every time.</p>"
                "<ul style=\"font-size:13px; line-height:1.7;\">"
                "<li><strong>273801118</strong> — package for a friend who is a current resident → should be <strong>Resident</strong>.</li>"
                "<li><strong>273055082</strong> — rental verification for a current resident → should be <strong>Resident</strong>.</li>"
                "<li><strong>270974570</strong> — broken shower work-order request from a resident → should be <strong>Resident</strong>.</li>"
                "<li><strong>268573697</strong> — resident escalating a safety complaint about her baby → should be <strong>Resident</strong>.</li>"
                "<li><strong>272773798</strong> — inquiry that should have been routed as a <strong>Lead</strong>.</li>"
                "<li><strong>269549141</strong> — inquiry that should have been routed as a <strong>Lead</strong>.</li>"
                "</ul>"
                "<p>Solicitor is a bucket for people trying to sell TO the property. It is not "
                "a fallback for 'agent didn't know what to do with this caller.'</p>"
            ),
            "protocols": [
                {
                    "label": "Q1 Protocol — replace face-value acceptance",
                    "current": (
                        "Reads the agent's submitted category and confirms whether it matches "
                        "the caller's stated intent."
                    ),
                    "recommended": (
                        "<strong>Do not use the agent's category as input.</strong> Read the transcript "
                        "first and independently classify the caller into one of the seven buckets "
                        "(Lead, Resident, Work Order, Market Survey, Solicitor, Vendor, Wrong Number). "
                        "Then compare to the agent's submission.<br><br>"
                        "<strong>Resident signals that override 'Solicitor':</strong> caller mentions "
                        "current or past tenancy, rent, lease, deposit, unit issues, package pickup, "
                        "callbacks about property staff, complaints, rental verification requests, "
                        "or gives a unit number that resolves in the system.<br><br>"
                        "<strong>Lead signals that override 'Solicitor':</strong> caller asks about "
                        "availability, pricing, floor plans, tours, applications, or is a locator/broker.<br><br>"
                        "<strong>Only mark Yes</strong> when the agent's category matches the "
                        "transcript-derived category."
                    ),
                    "labelOverride": "Q1 Recommended Protocol",
                },
            ],
        },
        {
            "num": 2,
            "title": "'Probing questions' means clarifying who the caller is — not just any question",
            "severity": "warning",
            "severityLabel": f"{q2['lenient']} lenient errors on 'no-question' calls",
            "owner": "Monica",
            "ownerClass": "warning",
            "problem_html": (
                "<p><strong>All 3 Q2 lenient errors</strong> came with the same human note: "
                "'no questions were asked to determine who the caller was.' "
                "The agent asked things like <em>What is this in regards to?</em> or "
                "confirmed a callback number — administrative logistics, not probing.</p>"
                "<ul style=\"font-size:13px; line-height:1.7;\">"
                "<li><strong>272954069</strong> — 'No Questions were asked to determine who the caller was'</li>"
                "<li><strong>271843479</strong> — verification call, agent went straight to reading the email</li>"
                "<li><strong>268637585</strong> — 'MD either way, rental verification for a current resident, but the agent did not ask for more info'</li>"
                "</ul>"
                "<p>The AI is treating any interrogative sentence as probing.</p>"
            ),
            "protocols": [
                {
                    "label": "Q2 Protocol — define probing narrowly",
                    "current": (
                        "Mark 'Yes' if the transcript shows the agent identifying the caller's "
                        "needs through probing questions, clarifying questions, or if the caller "
                        "clearly explained their reason for calling. Mark 'No' if the agent quickly "
                        "redirects the caller, requests a callback, or provides limited assistance "
                        "without first attempting to understand the caller's needs."
                    ),
                    "recommended": (
                        "Mark 'Yes' only if <strong>at least one of the following is true</strong>:<br><br>"
                        "1. The agent asked a question intended to identify WHO the caller is or "
                        "WHY they are calling (e.g., 'Are you a current resident?', "
                        "'What is this regarding?', 'Are you calling about your unit?', "
                        "'Are you a prospect?', 'Are you with a company that has a contract with us?'), "
                        "OR<br>"
                        "2. The caller volunteered a clear reason for the call in the first turn.<br><br>"
                        "<strong>Mark 'No'</strong> when the agent only asked administrative questions "
                        "(name, callback number, best time to reach you, email address to send info to) "
                        "without any attempt to determine the caller's identity or purpose. "
                        "Confirming a phone number back or reading an email address is not probing."
                    ),
                    "labelOverride": "Q2 Recommended Protocol",
                },
            ],
        },
        {
            "num": 3,
            "title": "Downstream impact — miscategorized Solicitors pollute the Resident dataset",
            "severity": "info",
            "severityLabel": "4 of 6 Q1 errors are hidden Resident calls",
            "owner": "Austin / Data",
            "ownerClass": "info",
            "problem_html": (
                "<p>4 of the 6 Q1 errors (273801118, 273055082, 270974570, 268573697) are "
                "resident calls submitted as Solicitor. That means:</p>"
                "<ul style=\"font-size:13px; line-height:1.7;\">"
                "<li>The Resident 490-call benchmark is <strong>under-counting resident calls</strong> "
                "by whatever fraction of Solicitor-labelled calls are actually residents.</li>"
                "<li>These calls never got the Resident scorecard (10 Qs + 2 DQs) applied — they "
                "were graded on Other's 2-question card instead, which does not check hold permission, "
                "closing, ownership, or the secure-info DQ.</li>"
                "<li>Fixing Q1 protocol will re-route these calls to the Resident bucket for future "
                "runs. Expect the Resident denominator to grow slightly and Solicitor volume to drop.</li>"
                "</ul>"
                "<p>Recommendation: after Q1 protocol is updated, re-pull the Other set and confirm "
                "how many former Solicitors are now Residents/Leads. Add those to the Resident "
                "benchmark rather than re-scoring them on Other.</p>"
            ),
            "protocols": [],
        },
    ]

    return key_findings, root_cause, recommendations


run1_narrative = author_run_1_narrative(all_runs[0]) if all_runs else (None, None, None)


def build_run_js(run_data, is_run_1):
    info = run_data["info"]
    meta = run_data["meta"]
    calls_data = run_data["calls"]
    questions_data = run_data["questions"]
    n_calls = len(calls_data)

    meta_js = f"""    meta: {{
      agreement: {meta['agreement']},
      totalDisagreements: {meta['totalDisagreements']},
      avgDelta: {meta['avgDelta']},
      strictErrors: {meta['strictErrors']},
      lenientErrors: {meta['lenientErrors']},
      target: 90
    }}"""

    if is_run_1 and run1_narrative[0]:
        key_findings, root_cause, recommendations = run1_narrative
    else:
        # Generic auto-generated
        agreement = meta["agreement"]
        key_findings = (
            f"<p>The AI disagrees with humans on <strong>{round(100 - agreement, 1)}%</strong> "
            f"of answers across these {n_calls} Other calls "
            f"({meta['totalDisagreements']} of {run_data['total_comparisons']} scored comparisons).</p>"
        )
        root_cause = {}
        recommendations = []

    key_findings_js = f"    keyFindings: `{key_findings}`"

    q_lines = []
    for q in questions_data:
        q_lines.append(
            f'      {{ short: "{q["short"]}", label: "{js_str_escape(q["label"])}", '
            f'weight: {q["weight"]}, agree: {q["agree"]}, disagree: {q["disagree"]}, '
            f'total: {q["total"]}, lenient: {q["lenient"]}, strict: {q["strict"]}, '
            f'dir: "{q["dir"]}",\n        evidence: []}}'
        )
    questions_js = "    questions: [\n" + ",\n".join(q_lines) + "\n    ]"

    calls_lines = []
    for c in calls_data:
        details_parts = []
        for d in c["details"]:
            details_parts.append(
                f'          {{ q: "{js_str_escape(d["q"])}", w: {d["w"]}, ai: "{d["ai"]}", h: "{d["h"]}", reason: "{js_str_escape(d["reason"])}" }}'
            )
        details_str = ",\n".join(details_parts)
        if details_str:
            details_str = "\n" + details_str + "\n        "
        calls_lines.append(
            f'      {{ id: "{c["id"]}", human: {c["human"]}, ai: {c["ai"]}, '
            f'strict: {c["strict"]}, lenient: {c["lenient"]}, weightLost: {c["weightLost"]}, '
            f'disqualifier: {"true" if c["disqualifier"] else "false"},\n'
            f"        note: '{js_str_escape(c['note'])}',\n"
            f"        details: [{details_str}],\n"
            f"        recommendation: '{js_str_escape(c['recommendation'])}' }}"
        )
    calls_js = "    calls: [\n" + ",\n".join(calls_lines) + "\n    ]"

    # Recommendations
    rec_lines = []
    for r in recommendations:
        protos_lines = []
        for p in r.get("protocols", []) or []:
            protos_lines.append(
                "          {\n"
                f"            label: \"{js_str_escape(p['label'])}\",\n"
                f"            current: \"{js_str_escape(p['current'])}\",\n"
                f"            recommended: \"{js_str_escape(p['recommended'])}\",\n"
                f"            labelOverride: \"{js_str_escape(p.get('labelOverride', p['label']))}\"\n"
                "          }"
            )
        protos_js = "[\n" + ",\n".join(protos_lines) + "\n        ]" if protos_lines else "[]"
        rec_lines.append(
            "      {\n"
            f"        num: {r['num']},\n"
            f"        title: \"{js_str_escape(r['title'])}\",\n"
            f"        severity: \"{r['severity']}\",\n"
            f"        severityLabel: \"{js_str_escape(r['severityLabel'])}\",\n"
            f"        owner: \"{r['owner']}\",\n"
            f"        ownerClass: \"{r['ownerClass']}\",\n"
            f"        problem: `{r['problem_html']}`,\n"
            f"        protocols: {protos_js}\n"
            "      }"
        )
    recommendations_js = "    recommendations: [\n" + ",\n".join(rec_lines) + "\n    ]"

    # Root cause block
    rc_lines = []
    for key, text in (root_cause or {}).items():
        rc_lines.append(f"      {key}: \"{js_str_escape(text)}\"")
    root_cause_js = "    rootCause: {\n" + ",\n".join(rc_lines) + "\n    }" if rc_lines else "    rootCause: {}"

    return (
        "  {\n"
        f"    id: {info['id']},\n"
        f"    label: \"{js_str_escape(info['label'])}\",\n"
        f"    date: \"{js_str_escape(info['date'])}\",\n"
        f"    description: \"{js_str_escape(info['description'])}\",\n"
        f"    changes: \"{js_str_escape(info['changes'])}\",\n"
        f"{meta_js},\n"
        f"{key_findings_js},\n"
        f"{questions_js},\n"
        f"{calls_js},\n"
        f"{recommendations_js},\n"
        f"{root_cause_js}\n"
        "  }"
    )


# --- Emit JS blocks ---
run_answer_data_parts = []
for rd in all_runs:
    run_id = rd["info"]["id"]
    ad = rd["answer_data"]
    ad_lines = []
    for call_id in sorted(ad.keys()):
        cd = ad[call_id]
        parts = []
        for q in ALL_KEYS:
            h, a = cd[q]
            h_str = str(h) if h is not None else "null"
            a_str = str(a) if a is not None else "null"
            parts.append(f"{q}:[{h_str},{a_str}]")
        ad_lines.append(f'    "{call_id}": {{{",".join(parts)}}}')
    run_answer_data_parts.append(f"  {run_id}: {{\n" + ",\n".join(ad_lines) + "\n  }")

other_run_answer_data_js = "const otherRunAnswerData = {\n" + ",\n".join(run_answer_data_parts) + "\n};"

ad = all_runs[0]["answer_data"]
ad_lines = []
for call_id in sorted(ad.keys()):
    cd = ad[call_id]
    parts = []
    for q in ALL_KEYS:
        h, a = cd[q]
        h_str = str(h) if h is not None else "null"
        a_str = str(a) if a is not None else "null"
        parts.append(f"{q}:[{h_str},{a_str}]")
    ad_lines.append(f'  "{call_id}": {{{",".join(parts)}}}')
other_answer_data_js = "const otherAnswerData = {\n" + ",\n".join(ad_lines) + "\n};"

runs_js_parts = [build_run_js(rd, i == 0) for i, rd in enumerate(all_runs)]
other_runs_js = "const otherRuns = [\n" + ",\n".join(runs_js_parts) + "\n];"

other_matrix_js = """const otherMatrixQuestions = [
  { short: "Submitted correctly", key: "correct_submission" },
  { short: "Probing questions", key: "probing_qs" },
];"""

n_calls = max(len(rd["calls"]) for rd in all_runs) if all_runs else len(benchmark_call_ids)
other_lead_config = f"""  other: {{
    label: 'Other',
    callCount: {n_calls},
    criteriaCount: 2,
    totalPoints: {TOTAL_POINTS},
    scoredQuestions: {len(SCORED_KEYS)},
    runs: otherRuns,
    answerData: otherAnswerData,
    runAnswerData: otherRunAnswerData,
    matrixQuestions: otherMatrixQuestions,
    subtitle: (r) => r ? `${{(r.calls && r.calls.length) ? r.calls.length : {n_calls}}} Other calls \\u00b7 2 scoring criteria \\u00b7 ${{r.label}} \\u2014 ${{r.date}}` : `{n_calls} Other calls \\u00b7 2 scoring criteria \\u00b7 ${{runs.length}} model runs tracked`,
    formulaDesc: 'Two Yes/No questions weighted 50 / 50 (from the AI Other tab of AI_ QA 2026 (22).xlsx). Score = (earned points / 100) \\u00d7 100%. No disqualifiers.',
    dqDesc: 'No disqualifiers apply to Other-category calls.',
  }}"""

# --- Patch HTML ---
print("\nPatching HTML...")
with open(HTML_PATH, "r") as f:
    html = f.read()


def replace_const_object(html_text, const_name, new_js):
    needle = f"const {const_name} = {{"
    if needle not in html_text:
        return html_text, False
    start = html_text.index(needle)
    depth = 0
    i = start + len(needle) - 1
    while i < len(html_text):
        ch = html_text[i]
        if ch == "{":
            depth += 1
        elif ch == "}":
            depth -= 1
            if depth == 0:
                end = i + 1
                if end < len(html_text) and html_text[end] == ";":
                    end += 1
                return html_text[:start] + new_js + html_text[end:], True
        i += 1
    return html_text, False


def replace_const_array(html_text, const_name, new_js):
    needle = f"const {const_name} = ["
    if needle not in html_text:
        return html_text, False
    start = html_text.index(needle)
    depth = 0
    i = start + len(needle) - 1
    while i < len(html_text):
        ch = html_text[i]
        if ch == "[":
            depth += 1
        elif ch == "]":
            depth -= 1
            if depth == 0:
                end = i + 1
                if end < len(html_text) and html_text[end] == ";":
                    end += 1
                return html_text[:start] + new_js + html_text[end:], True
        i += 1
    return html_text, False


block_marker = "// ═══ OTHER DATA ═══"
combined_block = (
    "\n\n" + block_marker + "\n"
    + other_answer_data_js + "\n\n"
    + other_run_answer_data_js + "\n\n"
    + other_runs_js + "\n\n"
    + other_matrix_js + "\n"
)

if block_marker in html:
    # Replace individual consts
    html, _ = replace_const_object(html, "otherAnswerData", other_answer_data_js)
    html, _ = replace_const_object(html, "otherRunAnswerData", other_run_answer_data_js)
    html, _ = replace_const_array(html, "otherRuns", other_runs_js)
    html, _ = replace_const_array(html, "otherMatrixQuestions", other_matrix_js)
    print("Refreshed existing OTHER DATA block")
else:
    # Insert after the resident matrix (or as fallback, warm matrix)
    for anchor in ("const residentMatrixQuestions = [", "const warmMatrixQuestions = ["):
        if anchor in html:
            start = html.index(anchor)
            depth = 0
            i = start + len(anchor) - 1
            while i < len(html):
                if html[i] == "[":
                    depth += 1
                elif html[i] == "]":
                    depth -= 1
                    if depth == 0:
                        end = i + 1
                        if end < len(html) and html[end] == ";":
                            end += 1
                        html = html[:end] + combined_block + html[end:]
                        print(f"Inserted OTHER DATA block after {anchor.split(' = ')[0]}")
                        break
                i += 1
            break

# leadConfig entry
if re.search(r"\n  other:\s*\{[\s\S]*?runs:\s*otherRuns", html):
    m = re.search(r"  other: \{[\s\S]*?\n  \}", html)
    if m:
        html = html[: m.start()] + other_lead_config + html[m.end():]
        print("Replaced other leadConfig entry")
else:
    # Insert into leadConfig before the closing };
    cfg_match = re.search(r"const leadConfig = \{[\s\S]*?\n\};", html)
    if not cfg_match:
        print("ERROR: Could not find leadConfig")
        sys.exit(1)
    resident_m = re.search(r"  resident: \{[\s\S]*?\n  \}", html)
    if resident_m and resident_m.end() < cfg_match.end():
        insert_at = resident_m.end()
        html = html[:insert_at] + ",\n" + other_lead_config + html[insert_at:]
        print("Inserted other leadConfig after resident")
    else:
        warm_m = re.search(r"  warm: \{[\s\S]*?\n  \}", html)
        if warm_m:
            insert_at = warm_m.end()
            html = html[:insert_at] + ",\n" + other_lead_config + html[insert_at:]
            print("Inserted other leadConfig after warm (resident not found)")
        else:
            print("ERROR: Could not find warm or resident block to anchor")
            sys.exit(1)

# Remove 'other' from PENDING_TYPES
pending_patterns = [
    (
        "const PENDING_TYPES = {\n"
        "  workorder: { label: 'Work Orders', icon: '🔧' },\n"
        "  other: { label: 'Other', icon: '📋' },\n"
        "};",
        "const PENDING_TYPES = {\n"
        "  workorder: { label: 'Work Orders', icon: '🔧' },\n"
        "};",
    ),
]
patched_pending = False
for old, new in pending_patterns:
    if old in html:
        html = html.replace(old, new)
        patched_pending = True
        break
if not patched_pending:
    # Regex fallback — drop any line that adds 'other' to PENDING_TYPES
    new_html, n = re.subn(r"\n\s*other: \{ label: 'Other', icon: '📋' \},?", "", html, count=1)
    if n:
        html = new_html
        patched_pending = True
if patched_pending:
    print("Removed 'other' from PENDING_TYPES")
else:
    print("(PENDING_TYPES already has no 'other')")

with open(HTML_PATH, "w") as f:
    f.write(html)

print(f"\nWrote updated HTML to {HTML_PATH}")
print(f"Other runs written: {len(all_runs)}")
for rd in all_runs:
    m = rd["meta"]
    print(
        f"  {rd['info']['label']}: {m['agreement']}% agreement, "
        f"{m['totalDisagreements']} disagreements "
        f"({m['strictErrors']} strict / {m['lenientErrors']} lenient), "
        f"avg delta {m['avgDelta']}%, {len(rd['calls'])} calls"
    )
