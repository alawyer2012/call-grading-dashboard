#!/usr/bin/env python3
"""
Full rebuild: Extract data from '20 Call Cold Lead Comparison' spreadsheet,
generate correct answerData, calls, questions, and meta for the dashboard.
Supports multiple AI run tabs (AI, AI 2, AI 3, ...).
Then patch index.html with the corrected data.
"""

import openpyxl
import json
import re
import sys
from datetime import date

XLSX_PATH = "/Users/alawyer/Downloads/20 Call Cold Lead Comparison (7).xlsx"
HTML_PATH = "/Users/alawyer/Entrata PM/Dashboard/call-grading/index.html"

# Which AI tabs to process (in order). Tab name → run metadata.
AI_TABS = [
    {"tab": "AI",   "id": 1, "label": "Run 1", "date": "June 5, 2026",
     "description": "Baseline — original prompts and protocols, no changes applied.",
     "changes": "Initial benchmark run"},
    {"tab": "AI 2", "id": 2, "label": "Run 2", "date": "June 8, 2026",
     "description": "Protocol & prompt updates — Monica's expanded recognition rules + Myles's transcription handling.",
     "changes": "Recs 1-5 applied: expanded recognition for occupants/pets/how-heard, redefined conversational/rapport, reclassified open-ended detection, lowered bar for feature/amenity + inclusive language, added transcription guidance"},
    {"tab": "AI 3", "id": 3, "label": "Run 3", "date": "June 9, 2026",
     "description": "Structural prompt additions — processing order, volunteered info rule, negative example override.",
     "changes": "Added protocol processing ORDER, VOLUNTEERED INFORMATION RULE, NEGATIVE EXAMPLE OVERRIDE, AVOIDING FALSE CREDIT guard to system prompt"},
    {"tab": "AI 4", "id": 4, "label": "Run 4", "date": "June 10, 2026",
     "description": "Post-correction baseline — first run measured against corrected human scores.",
     "changes": "Human scoring errors identified and corrected (100 answers across 19 calls). All runs re-evaluated against corrected 'New Manual' data."},
    {"tab": "AI 5", "id": 5, "label": "Run 5", "date": "June 30, 2026",
     "description": "Version 5.0 — latest model iteration with accumulated prompt and protocol improvements.",
     "changes": "Full prompt refinement pass incorporating all prior learnings from Runs 1-4"},
]

# ═══════════════════════════════════════════════════════════════════════════
# COLUMN MAPPINGS (from spreadsheet headers, verified above)
# ═══════════════════════════════════════════════════════════════════════════

# Manual tab: Col 1=Overall Score, Col 2=Call Id, Cols 3-26=answers
MANUAL_COL_MAP = {
    3:  "open_ended_qs",
    4:  "pricing_disclaimer",
    5:  "text_email_perm",
    6:  "disclaimers",
    7:  "feature_amenity",
    8:  "tour_offer",
    9:  "email",
    10: "pricing_value",
    11: "greeting",
    12: "tour_next_step",
    13: "acknowledged",
    14: "inclusive_lang",
    15: "name_usage",
    16: "conversational",
    17: "rapport",
    18: "phone",
    19: "occupants",
    20: "pets",
    21: "how_heard",
    22: "pet_breed",
    23: "fha",
    24: "secure_info",
    25: "no_contact",
    26: "closing",
}

# AI tab: Col 1=call_id, Cols 2-25=answers, Col 26=error
AI_COL_MAP = {
    2:  "fha",
    3:  "secure_info",
    4:  "no_contact",
    5:  "greeting",
    6:  "name_usage",
    7:  "conversational",
    8:  "rapport",
    9:  "phone",
    10: "occupants",
    11: "pets",
    12: "pet_breed",
    13: "how_heard",
    14: "open_ended_qs",
    15: "pricing_disclaimer",
    16: "text_email_perm",
    17: "disclaimers",
    18: "feature_amenity",
    19: "tour_offer",
    20: "email",
    21: "pricing_value",
    22: "tour_next_step",
    23: "acknowledged",
    24: "inclusive_lang",
    25: "closing",
}

WEIGHTS = {
    "greeting": 4, "name_usage": 3, "conversational": 3, "rapport": 4,
    "phone": 3, "occupants": 3, "pets": 3, "pet_breed": 2, "how_heard": 3,
    "open_ended_qs": 6, "pricing_disclaimer": 5, "text_email_perm": 2,
    "disclaimers": 5, "feature_amenity": 9, "tour_offer": 4, "email": 2,
    "pricing_value": 3, "tour_next_step": 5, "acknowledged": 6,
    "inclusive_lang": 3, "closing": 2,
}
DQ_KEYS = ["fha", "secure_info", "no_contact"]
SCORED_KEYS = list(WEIGHTS.keys())
ALL_KEYS = SCORED_KEYS + DQ_KEYS

Q_LABELS = {
    "greeting": "Greeting", "name_usage": "Name usage", "conversational": "Conversational",
    "rapport": "Rapport", "phone": "Phone number", "occupants": "Occupants",
    "pets": "Pets", "pet_breed": "Pet breed", "how_heard": "How heard",
    "open_ended_qs": "Open-ended questions", "pricing_disclaimer": "Urgency/pricing disclaimer",
    "text_email_perm": "Text/email permission", "disclaimers": "Required disclaimers",
    "feature_amenity": "Feature/amenity", "tour_offer": "Tour offer", "email": "Email",
    "pricing_value": "Pricing value", "tour_next_step": "Tour/next step",
    "acknowledged": "Acknowledged caller", "inclusive_lang": "Inclusive language",
    "closing": "Closing", "fha": "FHA violation (DQ)", "secure_info": "Secure info (DQ)",
    "no_contact": "No contact attempt (DQ)",
}

Q_FULL_LABELS = {
    "greeting": "Greeting (property name + intro)",
    "name_usage": "Asked for name + used it",
    "conversational": "Conversational info gathering",
    "rapport": "Rapport building / sentence variety",
    "phone": "Phone number gathered",
    "occupants": "Confirmed occupants",
    "pets": "Confirmed pets",
    "pet_breed": "Pet breed (if applicable)",
    "how_heard": "How heard about community",
    "open_ended_qs": "Two open-ended questions",
    "pricing_disclaimer": "Pricing disclaimer / urgency language",
    "text_email_perm": "Text/email permission",
    "disclaimers": "Required disclaimers stated",
    "feature_amenity": "Feature/amenity benefit selling",
    "tour_offer": "Offered a tour",
    "email": "Email address gathered",
    "pricing_value": "Pricing value language (if applicable)",
    "tour_next_step": "Tour scheduled / clear next step",
    "acknowledged": "Acknowledged caller / willingness to help",
    "inclusive_lang": "Inclusive language (we/us)",
    "closing": "Closing / confirm next steps",
}

STRICT_REASONS = {
    "name_usage": "Agent used caller's name but AI didn't detect it — likely transcription garble or exact-string matching failure.",
    "pricing_disclaimer": "Human credited urgency/pricing disclaimer language. AI applied narrow pattern matching.",
    "feature_amenity": "Agent mentioned features/amenities with value framing. AI required more explicit benefit language or missed due to transcription.",
    "pricing_value": "Human credited pricing value language. AI did not detect the phrasing used.",
    "inclusive_lang": "Agent used 'we/our' language but AI missed common phrasings like 'We do have...' or 'Our community...'",
    "how_heard": "Agent asked how caller heard about community. AI missed — possibly due to phrasing or position in call.",
    "open_ended_qs": "Human counted questions as open-ended. AI grammar-parsed too literally and rejected them.",
    "tour_offer": "Agent offered a tour or visit. AI may have required exact 'tour' wording.",
    "email": "Agent gathered email address. AI missed the phrasing used.",
    "occupants": "Occupant info surfaced in conversation. AI required a specific direct question format.",
    "pets": "Pet info surfaced in conversation. AI required agent to ask in a specific format.",
    "phone": "Agent gathered phone number. AI missed the phrasing.",
    "text_email_perm": "Agent asked for text/email permission. AI did not detect the request.",
    "disclaimers": "Agent stated required disclaimers. AI missed due to phrasing variation.",
    "acknowledged": "Agent acknowledged caller's question. AI did not credit it.",
    "closing": "Agent confirmed next steps. AI required more formal closing structure.",
    "conversational": "Human credited conversational gathering. AI did not detect sufficient conversational flow.",
    "rapport": "Human credited rapport building. AI did not detect sufficient personalization.",
    "greeting": "Agent greeted with property name and intro. AI did not detect it.",
    "pet_breed": "Agent asked about pet breed. AI missed.",
    "tour_next_step": "Agent provided a clear next step. AI did not credit it.",
}

LENIENT_REASONS = {
    "conversational": "AI credited info-gathering approach; human required genuine conversational flow beyond scripted exchanges.",
    "rapport": "AI counted agent's tone/variety as rapport; human required specific references to caller's situation.",
    "tour_next_step": "AI credited a vague statement as a next step; human required a definitive action.",
    "pricing_disclaimer": "AI flagged a basic statement as urgency language; human required explicit scarcity phrasing.",
    "pricing_value": "AI credited a basic pricing mention; human required more explicit value framing.",
    "closing": "AI counted a generic goodbye as closing; human required confirmed next steps.",
    "phone": "AI detected phone exchange but human says it wasn't properly gathered.",
    "occupants": "AI detected occupant mention but human says agent didn't properly confirm.",
    "pets": "AI detected pet mention but human says agent didn't properly confirm.",
    "how_heard": "AI detected mention but human says it wasn't properly asked.",
    "email": "AI credited email collection but human disagrees.",
    "disclaimers": "AI credited disclaimers; human did not consider them sufficient.",
    "no_contact": "AI incorrectly flagged no contact attempt; agent did gather contact info.",
    "feature_amenity": "AI credited a generic statement as feature selling; human required specific mentions.",
    "tour_offer": "AI credited a vague offer; human required explicit tour invitation.",
    "name_usage": "AI detected name usage but human says name wasn't properly used.",
    "text_email_perm": "AI credited permission request; human disagrees.",
    "acknowledged": "AI credited acknowledgment; human required more explicit recognition.",
    "inclusive_lang": "AI detected inclusive language; human says usage was insufficient.",
    "open_ended_qs": "AI counted questions as open-ended; human classified them as closed-ended.",
    "greeting": "AI credited greeting; human says it was insufficient.",
    "pet_breed": "AI credited breed inquiry; human disagrees.",
    "fha": "AI flagged an FHA violation; human did not find one.",
    "secure_info": "AI flagged secure info disclosure; human did not find one.",
}


def parse_yes_no(val):
    if val is None or str(val).strip() == "":
        return None
    return 1 if str(val).strip().lower() == "yes" else 0


def calc_score(answers):
    earned = sum(WEIGHTS[q] for q in SCORED_KEYS if answers.get(q) == 1)
    base = (earned / 80) * 100
    dq_count = sum(1 for dq in DQ_KEYS if answers.get(dq) == 1)
    return round(base * max(0, 1 - 0.20 * dq_count), 2)


# ═══════════════════════════════════════════════════════════════════════════
# 1. EXTRACT DATA FROM SPREADSHEET
# ═══════════════════════════════════════════════════════════════════════════

print("Reading spreadsheet...")
wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)

# Extract Manual (human) answers
ws_manual = wb['New Manual']
human_data = {}
for row in range(2, ws_manual.max_row + 1):
    call_id_raw = ws_manual.cell(row, 2).value
    if call_id_raw is None:
        continue
    call_id = str(int(call_id_raw))
    answers = {}
    for col, q_key in MANUAL_COL_MAP.items():
        answers[q_key] = parse_yes_no(ws_manual.cell(row, col).value)
    score = ws_manual.cell(row, 1).value
    human_data[call_id] = {"answers": answers, "sheet_score": score}

print(f"Extracted {len(human_data)} calls from Manual tab")

# Extract AI answers for each run tab
ai_runs = {}
for run_info in AI_TABS:
    tab_name = run_info["tab"]
    if tab_name not in wb.sheetnames:
        print(f"⚠️  Tab '{tab_name}' not found — skipping")
        continue
    ws_ai = wb[tab_name]
    ai_data = {}
    for row in range(2, ws_ai.max_row + 1):
        call_id_raw = ws_ai.cell(row, 1).value
        if call_id_raw is None:
            continue
        call_id = str(int(call_id_raw))
        answers = {}
        for col, q_key in AI_COL_MAP.items():
            answers[q_key] = parse_yes_no(ws_ai.cell(row, col).value)
        ai_data[call_id] = {"answers": answers}
    ai_runs[tab_name] = ai_data
    print(f"Extracted {len(ai_data)} calls from '{tab_name}' tab")

# Use Run 1 call IDs as the benchmark set
benchmark_call_ids = sorted(ai_runs[AI_TABS[0]["tab"]].keys())
print(f"\nBenchmark set: {len(benchmark_call_ids)} calls")

# ═══════════════════════════════════════════════════════════════════════════
# 2. VALIDATE SCORES AGAINST SPREADSHEET
# ═══════════════════════════════════════════════════════════════════════════

print("\n" + "=" * 70)
print("SCORE VALIDATION — Manual tab (benchmark calls only)")
print("=" * 70)

score_mismatches = 0
for call_id in benchmark_call_ids:
    if call_id not in human_data:
        print(f"⚠️  {call_id} not in Manual tab!")
        continue
    computed = calc_score(human_data[call_id]["answers"])
    sheet_raw = human_data[call_id]["sheet_score"]
    if sheet_raw is None:
        continue
    sheet_score = round(float(str(sheet_raw).replace("%", "")), 2)
    if abs(computed - sheet_score) > 0.1:
        print(f"❌ {call_id}: computed={computed}%, sheet={sheet_score}%")
        score_mismatches += 1
    else:
        print(f"✅ {call_id}: {computed}% ✓")

print(f"\nScore validation: {score_mismatches} mismatches out of {len(benchmark_call_ids)}")


# ═══════════════════════════════════════════════════════════════════════════
# 3. BUILD PER-RUN DATA (answerData, calls, questions, meta for each run)
# ═══════════════════════════════════════════════════════════════════════════

def build_run_data(human_data, ai_data, benchmark_ids):
    """Build all computed data for one run: answerData, calls, questions, meta."""
    answer_data = {}
    for call_id in benchmark_ids:
        h = human_data[call_id]["answers"]
        a = ai_data.get(call_id, {}).get("answers", {})
        call_answers = {}
        for q_key in ALL_KEYS:
            h_val = h.get(q_key)
            a_val = a.get(q_key)
            call_answers[q_key] = [h_val, a_val]
        answer_data[call_id] = call_answers

    # Per-call stats
    calls_data = []
    total_agree_scored = 0
    total_disagree_scored = 0
    total_strict_scored = 0
    total_lenient_scored = 0
    total_comparisons_scored = 0
    score_deltas = []

    for call_id in sorted(answer_data.keys()):
        cd = answer_data[call_id]
        h_score = calc_score({q: cd[q][0] for q in ALL_KEYS})
        ai_answers = {q: cd[q][1] for q in ALL_KEYS}
        has_ai = any(v is not None for q, v in ai_answers.items() if q in SCORED_KEYS)
        a_score = calc_score(ai_answers) if has_ai else None

        strict, lenient = 0, 0
        details = []
        scored_agree, scored_disagree = 0, 0

        for q in ALL_KEYS:
            h_val, a_val = cd[q]
            if a_val is None or h_val is None:
                continue
            is_dq = q in DQ_KEYS
            if not is_dq:
                total_comparisons_scored += 1
            if h_val == a_val:
                if not is_dq:
                    scored_agree += 1
                    total_agree_scored += 1
            else:
                if not is_dq:
                    scored_disagree += 1
                    total_disagree_scored += 1
                w = WEIGHTS.get(q, "DQ")
                if h_val == 1 and a_val == 0:
                    strict += 1
                    if not is_dq:
                        total_strict_scored += 1
                    details.append({"q": Q_LABELS[q], "w": "DQ" if is_dq else w, "ai": "No", "h": "Yes",
                                    "reason": STRICT_REASONS.get(q, "AI missed this behavior."), "key": q})
                else:
                    lenient += 1
                    if not is_dq:
                        total_lenient_scored += 1
                    details.append({"q": Q_LABELS[q], "w": "DQ" if is_dq else w, "ai": "Yes", "h": "No",
                                    "reason": LENIENT_REASONS.get(q, "AI over-credited this behavior."), "key": q})

        weight_lost = sum(d["w"] for d in details if d["w"] != "DQ")
        has_dq_disagree = any(d["w"] == "DQ" for d in details)

        if a_score is not None:
            score_deltas.append(abs(h_score - a_score))

        total_disagree_call = strict + lenient
        if total_disagree_call == 0:
            note = "Perfect agreement — AI and human matched on every question."
            rec = "No action needed — perfect agreement."
        else:
            if strict > 0 and lenient > 0:
                note = f"{total_disagree_call} disagreements: {strict} strict, {lenient} lenient."
            elif strict > 0:
                note = f"{strict} disagreement{'s' if strict > 1 else ''} — all AI too strict."
            else:
                note = f"{lenient} disagreement{'s' if lenient > 1 else ''} — all AI too lenient."

            strict_qs = sorted([d for d in details if d["ai"] == "No" and d["w"] != "DQ"], key=lambda x: -x["w"])[:3]
            lenient_qs = sorted([d for d in details if d["ai"] == "Yes" and d["w"] != "DQ"], key=lambda x: -x["w"])[:3]
            rec_parts = []
            if strict_qs:
                rec_parts.append("Strict fixes: " + ", ".join(f"{d['q']} ({d['w']}pts)" for d in strict_qs))
            if lenient_qs:
                rec_parts.append("Lenient fixes: " + ", ".join(f"{d['q']} ({d['w']}pts)" for d in lenient_qs))
            rec = ". ".join(rec_parts) + "."

        calls_data.append({
            "id": call_id, "human": h_score, "ai": a_score if a_score is not None else 0,
            "strict": strict, "lenient": lenient, "weightLost": weight_lost,
            "disqualifier": has_dq_disagree, "note": note, "details": details, "recommendation": rec,
            "incomplete_ai": not has_ai,
        })

    # Per-question stats
    questions_data = []
    for q_key in ALL_KEYS:
        agree, disagree, strict, lenient, total = 0, 0, 0, 0, 0
        for call_id, cd in answer_data.items():
            h, a = cd[q_key]
            if a is None or h is None:
                continue
            total += 1
            if h == a:
                agree += 1
            elif h == 1 and a == 0:
                disagree += 1
                strict += 1
            else:
                disagree += 1
                lenient += 1
        if disagree == 0:
            continue
        if strict > lenient:
            direction = "strict"
        elif lenient > strict:
            direction = "lenient"
        else:
            direction = "mixed"
        label = Q_FULL_LABELS.get(q_key, Q_LABELS[q_key])
        weight = WEIGHTS.get(q_key, 0)
        questions_data.append({
            "short": q_key, "label": label, "weight": weight,
            "agree": agree, "disagree": disagree, "total": total,
            "strict": strict, "lenient": lenient, "dir": direction,
        })
    questions_data.sort(key=lambda q: -q["disagree"])

    # Meta
    agreement_pct = round((total_agree_scored / total_comparisons_scored) * 100, 1) if total_comparisons_scored else 0
    avg_delta = round(sum(score_deltas) / len(score_deltas), 1) if score_deltas else 0

    meta = {
        "agreement": agreement_pct,
        "totalDisagreements": total_disagree_scored,
        "avgDelta": avg_delta,
        "strictErrors": total_strict_scored,
        "lenientErrors": total_lenient_scored,
        "target": 90,
    }

    return {
        "answer_data": answer_data,
        "calls": calls_data,
        "questions": questions_data,
        "meta": meta,
        "total_comparisons": total_comparisons_scored,
    }


# Build each run
all_runs = []
for run_info in AI_TABS:
    tab_name = run_info["tab"]
    if tab_name not in ai_runs:
        continue
    print(f"\n{'=' * 70}")
    print(f"BUILDING {run_info['label']} (tab: {tab_name})")
    print(f"{'=' * 70}")
    rd = build_run_data(human_data, ai_runs[tab_name], benchmark_call_ids)
    rd["info"] = run_info
    all_runs.append(rd)

    m = rd["meta"]
    print(f"  Agreement: {m['agreement']}% ({m['agreement']*rd['total_comparisons']//100}/{rd['total_comparisons']})")
    print(f"  Total disagreements: {m['totalDisagreements']}")
    print(f"  Strict errors: {m['strictErrors']}")
    print(f"  Lenient errors: {m['lenientErrors']}")
    print(f"  Avg score delta: {m['avgDelta']}%")
    print(f"  Questions with disagreements: {len(rd['questions'])}")

# Comparison
if len(all_runs) >= 2:
    r1 = all_runs[0]["meta"]
    r2 = all_runs[1]["meta"]
    print(f"\n{'=' * 70}")
    print("RUN 1 → RUN 2 COMPARISON")
    print(f"{'=' * 70}")
    print(f"  Agreement: {r1['agreement']}% → {r2['agreement']}% ({r2['agreement'] - r1['agreement']:+.1f})")
    print(f"  Disagreements: {r1['totalDisagreements']} → {r2['totalDisagreements']} ({r2['totalDisagreements'] - r1['totalDisagreements']:+d})")
    print(f"  Strict: {r1['strictErrors']} → {r2['strictErrors']} ({r2['strictErrors'] - r1['strictErrors']:+d})")
    print(f"  Lenient: {r1['lenientErrors']} → {r2['lenientErrors']} ({r2['lenientErrors'] - r1['lenientErrors']:+d})")
    print(f"  Avg Delta: {r1['avgDelta']}% → {r2['avgDelta']}% ({r2['avgDelta'] - r1['avgDelta']:+.1f})")

# ═══════════════════════════════════════════════════════════════════════════
# 4. GENERATE JS AND PATCH HTML
# ═══════════════════════════════════════════════════════════════════════════

def js_escape(s):
    return s.replace("\\", "\\\\").replace("'", "\\'").replace('"', '\\"').replace("`", "\\`")


def build_answer_data_js(answer_data):
    """Generate answerData constant (shared — uses Run 1 for backward compat, but we'll make per-run)."""
    ad_lines = []
    for call_id in sorted(answer_data.keys()):
        cd = answer_data[call_id]
        parts = []
        for q in ALL_KEYS:
            h, a = cd[q]
            h_str = str(h) if h is not None else "null"
            a_str = str(a) if a is not None else "null"
            parts.append(f"{q}:[{h_str},{a_str}]")
        ad_lines.append(f'  "{call_id}": {{{",".join(parts)}}}')
    return "const coldAnswerData = {\n" + ",\n".join(ad_lines) + "\n};"


def build_run_js(run_data):
    """Build a single run object in JS (meta, keyFindings, questions, calls)."""
    info = run_data["info"]
    meta = run_data["meta"]
    calls_data = run_data["calls"]
    questions_data = run_data["questions"]

    # Meta
    meta_js = f"""    meta: {{
      agreement: {meta['agreement']},
      totalDisagreements: {meta['totalDisagreements']},
      avgDelta: {meta['avgDelta']},
      strictErrors: {meta['strictErrors']},
      lenientErrors: {meta['lenientErrors']},
      target: 90
    }}"""

    # Key findings
    total_comparisons = run_data["total_comparisons"]
    total_disagree = meta["totalDisagreements"]
    strict_top = sorted([q for q in questions_data if q["strict"] > q["lenient"]], key=lambda x: -x["disagree"])[:5]
    lenient_top = sorted([q for q in questions_data if q["lenient"] > q["strict"]], key=lambda x: -x["disagree"])[:3]

    strict_summary = ", ".join(
        f"<em>{q['label']}</em> ({q['agree']}/{q['total']} = {round(q['agree']/q['total']*100)}%)"
        for q in strict_top[:3]
    )
    lenient_summary = ", ".join(
        f"<em>{q['label']}</em> ({q['agree']}/{q['total']} = {round(q['agree']/q['total']*100)}%)"
        for q in lenient_top[:3]
    )

    bias_pct = round(meta["strictErrors"] / total_disagree * 100) if total_disagree else 0
    dominant = "too strict" if meta["strictErrors"] > meta["lenientErrors"] else "too lenient"
    key_findings = (
        f'<p>The AI disagrees with humans on <strong>{round(100-meta["agreement"], 1)}% of question-level answers</strong> '
        f'across these 20 calls ({total_disagree} of {total_comparisons} comparisons). '
        f'The dominant pattern is the AI being <strong>{dominant}</strong> '
        f'— {bias_pct}% of disagreements ({meta["strictErrors"]} of {total_disagree}).</p>'
        f'<p>Worst strict questions: {strict_summary}. '
        f'Worst lenient questions: {lenient_summary}.</p>'
        f'<p>Average absolute score delta is <strong>{meta["avgDelta"]} percentage points</strong> (weighted).</p>'
    )
    key_findings_js = f"    keyFindings: `{key_findings}`"

    # Questions array
    q_lines = []
    for q in questions_data:
        q_lines.append(
            f'      {{ short: "{q["short"]}", label: "{js_escape(q["label"])}", weight: {q["weight"]}, '
            f'agree: {q["agree"]}, disagree: {q["disagree"]}, total: {q["total"]}, '
            f'lenient: {q["lenient"]}, strict: {q["strict"]}, dir: "{q["dir"]}",\n'
            f'        evidence: []}}'
        )
    questions_js = "    questions: [\n" + ",\n".join(q_lines) + "\n    ]"

    # Calls array
    calls_lines = []
    for c in calls_data:
        details_parts = []
        for d in c["details"]:
            w_js = f'"{d["w"]}"' if d["w"] == "DQ" else str(d["w"])
            details_parts.append(
                f'          {{ q: "{js_escape(d["q"])}", w: {w_js}, ai: "{d["ai"]}", h: "{d["h"]}", reason: "{js_escape(d["reason"])}" }}'
            )
        details_str = ",\n".join(details_parts)
        if details_str:
            details_str = "\n" + details_str + "\n        "
        calls_lines.append(
            f'      {{ id: "{c["id"]}", human: {c["human"]}, ai: {c["ai"]}, strict: {c["strict"]}, lenient: {c["lenient"]}, weightLost: {c["weightLost"]}, disqualifier: {"true" if c["disqualifier"] else "false"},\n'
            f'        note: \'{js_escape(c["note"])}\',\n'
            f'        details: [{details_str}],\n'
            f'        recommendation: \'{js_escape(c["recommendation"])}\' }}'
        )
    calls_js = "    calls: [\n" + ",\n".join(calls_lines) + "\n    ]"

    # Assemble run object
    return f"""  {{
    id: {info['id']},
    label: "{info['label']}",
    date: "{info['date']}",
    description: "{js_escape(info['description'])}",
    changes: "{js_escape(info['changes'])}",
{meta_js},
{key_findings_js},
{questions_js},
{calls_js},
    recommendations: [],
    rootCause: {{}}
  }}"""


# Build per-run answerData objects (the dashboard uses the active run's answerData)
# For backward compat, we'll output a combined answerData that uses Run 1,
# and also add per-run answerData to each run object.
# Actually the dashboard's matrix reads from the global answerData — we need per-run.
# Let's output a `runAnswerData` object keyed by run ID.

run_answer_data_parts = []
for rd in all_runs:
    run_id = rd["info"]["id"]
    ad = rd["answer_data"]
    ad_lines = []
    for call_id in sorted(ad.keys()):
        cd = ad[call_id]
        parts = []
        for q in ALL_KEYS:
            h, a = cd[q]
            h_str = str(h) if h is not None else "null"
            a_str = str(a) if a is not None else "null"
            parts.append(f"{q}:[{h_str},{a_str}]")
        ad_lines.append(f'    "{call_id}": {{{",".join(parts)}}}')
    run_answer_data_parts.append(f"  {run_id}: {{\n" + ",\n".join(ad_lines) + "\n  }")

run_answer_data_js = "const coldRunAnswerData = {\n" + ",\n".join(run_answer_data_parts) + "\n};"

# Also keep the global answerData pointing to Run 1 for backward compat
global_answer_data_js = build_answer_data_js(all_runs[0]["answer_data"])

# Build full runs array
runs_js_parts = [build_run_js(rd) for rd in all_runs]
runs_js = "const coldRuns = [\n" + ",\n".join(runs_js_parts) + "\n];"

# ── Now patch the HTML ──
print(f"\nPatching HTML...")
with open(HTML_PATH, "r") as f:
    html = f.read()

# Replace answerData
old_ad_start = html.index("const coldAnswerData = {")
old_ad_end = html.index("};", old_ad_start) + 2
html = html[:old_ad_start] + global_answer_data_js + html[old_ad_end:]
print("✅ Replaced coldAnswerData")

if "const coldRunAnswerData = {" in html:
    old_rad_start = html.index("const coldRunAnswerData = {")
    old_rad_end = html.index("};", old_rad_start) + 2
    html = html[:old_rad_start] + run_answer_data_js + html[old_rad_end:]
else:
    insert_after = html.index("};", html.index("const coldAnswerData = {")) + 2
    html = html[:insert_after] + "\n\n" + run_answer_data_js + html[insert_after:]
print("✅ Inserted/replaced coldRunAnswerData")

old_runs_start = html.index("const coldRuns = [")
bracket_depth = 0
found = False
for i in range(old_runs_start + len("const coldRuns = ["), len(html)):
    if html[i] == '[':
        bracket_depth += 1
    elif html[i] == ']':
        if bracket_depth == 0:
            old_runs_end = i + 2
            break
        bracket_depth -= 1
html = html[:old_runs_start] + runs_js + html[old_runs_end:]
print("✅ Replaced coldRuns array")

with open(HTML_PATH, "w") as f:
    f.write(html)
print(f"\n✅ Wrote updated HTML to {HTML_PATH}")
print(f"   Runs: {len(all_runs)}")
for rd in all_runs:
    m = rd["meta"]
    print(f"   {rd['info']['label']}: {m['agreement']}% agreement, {m['totalDisagreements']} disagreements, {m['avgDelta']}% avg delta")
