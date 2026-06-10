#!/usr/bin/env python3
"""
Warm lead rebuild: Extract data from '20 Call Warm Lead Comparison' spreadsheet,
generate warmRuns, warmAnswerData, warmRunAnswerData for the dashboard.
Then patch index.html with the warm lead data.
"""

import openpyxl
import json
import re
import sys
from datetime import date

XLSX_PATH = "/Users/alawyer/Downloads/20 Call Warm Lead Comparison (1).xlsx"
HTML_PATH = "/Users/alawyer/Entrata PM/Dashboard/call-grading/index.html"

AI_TABS = [
    {"tab": "AI",   "id": 1, "label": "Run 1", "date": date.today().strftime("%B %-d, %Y"),
     "description": "Baseline — original prompts and protocols for warm leads.",
     "changes": "Initial warm lead benchmark run"},
]

# Manual tab: Col 3=Call Id, Col 10=Overall Score, Cols 11-26=answers
MANUAL_COL_MAP = {
    11: "disclaimers",
    12: "tour_offer",
    13: "email",
    14: "acknowledged",
    15: "inclusive_lang",
    16: "closing",
    17: "greeting",
    18: "name_usage",
    19: "conversational",
    20: "rapport",
    21: "phone",
    22: "open_ended_qs",
    23: "pricing_disclaimer",
    24: "text_email_perm",
    25: "fha",
    26: "secure_info",
}

# AI tab: Col 1=call_id, Cols 2-17=answers, Col 18=error
AI_COL_MAP = {
    2:  "greeting",
    3:  "name_usage",
    4:  "conversational",
    5:  "rapport",
    6:  "phone",
    7:  "open_ended_qs",
    8:  "pricing_disclaimer",
    9:  "text_email_perm",
    10: "disclaimers",
    11: "tour_offer",
    12: "email",
    13: "acknowledged",
    14: "inclusive_lang",
    15: "closing",
    16: "fha",
    17: "secure_info",
}

WEIGHTS = {
    "greeting": 4, "name_usage": 3, "conversational": 3, "rapport": 4,
    "phone": 3, "open_ended_qs": 6, "pricing_disclaimer": 5,
    "disclaimers": 5, "tour_offer": 4,
    "email": 2, "acknowledged": 6, "inclusive_lang": 3, "closing": 2,
}
TOTAL_POINTS = 13
# text_email_perm excluded entirely: the protocol says "If this was never offered,
# always answer YES" — AI follows this (always Yes), humans mark No (meaning N/A).
# This creates 20/20 false disagreements. Excluded from scoring AND comparison.
DQ_KEYS = ["fha", "secure_info"]
SCORED_KEYS = list(WEIGHTS.keys())
ALL_KEYS = SCORED_KEYS + DQ_KEYS

Q_LABELS = {
    "greeting": "Greeting", "name_usage": "Name usage", "conversational": "Conversational",
    "rapport": "Rapport", "phone": "Phone number", "open_ended_qs": "Open-ended questions",
    "pricing_disclaimer": "Pricing disclaimer",
    "disclaimers": "Required disclaimers", "tour_offer": "Tour offer", "email": "Email",
    "acknowledged": "Acknowledged/ownership", "inclusive_lang": "Inclusive language",
    "closing": "Closing", "fha": "FHA violation (DQ)", "secure_info": "Secure info (DQ)",
}

Q_FULL_LABELS = {
    "greeting": "Greeting (property name + intro)",
    "name_usage": "Asked for name + used it",
    "conversational": "Conversational info gathering",
    "rapport": "Rapport building / sentence variety",
    "phone": "Phone number gathered",
    "open_ended_qs": "Open-ended questions (2+ distinct)",
    "pricing_disclaimer": "Pricing disclaimer / legal language",
    "disclaimers": "Required disclaimers stated",
    "tour_offer": "Offered a tour",
    "email": "Email address gathered",
    "acknowledged": "Acknowledged caller / took ownership",
    "inclusive_lang": "Inclusive language (we/us)",
    "closing": "Closing / confirm next steps",
}

STRICT_REASONS = {
    "name_usage": "Agent used caller's name but AI didn't detect it.",
    "pricing_disclaimer": "Human credited pricing disclaimer. AI applied narrow pattern matching.",
    "inclusive_lang": "Agent used 'we/our' language but AI missed common phrasings.",
    "open_ended_qs": "Human counted questions as open-ended. AI grammar-parsed too literally.",
    "tour_offer": "Agent offered a tour or visit. AI may have required exact wording.",
    "email": "Agent gathered email address. AI missed the phrasing used.",
    "phone": "Agent gathered phone number. AI missed the phrasing.",
    "disclaimers": "Agent stated required disclaimers. AI missed due to phrasing variation.",
    "acknowledged": "Agent acknowledged caller's question. AI did not credit it.",
    "closing": "Agent confirmed next steps. AI required more formal closing structure.",
    "conversational": "Human credited conversational gathering. AI did not detect sufficient flow.",
    "rapport": "Human credited rapport building. AI did not detect sufficient personalization.",
    "greeting": "Agent greeted with property name and intro. AI did not detect it.",
    "tour_offer": "Agent offered a tour. AI did not credit it.",
    "feature_amenity": "Agent mentioned features/amenities. AI required more explicit language.",
}

LENIENT_REASONS = {
    "conversational": "AI credited info-gathering; human required genuine conversational flow.",
    "rapport": "AI counted tone as rapport; human required specific references to caller's situation.",
    "closing": "AI counted a generic goodbye; human required confirmed next steps.",
    "phone": "AI detected phone exchange but human says it wasn't properly gathered.",
    "email": "AI credited email collection but human disagrees.",
    "disclaimers": "AI credited disclaimers; human did not consider them sufficient.",
    "tour_offer": "AI credited a vague offer; human required explicit tour invitation.",
    "name_usage": "AI detected name usage but human says name wasn't properly used.",
    "acknowledged": "AI credited acknowledgment; human required more explicit recognition.",
    "inclusive_lang": "AI detected inclusive language; human says usage was insufficient.",
    "open_ended_qs": "AI counted questions as open-ended; human classified them as closed-ended.",
    "greeting": "AI credited greeting; human says it was insufficient.",
    "pricing_disclaimer": "AI flagged a basic statement; human required explicit disclaimer.",
    "fha": "AI flagged an FHA violation; human did not find one.",
    "secure_info": "AI flagged secure info disclosure; human did not find one.",
}


def parse_yes_no(val):
    if val is None or str(val).strip() == "":
        return None
    return 1 if str(val).strip().lower() == "yes" else 0


def calc_score(answers):
    # Unweighted: Yes count / 13. No DQ penalties (warm lead sheet doesn't apply them).
    yes_count = sum(1 for q in SCORED_KEYS if answers.get(q) == 1)
    return round((yes_count / TOTAL_POINTS) * 100, 2)


print("Reading spreadsheet...")
wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)

ws_manual = wb['Manual']
human_data = {}
for row in range(2, ws_manual.max_row + 1):
    call_id_raw = ws_manual.cell(row, 3).value  # Col 3 = Call Id
    if call_id_raw is None:
        continue
    call_id = str(int(call_id_raw))
    answers = {}
    for col, q_key in MANUAL_COL_MAP.items():
        answers[q_key] = parse_yes_no(ws_manual.cell(row, col).value)
    score = ws_manual.cell(row, 10).value  # Col 10 = Overall Score
    human_data[call_id] = {"answers": answers, "sheet_score": score}

print(f"Extracted {len(human_data)} calls from Manual tab")

ai_runs = {}
for run_info in AI_TABS:
    tab_name = run_info["tab"]
    if tab_name not in wb.sheetnames:
        print(f"⚠️  Tab '{tab_name}' not found — skipping")
        continue
    ws_ai = wb[tab_name]
    ai_data = {}
    for row in range(2, ws_ai.max_row + 1):
        call_id_raw = ws_ai.cell(row, 1).value
        if call_id_raw is None:
            continue
        call_id = str(int(call_id_raw))
        answers = {}
        for col, q_key in AI_COL_MAP.items():
            answers[q_key] = parse_yes_no(ws_ai.cell(row, col).value)
        ai_data[call_id] = {"answers": answers}
    ai_runs[tab_name] = ai_data
    print(f"Extracted {len(ai_data)} calls from '{tab_name}' tab")

benchmark_call_ids = sorted(ai_runs[AI_TABS[0]["tab"]].keys())
print(f"\nBenchmark set: {len(benchmark_call_ids)} calls")

# ═══════════════════════════════════════════════════════════════════════════
# VALIDATE SCORES
# ═══════════════════════════════════════════════════════════════════════════

print("\n" + "=" * 70)
print("SCORE VALIDATION — Manual tab (benchmark calls only)")
print("=" * 70)

score_mismatches = 0
for call_id in benchmark_call_ids:
    if call_id not in human_data:
        print(f"⚠️  {call_id} not in Manual tab!")
        continue
    computed = calc_score(human_data[call_id]["answers"])
    sheet_raw = human_data[call_id]["sheet_score"]
    if sheet_raw is None:
        continue
    sheet_str = str(sheet_raw).replace("%", "")
    sheet_score = round(float(sheet_str) * (100 if float(sheet_str) <= 1 else 1), 2)
    if abs(computed - sheet_score) > 0.5:
        print(f"❌ {call_id}: computed={computed}%, sheet={sheet_score}%")
        score_mismatches += 1
    else:
        print(f"✅ {call_id}: {computed}% ✓")

print(f"\nScore validation: {score_mismatches} mismatches out of {len(benchmark_call_ids)}")


# ═══════════════════════════════════════════════════════════════════════════
# BUILD PER-RUN DATA
# ═══════════════════════════════════════════════════════════════════════════

def build_run_data(human_data, ai_data, benchmark_ids):
    answer_data = {}
    for call_id in benchmark_ids:
        h = human_data[call_id]["answers"]
        a = ai_data.get(call_id, {}).get("answers", {})
        call_answers = {}
        for q_key in ALL_KEYS:
            h_val = h.get(q_key)
            a_val = a.get(q_key)
            call_answers[q_key] = [h_val, a_val]
        answer_data[call_id] = call_answers

    calls_data = []
    total_agree_scored = 0
    total_disagree_scored = 0
    total_strict_scored = 0
    total_lenient_scored = 0
    total_comparisons_scored = 0
    score_deltas = []

    for call_id in sorted(answer_data.keys()):
        cd = answer_data[call_id]
        h_score = calc_score({q: cd[q][0] for q in ALL_KEYS})
        ai_answers = {q: cd[q][1] for q in ALL_KEYS}
        has_ai = any(v is not None for q, v in ai_answers.items() if q in SCORED_KEYS)
        a_score = calc_score(ai_answers) if has_ai else None

        strict, lenient = 0, 0
        details = []
        scored_agree, scored_disagree = 0, 0

        for q in ALL_KEYS:
            h_val, a_val = cd[q]
            if a_val is None or h_val is None:
                continue
            is_dq = q in DQ_KEYS
            if not is_dq:
                total_comparisons_scored += 1
            if h_val == a_val:
                if not is_dq:
                    scored_agree += 1
                    total_agree_scored += 1
            else:
                if not is_dq:
                    scored_disagree += 1
                    total_disagree_scored += 1
                w = WEIGHTS.get(q, "DQ")
                if h_val == 1 and a_val == 0:
                    strict += 1
                    if not is_dq:
                        total_strict_scored += 1
                    details.append({"q": Q_LABELS[q], "w": "DQ" if is_dq else w, "ai": "No", "h": "Yes",
                                    "reason": STRICT_REASONS.get(q, "AI missed this behavior."), "key": q})
                else:
                    lenient += 1
                    if not is_dq:
                        total_lenient_scored += 1
                    details.append({"q": Q_LABELS[q], "w": "DQ" if is_dq else w, "ai": "Yes", "h": "No",
                                    "reason": LENIENT_REASONS.get(q, "AI over-credited this behavior."), "key": q})

        weight_lost = sum(d["w"] for d in details if d["w"] != "DQ")
        has_dq_disagree = any(d["w"] == "DQ" for d in details)

        if a_score is not None:
            score_deltas.append(abs(h_score - a_score))

        total_disagree_call = strict + lenient
        if total_disagree_call == 0:
            note = "Perfect agreement — AI and human matched on every question."
            rec = "No action needed — perfect agreement."
        else:
            if strict > 0 and lenient > 0:
                note = f"{total_disagree_call} disagreements: {strict} strict, {lenient} lenient."
            elif strict > 0:
                note = f"{strict} disagreement{'s' if strict > 1 else ''} — all AI too strict."
            else:
                note = f"{lenient} disagreement{'s' if lenient > 1 else ''} — all AI too lenient."

            strict_qs = sorted([d for d in details if d["ai"] == "No" and d["w"] != "DQ"], key=lambda x: -x["w"])[:3]
            lenient_qs = sorted([d for d in details if d["ai"] == "Yes" and d["w"] != "DQ"], key=lambda x: -x["w"])[:3]
            rec_parts = []
            if strict_qs:
                rec_parts.append("Strict fixes: " + ", ".join(f"{d['q']} ({d['w']}pts)" for d in strict_qs))
            if lenient_qs:
                rec_parts.append("Lenient fixes: " + ", ".join(f"{d['q']} ({d['w']}pts)" for d in lenient_qs))
            rec = ". ".join(rec_parts) + "."

        calls_data.append({
            "id": call_id, "human": h_score, "ai": a_score if a_score is not None else 0,
            "strict": strict, "lenient": lenient, "weightLost": weight_lost,
            "disqualifier": has_dq_disagree, "note": note, "details": details, "recommendation": rec,
            "incomplete_ai": not has_ai,
        })

    questions_data = []
    for q_key in ALL_KEYS:
        agree, disagree, strict, lenient, total = 0, 0, 0, 0, 0
        for call_id, cd in answer_data.items():
            h, a = cd[q_key]
            if a is None or h is None:
                continue
            total += 1
            if h == a:
                agree += 1
            elif h == 1 and a == 0:
                disagree += 1
                strict += 1
            else:
                disagree += 1
                lenient += 1
        if disagree == 0:
            continue
        if strict > lenient:
            direction = "strict"
        elif lenient > strict:
            direction = "lenient"
        else:
            direction = "mixed"
        label = Q_FULL_LABELS.get(q_key, Q_LABELS[q_key])
        weight = WEIGHTS.get(q_key, 0)
        questions_data.append({
            "short": q_key, "label": label, "weight": weight,
            "agree": agree, "disagree": disagree, "total": total,
            "strict": strict, "lenient": lenient, "dir": direction,
        })
    questions_data.sort(key=lambda q: -q["disagree"])

    agreement_pct = round((total_agree_scored / total_comparisons_scored) * 100, 1) if total_comparisons_scored else 0
    avg_delta = round(sum(score_deltas) / len(score_deltas), 1) if score_deltas else 0

    meta = {
        "agreement": agreement_pct,
        "totalDisagreements": total_disagree_scored,
        "avgDelta": avg_delta,
        "strictErrors": total_strict_scored,
        "lenientErrors": total_lenient_scored,
        "target": 90,
    }

    return {
        "answer_data": answer_data,
        "calls": calls_data,
        "questions": questions_data,
        "meta": meta,
        "total_comparisons": total_comparisons_scored,
    }


all_runs = []
for run_info in AI_TABS:
    tab_name = run_info["tab"]
    if tab_name not in ai_runs:
        continue
    print(f"\n{'=' * 70}")
    print(f"BUILDING {run_info['label']} (tab: {tab_name})")
    print(f"{'=' * 70}")
    rd = build_run_data(human_data, ai_runs[tab_name], benchmark_call_ids)
    rd["info"] = run_info
    all_runs.append(rd)

    m = rd["meta"]
    print(f"  Agreement: {m['agreement']}% ({m['agreement']*rd['total_comparisons']//100}/{rd['total_comparisons']})")
    print(f"  Total disagreements: {m['totalDisagreements']}")
    print(f"  Strict errors: {m['strictErrors']}")
    print(f"  Lenient errors: {m['lenientErrors']}")
    print(f"  Avg score delta: {m['avgDelta']}%")
    print(f"  Questions with disagreements: {len(rd['questions'])}")


# ═══════════════════════════════════════════════════════════════════════════
# GENERATE JS AND PATCH HTML
# ═══════════════════════════════════════════════════════════════════════════

def js_escape(s):
    return s.replace("\\", "\\\\").replace("'", "\\'").replace('"', '\\"').replace("`", "\\`")


def build_run_js(run_data):
    info = run_data["info"]
    meta = run_data["meta"]
    calls_data = run_data["calls"]
    questions_data = run_data["questions"]

    meta_js = f"""    meta: {{
      agreement: {meta['agreement']},
      totalDisagreements: {meta['totalDisagreements']},
      avgDelta: {meta['avgDelta']},
      strictErrors: {meta['strictErrors']},
      lenientErrors: {meta['lenientErrors']},
      target: 90
    }}"""

    total_comparisons = run_data["total_comparisons"]
    total_disagree = meta["totalDisagreements"]
    strict_top = sorted([q for q in questions_data if q["strict"] > q["lenient"]], key=lambda x: -x["disagree"])[:5]
    lenient_top = sorted([q for q in questions_data if q["lenient"] > q["strict"]], key=lambda x: -x["disagree"])[:3]

    strict_summary = ", ".join(
        f"<em>{q['label']}</em> ({q['agree']}/{q['total']} = {round(q['agree']/q['total']*100)}%)"
        for q in strict_top[:3]
    )
    lenient_summary = ", ".join(
        f"<em>{q['label']}</em> ({q['agree']}/{q['total']} = {round(q['agree']/q['total']*100)}%)"
        for q in lenient_top[:3]
    )

    bias_pct = round(meta["strictErrors"] / total_disagree * 100) if total_disagree else 0
    dominant = "too strict" if meta["strictErrors"] > meta["lenientErrors"] else "too lenient"
    key_findings = (
        f'<p>The AI disagrees with humans on <strong>{round(100-meta["agreement"], 1)}% of question-level answers</strong> '
        f'across these 20 calls ({total_disagree} of {total_comparisons} comparisons). '
        f'The dominant pattern is the AI being <strong>{dominant}</strong> '
        f'— {bias_pct}% of disagreements ({meta["strictErrors"]} of {total_disagree}).</p>'
        f'<p>Worst strict questions: {strict_summary}. '
        f'Worst lenient questions: {lenient_summary}.</p>'
        f'<p>Average absolute score delta is <strong>{meta["avgDelta"]} percentage points</strong> (weighted).</p>'
    )
    key_findings_js = f"    keyFindings: `{key_findings}`"

    q_lines = []
    for q in questions_data:
        q_lines.append(
            f'      {{ short: "{q["short"]}", label: "{js_escape(q["label"])}", weight: {q["weight"]}, '
            f'agree: {q["agree"]}, disagree: {q["disagree"]}, total: {q["total"]}, '
            f'lenient: {q["lenient"]}, strict: {q["strict"]}, dir: "{q["dir"]}",\n'
            f'        evidence: []}}'
        )
    questions_js = "    questions: [\n" + ",\n".join(q_lines) + "\n    ]"

    calls_lines = []
    for c in calls_data:
        details_parts = []
        for d in c["details"]:
            w_js = f'"{d["w"]}"' if d["w"] == "DQ" else str(d["w"])
            details_parts.append(
                f'          {{ q: "{js_escape(d["q"])}", w: {w_js}, ai: "{d["ai"]}", h: "{d["h"]}", reason: "{js_escape(d["reason"])}" }}'
            )
        details_str = ",\n".join(details_parts)
        if details_str:
            details_str = "\n" + details_str + "\n        "
        calls_lines.append(
            f'      {{ id: "{c["id"]}", human: {c["human"]}, ai: {c["ai"]}, strict: {c["strict"]}, lenient: {c["lenient"]}, weightLost: {c["weightLost"]}, disqualifier: {"true" if c["disqualifier"] else "false"},\n'
            f'        note: \'{js_escape(c["note"])}\',\n'
            f'        details: [{details_str}],\n'
            f'        recommendation: \'{js_escape(c["recommendation"])}\' }}'
        )
    calls_js = "    calls: [\n" + ",\n".join(calls_lines) + "\n    ]"

    return f"""  {{
    id: {info['id']},
    label: "{info['label']}",
    date: "{info['date']}",
    description: "{js_escape(info['description'])}",
    changes: "{js_escape(info['changes'])}",
{meta_js},
{key_findings_js},
{questions_js},
{calls_js},
    recommendations: [],
    rootCause: {{}}
  }}"""


run_answer_data_parts = []
for rd in all_runs:
    run_id = rd["info"]["id"]
    ad = rd["answer_data"]
    ad_lines = []
    for call_id in sorted(ad.keys()):
        cd = ad[call_id]
        parts = []
        for q in ALL_KEYS:
            h, a = cd[q]
            h_str = str(h) if h is not None else "null"
            a_str = str(a) if a is not None else "null"
            parts.append(f"{q}:[{h_str},{a_str}]")
        ad_lines.append(f'    "{call_id}": {{{",".join(parts)}}}')
    run_answer_data_parts.append(f"  {run_id}: {{\n" + ",\n".join(ad_lines) + "\n  }")

warm_run_answer_data_js = "const warmRunAnswerData = {\n" + ",\n".join(run_answer_data_parts) + "\n};"

# Global warmAnswerData = Run 1 data
ad = all_runs[0]["answer_data"]
ad_lines = []
for call_id in sorted(ad.keys()):
    cd = ad[call_id]
    parts = []
    for q in ALL_KEYS:
        h, a = cd[q]
        h_str = str(h) if h is not None else "null"
        a_str = str(a) if a is not None else "null"
        parts.append(f"{q}:[{h_str},{a_str}]")
    ad_lines.append(f'  "{call_id}": {{{",".join(parts)}}}')
warm_answer_data_js = "const warmAnswerData = {\n" + ",\n".join(ad_lines) + "\n};"

runs_js_parts = [build_run_js(rd) for rd in all_runs]
warm_runs_js = "const warmRuns = [\n" + ",\n".join(runs_js_parts) + "\n];"

# ── Patch the HTML ──
print(f"\nPatching HTML...")
with open(HTML_PATH, "r") as f:
    html = f.read()

# Replace or insert warmAnswerData
if "const warmAnswerData = {" in html:
    old_start = html.index("const warmAnswerData = {")
    old_end = html.index("};", old_start) + 2
    html = html[:old_start] + warm_answer_data_js + html[old_end:]
    print("✅ Replaced warmAnswerData")
else:
    marker = "// ═══ WARM LEAD DATA ═══"
    if marker in html:
        insert_pos = html.index(marker) + len(marker)
        html = html[:insert_pos] + "\n" + warm_answer_data_js + html[insert_pos:]
    else:
        print("❌ Could not find warm data insertion point. Add '// ═══ WARM LEAD DATA ═══' marker to index.html first.")
        sys.exit(1)
    print("✅ Inserted warmAnswerData")

# Replace or insert warmRunAnswerData
if "const warmRunAnswerData = {" in html:
    old_start = html.index("const warmRunAnswerData = {")
    old_end = html.index("};", old_start) + 2
    html = html[:old_start] + warm_run_answer_data_js + html[old_end:]
    print("✅ Replaced warmRunAnswerData")
else:
    insert_after = html.index("};", html.index("const warmAnswerData = {")) + 2
    html = html[:insert_after] + "\n\n" + warm_run_answer_data_js + html[insert_after:]
    print("✅ Inserted warmRunAnswerData")

# Replace or insert warmRuns
if "const warmRuns = [" in html:
    old_start = html.index("const warmRuns = [")
    bracket_depth = 0
    for i in range(old_start + len("const warmRuns = ["), len(html)):
        if html[i] == '[':
            bracket_depth += 1
        elif html[i] == ']':
            if bracket_depth == 0:
                old_end = i + 2
                break
            bracket_depth -= 1
    html = html[:old_start] + warm_runs_js + html[old_end:]
    print("✅ Replaced warmRuns")
else:
    insert_after = html.index("};", html.index("const warmRunAnswerData = {")) + 2
    html = html[:insert_after] + "\n\n" + warm_runs_js + html[insert_after:]
    print("✅ Inserted warmRuns")

with open(HTML_PATH, "w") as f:
    f.write(html)
print(f"\n✅ Wrote updated HTML to {HTML_PATH}")
print(f"   Warm lead runs: {len(all_runs)}")
for rd in all_runs:
    m = rd["meta"]
    print(f"   {rd['info']['label']}: {m['agreement']}% agreement, {m['totalDisagreements']} disagreements, {m['avgDelta']}% avg delta")
