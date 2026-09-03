#!/usr/bin/env python3
"""
Append (or refresh) the Residents → Large Test dashboard runs from the
July/August 490-call comparison tabs.

Two runs today:
  - Large Test   (id 90) — AI JulyAugust Resident Simuluat (baseline / file 17 protocols)
  - Large Test 2 (id 91) — Sheet14 (file 20 protocols: I3 last-name drop, I6 revert, I8 expansion, I11 written)

Both share the same human/manual tab and column layout. Only the AI tab changes.

Does NOT rebuild Runs 1–8. Hand-written recs / rootCause on those runs stay put.

Source grades: ~/Downloads/20 Call Resident Comparison (12).xlsx
Live protocols (column I / Other Prompt): ~/Downloads/AI_ QA 2026 (20).xlsx
  - Sheet: AI Resident Fundamentals
  - Column H (Protocols) is empty. Recs quote column I, not the question stem.
"""

import openpyxl
import re
from datetime import date

XLSX_PATH = "/Users/alawyer/Downloads/20 Call Resident Comparison (12).xlsx"
HTML_PATH = "/Users/alawyer/Entrata PM/Dashboard/call-grading/index.html"

MANUAL_TAB = "Manual Grades JulyAugust Reside"

# Each entry becomes one Large Test run. Order matters — the runs are emitted
# in this order after residentRuns baseline.
LARGE_TESTS = [
    {
        "run_id": 90,
        "label": "Large Test",
        "date": "September 1, 2026",
        "ai_tab": "AI JulyAugust Resident Simuluat",
        "protocol_file": "AI_ QA 2026 (17).xlsx",
        "description": "July/August resident simulation — 490 matching human vs AI grades on the overlapping 9-question card (reason-for-call scorecard, not the 20-call protocol series).",
        "changes": "New 490-call eval, not a protocol run. Compared Manual Grades JulyAugust Reside vs AI JulyAugust Resident Simuluat. Shared questions only (37 pts). Human-only final-closing and AI-only neutral language excluded.",
    },
    {
        "run_id": 91,
        "label": "Large Test 2",
        "date": "September 3, 2026",
        "ai_tab": "Sheet14",
        "protocol_file": "AI_ QA 2026 (20).xlsx",
        "description": "Second 490-call July/August resident simulation. Same manual tab, same 490 call IDs. AI tab is Sheet14 (renamed 'AI JulyAugust Resident Simulation 2.0').",
        "changes": "Live protocol changes (column I) shipped between runs: I3 dropped the last-name requirement, I6 reverted the (17) tightening (specific reason + next-step action = Yes), I8 expanded next-steps YES list, I11 (validate) was written from scratch (was empty in file 17).",
    },
]

# Human tab column order is NOT the same as the AI tab.
HUMAN_COL_MAP = {
    3: "hold_permission",
    4: "validate_concern",
    5: "greeting",
    6: "name_usage",
    7: "contact_info",
    8: "unit_number",
    9: "reason_for_call",
    10: "acknowledged",
    11: "next_steps",
    12: "final_closing_q",  # human-only — not in AI tab
    13: "fha",
    14: "secure_info",
}

AI_COL_MAP = {
    2: "greeting",
    3: "name_usage",
    4: "contact_info",
    5: "unit_number",
    6: "reason_for_call",
    7: "acknowledged",
    8: "next_steps",
    9: "neutral_language",  # AI-only — not in human tab
    10: "hold_permission",
    11: "validate_concern",
    12: "fha",
    13: "secure_info",
}

# Overlapping scored questions (37 pts). Neutral (5) and final-closing are
# one-sided, so they are excluded from agreement and from the weighted score.
WEIGHTS = {
    "greeting": 4,
    "name_usage": 3,
    "contact_info": 4,
    "unit_number": 4,
    "reason_for_call": 7,
    "acknowledged": 5,
    "next_steps": 3,
    "hold_permission": 2,
    "validate_concern": 5,
}
SCORED_KEYS = list(WEIGHTS.keys())
DQ_KEYS = ["fha", "secure_info"]
ALL_KEYS = SCORED_KEYS + DQ_KEYS
TOTAL_POINTS = sum(WEIGHTS.values())  # 37

Q_LABELS = {
    "greeting": "Greeting",
    "name_usage": "Name usage",
    "contact_info": "Contact info",
    "unit_number": "Unit number",
    "reason_for_call": "Reason for the call",
    "acknowledged": "Acknowledged/ownership",
    "next_steps": "Next steps",
    "hold_permission": "Hold permission",
    "validate_concern": "Validate concern",
    "fha": "FHA violation (DQ)",
    "secure_info": "Secure info (DQ)",
}

Q_FULL_LABELS = {
    "greeting": "Greeting (property name + intro)",
    "name_usage": "Asked for name + used it",
    "contact_info": "Contact info confirmed (if update requested)",
    "unit_number": "Unit / apartment number confirmed",
    "reason_for_call": "Reason for the call",
    "acknowledged": "Acknowledged caller / took ownership",
    "next_steps": "Next steps / role / what to expect",
    "hold_permission": "Asked permission before hold",
    "validate_concern": "Validated caller concern (if expressed)",
}

STRICT_REASONS = {
    "greeting": "Agent greeted with property name and intro. AI did not detect it.",
    "name_usage": "Human credited asked-for + used name. AI marked No.",
    "contact_info": "Agent confirmed contact info. AI did not credit it.",
    "unit_number": "Unit/apartment number was confirmed. AI missed it.",
    "reason_for_call": "Human credited capturing the reason for the call. AI marked No.",
    "acknowledged": "Agent acknowledged caller / took ownership. AI did not credit it.",
    "next_steps": "Human credited next steps / role / expectation. AI marked No.",
    "hold_permission": "Agent asked hold permission (or hold not needed). AI did not credit it.",
    "validate_concern": "Human credited concern validation. AI did not.",
}

LENIENT_REASONS = {
    "greeting": "AI credited greeting; human says it was insufficient.",
    "name_usage": "AI credited name usage; human says name was not properly used.",
    "contact_info": "AI credited contact confirmation; human disagrees.",
    "unit_number": "AI credited unit confirmation; human disagrees.",
    "reason_for_call": "AI credited capturing the reason; human marked No.",
    "acknowledged": "AI credited acknowledgment; human required more explicit ownership.",
    "next_steps": "AI credited next steps; human did not.",
    "hold_permission": "AI credited hold permission; human says it was not properly asked.",
    "validate_concern": "AI credited concern validation; human did not.",
    "fha": "AI flagged an FHA violation; human did not find one.",
    "secure_info": "AI flagged secure info disclosure; human did not find one.",
}


def parse_yes_no(val):
    if val is None or str(val).strip() == "":
        return None
    return 1 if str(val).strip().lower().startswith("y") else 0


def call_id_str(val):
    if val is None:
        return None
    try:
        return str(int(float(val)))
    except (TypeError, ValueError):
        s = str(val).strip()
        return s or None


def calc_score(answers):
    earned = sum(WEIGHTS[q] for q in SCORED_KEYS if answers.get(q) == 1)
    base = (earned / TOTAL_POINTS) * 100 if TOTAL_POINTS else 0
    dq_count = sum(1 for dq in DQ_KEYS if answers.get(dq) == 1)
    return round(base * max(0, 1 - 0.20 * dq_count), 2)


def js_escape(s):
    return (
        s.replace("\\", "\\\\")
        .replace("'", "\\'")
        .replace('"', '\\"')
        .replace("`", "\\`")
        .replace("\n", " ")
    )


def load_tab(ws, id_col, col_map, skip_dups_label):
    data = {}
    dups = 0
    for row in range(2, ws.max_row + 1):
        cid = call_id_str(ws.cell(row, id_col).value)
        if not cid:
            continue
        if cid in data:
            dups += 1
            continue
        answers = {k: parse_yes_no(ws.cell(row, c).value) for c, k in col_map.items()}
        data[cid] = answers
    print(f"  {skip_dups_label}: {len(data)} unique, {dups} duplicate rows skipped")
    return data


print("Reading spreadsheet...")
wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
print(f"  Tabs: {wb.sheetnames}")
human_raw = load_tab(wb[MANUAL_TAB], 1, HUMAN_COL_MAP, "Human")


def build_run(cfg, human_raw, wb):
    print("\n" + "=" * 70)
    print(f"Building {cfg['label']}  (id {cfg['run_id']})  from tab '{cfg['ai_tab']}'")
    print("=" * 70)
    ai_raw = load_tab(wb[cfg["ai_tab"]], 1, AI_COL_MAP, "AI")
    both = sorted(set(human_raw) & set(ai_raw))
    print(
        f"  Matching IDs: {len(both)}  "
        f"human-only={len(set(human_raw)-set(ai_raw))}  "
        f"ai-only={len(set(ai_raw)-set(human_raw))}"
    )

    answer_data = {}
    for cid in both:
        h = human_raw[cid]
        a = ai_raw[cid]
        answer_data[cid] = {q: [h.get(q), a.get(q)] for q in ALL_KEYS}

    total_agree = total_dis = total_s = total_l = total_comp = 0
    dq_dis = 0
    score_deltas = []
    calls_data = []
    q_acc = {
        q: {"agree": 0, "dis": 0, "strict": 0, "lenient": 0, "total": 0,
            "strict_ids": [], "lenient_ids": [], "h_yes": 0, "ai_yes": 0}
        for q in ALL_KEYS
    }

    for cid in both:
        cd = answer_data[cid]
        h_score = calc_score({q: cd[q][0] for q in ALL_KEYS})
        a_score = calc_score({q: cd[q][1] for q in ALL_KEYS})
        score_deltas.append(abs(h_score - a_score))
        strict = lenient = 0
        details = []
        for q in ALL_KEYS:
            h_val, a_val = cd[q]
            if h_val is None or a_val is None:
                continue
            is_dq = q in DQ_KEYS
            q_acc[q]["total"] += 1
            if h_val == 1:
                q_acc[q]["h_yes"] += 1
            if a_val == 1:
                q_acc[q]["ai_yes"] += 1
            if not is_dq:
                total_comp += 1
            if h_val == a_val:
                q_acc[q]["agree"] += 1
                if not is_dq:
                    total_agree += 1
                continue
            q_acc[q]["dis"] += 1
            if is_dq:
                dq_dis += 1
            else:
                total_dis += 1
            w = WEIGHTS.get(q, "DQ")
            if h_val == 1 and a_val == 0:
                strict += 1
                q_acc[q]["strict"] += 1
                q_acc[q]["strict_ids"].append(cid)
                if not is_dq:
                    total_s += 1
                details.append(
                    {
                        "q": Q_LABELS[q],
                        "w": "DQ" if is_dq else w,
                        "ai": "No",
                        "h": "Yes",
                        "reason": STRICT_REASONS.get(q, "AI missed this behavior."),
                        "key": q,
                    }
                )
            else:
                lenient += 1
                q_acc[q]["lenient"] += 1
                q_acc[q]["lenient_ids"].append(cid)
                if not is_dq:
                    total_l += 1
                details.append(
                    {
                        "q": Q_LABELS[q],
                        "w": "DQ" if is_dq else w,
                        "ai": "Yes",
                        "h": "No",
                        "reason": LENIENT_REASONS.get(q, "AI over-credited this behavior."),
                        "key": q,
                    }
                )

        weight_lost = sum(d["w"] for d in details if d["w"] != "DQ")
        has_dq = any(d["w"] == "DQ" for d in details)
        n_dis = strict + lenient
        if n_dis == 0:
            note = "Perfect agreement on every overlapping question."
            rec = "No action needed — perfect agreement."
        else:
            bits = []
            if strict:
                bits.append(f"{strict} strict")
            if lenient:
                bits.append(f"{lenient} lenient")
            note = f"{n_dis} disagreement{'s' if n_dis != 1 else ''}: {', '.join(bits)}."
            rec_parts = []
            s_qs = sorted(
                [d for d in details if d["ai"] == "No" and d["w"] != "DQ"],
                key=lambda x: -x["w"],
            )[:3]
            l_qs = sorted(
                [d for d in details if d["ai"] == "Yes" and d["w"] != "DQ"],
                key=lambda x: -x["w"],
            )[:3]
            if s_qs:
                rec_parts.append("Strict: " + ", ".join(f"{d['q']} ({d['w']}pts)" for d in s_qs))
            if l_qs:
                rec_parts.append("Lenient: " + ", ".join(f"{d['q']} ({d['w']}pts)" for d in l_qs))
            rec = ". ".join(rec_parts) + "."

        calls_data.append(
            {
                "id": cid,
                "human": h_score,
                "ai": a_score,
                "strict": strict,
                "lenient": lenient,
                "weightLost": weight_lost,
                "disqualifier": has_dq,
                "note": note,
                "details": details,
                "recommendation": rec,
            }
        )

    questions_data = []
    for q in ALL_KEYS:
        acc = q_acc[q]
        if acc["dis"] == 0:
            direction = "mixed"
        elif acc["strict"] > acc["lenient"]:
            direction = "strict"
        elif acc["lenient"] > acc["strict"]:
            direction = "lenient"
        else:
            direction = "mixed"
        questions_data.append(
            {
                "short": q,
                "label": Q_FULL_LABELS.get(q, Q_LABELS[q]),
                "weight": WEIGHTS.get(q, 0),
                "agree": acc["agree"],
                "disagree": acc["dis"],
                "total": acc["total"],
                "strict": acc["strict"],
                "lenient": acc["lenient"],
                "dir": direction,
                "strict_ids": acc["strict_ids"],
                "lenient_ids": acc["lenient_ids"],
                "h_yes": acc["h_yes"],
                "ai_yes": acc["ai_yes"],
            }
        )
    questions_data.sort(key=lambda q: (-q["disagree"], -q["weight"]))

    agreement_pct = round((total_agree / total_comp) * 100, 1) if total_comp else 0
    avg_delta = round(sum(score_deltas) / len(score_deltas), 1) if score_deltas else 0
    perfect = sum(1 for c in calls_data if c["strict"] == 0 and c["lenient"] == 0)
    mean_h = round(sum(c["human"] for c in calls_data) / len(calls_data), 1)
    mean_a = round(sum(c["ai"] for c in calls_data) / len(calls_data), 1)
    need_90 = max(0, int(0.9 * total_comp - total_agree + 0.999))

    return {
        "cfg": cfg,
        "both": both,
        "answer_data": answer_data,
        "questions_data": questions_data,
        "calls_data": calls_data,
        "metrics": {
            "agreement_pct": agreement_pct,
            "avg_delta": avg_delta,
            "total_agree": total_agree,
            "total_dis": total_dis,
            "total_s": total_s,
            "total_l": total_l,
            "total_comp": total_comp,
            "dq_dis": dq_dis,
            "perfect": perfect,
            "mean_h": mean_h,
            "mean_a": mean_a,
            "need_90": need_90,
        },
    }


runs_built = [build_run(cfg, human_raw, wb) for cfg in LARGE_TESTS]


# ── Report metrics for each ─────────────────────────────────────────────────
for run in runs_built:
    m = run["metrics"]
    cfg = run["cfg"]
    print("\n" + "-" * 60)
    print(f"{cfg['label']} metrics")
    print("-" * 60)
    print(f"  Calls: {len(run['both'])}")
    print(f"  Agreement: {m['agreement_pct']}% ({m['total_agree']}/{m['total_comp']})")
    print(f"  Disagreements: {m['total_dis']}  strict {m['total_s']}  "
          f"lenient {m['total_l']}  DQ {m['dq_dis']}")
    print(f"  Avg |delta|: {m['avg_delta']}%   mean H {m['mean_h']}  mean AI {m['mean_a']}")
    print(f"  Perfect: {m['perfect']}/{len(run['both'])}")
    print(f"  90% needs {m['need_90']} more scored agreements")
    for q in run["questions_data"]:
        print(
            f"  {q['short']:20} {q['agree']}/{q['total']} = "
            f"{round(q['agree']/q['total']*100,1)}%  S{q['strict']} L{q['lenient']}"
        )


# ── Run 1 (Large Test) key findings / root cause / recs ────────────────────
def build_run1_content(run):
    m = run["metrics"]
    cfg = run["cfg"]
    qmap = {q["short"]: q for q in run["questions_data"]}
    name_q, next_q, reason_q, val_q = qmap["name_usage"], qmap["next_steps"], qmap["reason_for_call"], qmap["validate_concern"]
    greet_q, unit_q, ack_q, contact_q, hold_q = qmap["greeting"], qmap["unit_number"], qmap["acknowledged"], qmap["contact_info"], qmap["hold_permission"]

    name_pct = round(name_q["agree"] / name_q["total"] * 100, 1)
    next_pct = round(next_q["agree"] / next_q["total"] * 100, 1)
    reason_pct = round(reason_q["agree"] / reason_q["total"] * 100, 1)
    val_pct = round(val_q["agree"] / val_q["total"] * 100, 1)
    strict_share = round(m["total_s"] / m["total_dis"] * 100) if m["total_dis"] else 0
    name_w = name_q["weight"] * name_q["strict"]
    reason_w = reason_q["weight"] * reason_q["strict"]
    next_w = next_q["weight"] * next_q["strict"]
    val_w = val_q["weight"] * val_q["lenient"]
    top3_s = name_q["strict"] + next_q["strict"] + reason_q["strict"]

    key_findings = f"""<p><strong>Large Test holds at {m['agreement_pct']}% on 490 calls — the 20-call Run 8.0 result was not a small-n illusion.</strong> Spreadsheet tabs <strong>Manual Grades JulyAugust Reside</strong> vs <strong>AI JulyAugust Resident Simuluat</strong>. Scored agreement <strong>{m['agreement_pct']}%</strong> ({m['total_agree']}/{m['total_comp']}) across 9 overlapping questions. Disagreements: <strong>{m['total_dis']}</strong> ({m['total_s']} strict / {m['total_l']} lenient). Avg absolute score delta <strong>{m['avg_delta']} pp</strong> on a 37-pt overlapping card. Perfect agreement: <strong>{m['perfect']} of {len(run['both'])}</strong> ({round(m['perfect']/len(run['both'])*100,1)}%). Mean AI {m['mean_a']}% vs mean human {m['mean_h']}%.</p>
<p><strong>The model is too strict at scale, not too loose.</strong> {strict_share}% of disagreements are AI=No / human=Yes ({m['total_s']} of {m['total_dis']}). Human is the reference, not the truth — but a {strict_share}/{100-strict_share} split that large is the model under-crediting agents. Run 8.0 on 20 calls was roughly even (10 strict / 11 lenient). The 490-call set exposes the real bias.</p>
<p><strong>Three questions are the work.</strong> Name usage <strong>{name_pct}%</strong> ({name_q["agree"]}/{name_q["total"]}, {name_q["strict"]}S/{name_q["lenient"]}L) — worst agreement, AI Yes-rate 69% vs human 91%. Next steps <strong>{next_pct}%</strong> ({next_q["agree"]}/{next_q["total"]}, {next_q["strict"]}S/{next_q["lenient"]}L) — AI 81% vs human 97%. Reason for the call <strong>{reason_pct}%</strong> ({reason_q["agree"]}/{reason_q["total"]}, {reason_q["strict"]}S/{reason_q["lenient"]}L) — AI 85% vs human 95%, and the highest weighted miss because it is 7 pts ({reason_w} strict-pts). Those three are {top3_s} of {m['total_s']} strict errors ({round(top3_s/m['total_s']*100) if m['total_s'] else 0}%).</p>
<p><strong>Live protocol is column I (Other Prompt)</strong> on <em>AI Resident Fundamentals</em> in <code>AI_ QA 2026 (17).xlsx</code>. Column H is empty. Recs below are vs that text, not vs the question stem. File (17) tightened reason-for-call vs file (16): caller stating the reason is no longer enough, and a generic callback is No. That tightening is the opposite of the 95% human Yes-rate.</p>
<p><strong>Validate concern is the only big lenient miss</strong> — {val_pct}% ({val_q["agree"]}/{val_q["total"]}, {val_q["strict"]}S/{val_q["lenient"]}L). AI Yes-rate 99% vs human 86%. Column I for validate is <strong>empty</strong> — there is no protocol to edit, only the question stem. Contact {round(contact_q["agree"]/contact_q["total"]*100,1)}%, hold {round(hold_q["agree"]/hold_q["total"]*100,1)}%, FHA 100%, secure-info 99.4% — leave them alone.</p>
<p><strong>Scorecard caveat:</strong> humans also graded a 10th item (final closing question, 92% Yes) that AI never answered. AI also graded neutral language (99.8% Yes) that humans never answered. Agreement below uses the 9 shared scored questions + 2 DQs (37 pts, not the 42-pt 20-call card). Graded-stats Yes-rates on the sheet match these Yes-rates. <strong>→ Open Recommendations</strong> — 90% is {m['need_90']} more scored agreements. Recovering half of name strict + half of next-steps strict clears it.</p>"""

    root_strict = f"""<p><strong>Name usage ({name_q["strict"]} strict / {name_q["lenient"]} lenient, {name_w} weighted pts):</strong> Lowest agreement on the card ({name_pct}%). Column I requires <em>full name (first and last)</em> asked or volunteered, then one first-name use. Humans graded the weaker stem (name asked/offered + used). Resident callers often give a first name only. That I vs stem mismatch is the leading hypothesis — not sales-style repeated use. Keep the refusal and anonymous Yeses already in I.</p>
<p><strong>Next steps ({next_q["strict"]} strict / {next_q["lenient"]} lenient, {next_w} weighted pts):</strong> Column I already credits callback, notes to property, voicemail, “someone will reach out,” cannot-transfer + what they can do, cooperative ending, and declined callback. “Anything else?” is sufficient but not required. Truncation is not a No. The {next_q["strict"]} false Nos are the model <em>not following</em> I8, not I8 being too tight. Do not rewrite I8. Do not add a second next-steps protocol.</p>
<p><strong>Reason for the call ({reason_q["strict"]} strict / {reason_q["lenient"]} lenient, {reason_w} weighted pts):</strong> Highest weighted miss. File (17) column I is a tightening vs file (16): “A caller stating the reason is not sufficient by itself,” and the package + “someone will call you back” example is explicitly No. Humans Yes-rate 95% — they are still grading like file (16). If the goal is 90% vs humans, I6 (17) is fighting the reference graders. Hold the {reason_q["lenient"]} false Yeses — do not copy AI 7’s over-credit, and do not revert to open-ended.</p>
<p><strong>Second-tier strict cluster:</strong> Greeting {greet_q["strict"]}S/0L, unit {unit_q["strict"]}S/0L, ownership {ack_q["strict"]}S/{ack_q["lenient"]}L. Column I for all three is already Yes-friendly (standard greeting counts; asking for unit is enough; first-person callback/note language counts). Remaining misses are follow-through, not missing protocol. Do not rewrite I2 / I5 / I7 in the same paste.</p>"""

    root_lenient = f"""<p><strong>Validate concern ({val_q["lenient"]} lenient / {val_q["strict"]} strict, {val_w} weighted over-credit pts):</strong> The only question where AI is systematically generous. Column I is empty — the model only has the question stem (“If no concerns were expressed always answer yes”). Humans still marked No on {val_q["lenient"]} of these calls (human Yes-rate 85.5%). Write I11 rather than editing a protocol that does not exist. Do not auto-Yes through frustration / urgency / a stated problem. Specific validating phrase required when concern language is present.</p>
<p><strong>Everything else lenient is noise.</strong> Contact {contact_q["lenient"]}L, reason {reason_q["lenient"]}L, next steps {next_q["lenient"]}L, name {name_q["lenient"]}L, secure-info 3, hold 2. Do not write protocols for these. Tightening them will cost Yeses we need on the strict side.</p>
<p>Human is the reference. On validate, humans may be the ones being harsh. On name / next steps / reason, the Yes-rate gaps (22 / 16 / 10 pp) are too large to pin on grader noise — and on reason, file (17) I6 is also fighting the humans.</p>"""

    recs = [
        {
            "num": 1,
            "title": f"Name usage — {name_pct}% is the largest gap ({name_q['strict']} strict)",
            "severity": "critical",
            "severityLabel": f"{name_q['agree']}/{name_q['total']} · {name_q['strict']} strict / {name_q['lenient']} lenient · AI Yes 69% vs human 91%",
            "owner": "AI Engineering",
            "ownerClass": "info",
            "problem": f"""<p>Name is the worst question on 490 calls. Agreement {name_pct}%. AI misses <strong>26%</strong> of human-Yes names ({name_q['strict']}/{name_q['agree'] + name_q['strict']}). False Yeses are only {name_q['lenient']} — this is not a loosen-vs-tighten toss-up. Recovering all {name_q['strict']} stricts would land at {round((m['total_agree'] + name_q['strict']) / m['total_comp'] * 100, 1)}%. Recovering half ({name_q['strict'] // 2}) gets to {round((m['total_agree'] + name_q['strict'] // 2) / m['total_comp'] * 100, 1)}%, one point shy of 90% by itself.</p>
<p>Column I (Other Prompt), row 3, is stricter than the question humans graded. Humans see “asked for the caller’s name, or it was offered, and used it at least once.” I3 requires <strong>full name (first and last)</strong> asked or volunteered before the first-name-use counts. Refusal and anonymous are already Yes in I3 — keep those. Resident callers often give a first name only. That I vs stem mismatch is the leading hypothesis for the {name_q['strict']} stricts — not “sales-style repeated use.”</p>""",
            "protocols": [
                {
                    "label": "Name usage — drop the last-name requirement; keep ask-or-offer + one first-name use",
                    "current": (
                        "The transcript must show that the agent either asked for the caller’s full name (first and last name) "
                        "or the caller voluntarily provided their full name without being asked, and the agent must use the caller’s "
                        "first name at least once during the conversation. Both actions must occur for a “Yes” answer. If the agent "
                        "requests the caller’s full name and the caller refuses to provide it then mark \"yes\"<br><br>"
                        "If the caller wants to remain anonymous and does not give name, phone, email, and/or apartment number, mark Yes."
                    ),
                    "recommended": (
                        "Keep ask-or-offer + one first-name use. <strong style=\"color:var(--red);\">Do not require last name.</strong><br><br>"
                        "The transcript must show that the agent either asked for the caller’s name (first name is enough) or the caller "
                        "volunteered it, and the agent used the caller’s first name at least once. Both must occur for Yes.<br><br>"
                        "<strong style=\"color:var(--red);\">Keep these Yeses:</strong> agent asks for the name and the caller refuses; "
                        "caller stays anonymous and gives no name / phone / email / unit.<br><br>"
                        "<strong style=\"color:var(--red);\">Also Yes:</strong> first-name-only volunteer (“This is Mia”) + one use; "
                        "“Can I get your name / first name?” + one use; transcript-garbled name if the agent is clearly addressing the caller by name.<br><br>"
                        f"Do not require a second reuse. Gate: name ≥ 85% on this 490, strict &lt; 70. Leave the {name_q['lenient']} false Yeses unless they move with the loosen."
                    ),
                }
            ],
        },
        {
            "num": 2,
            "title": f"Next steps — {next_pct}% , {next_q['strict']} false Nos (I8 already credits callback)",
            "severity": "critical",
            "severityLabel": f"{next_q['agree']}/{next_q['total']} · {next_q['strict']} strict / {next_q['lenient']} lenient · AI Yes 81% vs human 97%",
            "owner": "AI Engineering",
            "ownerClass": "info",
            "problem": f"""<p>Human Yes-rate on next steps is 96.9%. AI is 81.0%. Almost every miss is a false No ({next_q['strict']} vs {next_q['lenient']} false Yes). Half of these plus half of name strict is {name_q['strict'] // 2 + next_q['strict'] // 2} agreements → {round((m['total_agree'] + name_q['strict'] // 2 + next_q['strict'] // 2) / m['total_comp'] * 100, 1)}%, which clears 90% without touching reason.</p>
<p>Column I row 8 already says the thing the previous rec asked to add: callback, notes to property, voicemail, someone will reach out, cannot-transfer + what they can do, cooperative ending, declined callback still Yes, “anything else?” not required, truncation not a No. The {next_q['strict']} false Nos are the model <strong>not following I8</strong>, not a missing protocol. Do not rewrite I8. Adding competing next-steps text is how Runs 3–5 lost ground.</p>""",
            "protocols": [
                {
                    "label": "Next steps — keep I8 verbatim; enforce it, do not rewrite it",
                    "current": (
                        "Mark YES if the agent clearly states the next step (callback arranged, notes sent to property, "
                        "voicemail/message being placed, someone will reach out, or they cannot transfer and this is what they "
                        "can do instead) and the call ends cooperatively (caller okay/thanks, caller declines the offered next step, "
                        "or agent asks \"anything else?\").<br><br>"
                        "\"Is there anything else?\" is sufficient but not required when next steps are already confirmed and the "
                        "caller winds down the call.<br><br>"
                        "If the agent offers a callback or transfer and the caller declines, still mark YES — the next step was explained.<br><br>"
                        "If the agent commits to forwarding or placing a voicemail, that is the next step. Do not mark NO just because "
                        "the recording ends while the caller is still talking.<br><br>"
                        "Mark NO only when the agent ends abruptly with no next step and no confirmation that the caller's request was handled."
                    ),
                    "recommended": (
                        "<strong style=\"color:var(--red);\">Do not change this protocol.</strong> Paste I8 as-is.<br><br>"
                        "Enforcement only: if any Yes condition above is present, mark YES. The 88 false Nos on this 490 are misses "
                        "against this text (callback / note / voicemail / role with no \"anything else,\" truncated voicemail-forward). "
                        "Put 2–3 of those call IDs in column J (Additional Prompts), not a second I8.<br><br>"
                        "Do not add negative examples first. Gate: next steps ≥ 90% on this 490, strict &lt; 40."
                    ),
                }
            ],
        },
        {
            "num": 3,
            "title": f"Reason for the call — {reason_pct}%, I6 (17) is fighting the humans ({reason_w} pts)",
            "severity": "critical",
            "severityLabel": f"{reason_q['agree']}/{reason_q['total']} · {reason_q['strict']} strict / {reason_q['lenient']} lenient · 7-pt slot",
            "owner": "AI Engineering",
            "ownerClass": "info",
            "problem": f"""<p>Unlike Run 8.0, both sides graded “Did the agent capture the reason for the call?” Agreement is {reason_pct}% with {reason_q['strict']} false Nos. At 7 pts this is the most expensive strict question ({reason_w} weighted pts vs name {name_w} vs next steps {next_w}).</p>
<p>File (17) column I is a <strong>tightening vs file (16)</strong>. (16) said caller stating the reason is enough, and acting on it counts. (17) says the opposite: “A caller stating the reason is not sufficient by itself,” and the package + “someone will call you back” example is explicitly No. Humans Yes-rate 95% — they are still grading like (16). If the goal is 90% vs humans, keep (17) I6 and you will keep most of these {reason_q['strict']} stricts.</p>
<p>Keep the question. Do not revert to 2+ open-ended. Do not ship the AI 7 over-credit pattern ({reason_q['lenient']} false Yeses already exist; do not add more).</p>""",
            "protocols": [
                {
                    "label": "Reason for the call — revert I6 (17) toward the human / file (16) bar",
                    "current": (
                        "A caller stating the reason is not sufficient by itself.<br><br>"
                        "Mark YES only when:<br>"
                        "1. The caller’s specific reason is clear; AND<br>"
                        "2. The agent explicitly captures that reason by repeating, summarizing, confirming, asking a related "
                        "follow-up, or taking an action that clearly references the specific reason.<br><br>"
                        "A generic statement such as “I will forward the message,” “someone will call you back,” or “I will send this "
                        "to the property” does not prove that the reason was captured unless the agent also identifies what the "
                        "message or callback is about.<br><br>"
                        "Examples:<br>"
                        "- Caller: “My package has not arrived.” Agent: “I’ll send a package-delivery request to the property.” → YES<br>"
                        "- Caller: “My package has not arrived.” Agent: “Okay, may I have your name and phone number? Someone will call you back.” → NO<br>"
                        "- Caller: “I need a walk-through and carpet-cleaning information.” Agent: “I’ll ask the leasing specialist to arrange the walk-through and provide the carpet-cleaning information.” → YES<br>"
                        "- Caller: “I need help.” Agent: “Someone will call you back.” → NO<br><br>"
                        "Identity and contact collection alone never proves that the reason was captured."
                    ),
                    "recommended": (
                        "<strong style=\"color:var(--red);\">Revert the file (17) tightening.</strong> Keep the question. Do not go back to 2+ open-ended.<br><br>"
                        "The reason for the call is why the person is calling: the request, concern, issue, or who they need to reach.<br><br>"
                        "Mark YES if, by the end of the call, the specific reason is known from the transcript. Any one of these is enough:<br>"
                        "- The caller states a specific reason (package, lockout, work order, inspection, notice, payment, speak to the office).<br>"
                        "- The agent asks what the call is regarding and gets an answer.<br>"
                        "- The agent restates, confirms, OR takes a next-step action on that issue (note, callback, voicemail, send to property) "
                        "<strong style=\"color:var(--red);\">even without repeating the reason</strong>.<br><br>"
                        "“How may I help you?” plus the caller explaining is YES. A second probe is not required.<br><br>"
                        "<strong style=\"color:var(--red);\">Flip this example:</strong> Caller: “My package has not arrived.” "
                        "Agent: “Okay, may I have your name and phone number? Someone will call you back.” → YES "
                        "(specific reason was stated; the callback is capture).<br><br>"
                        "Mark NO only if:<br>"
                        "- The call ends and it is still unclear why they called (agent only collected name / unit / phone, caller never stated a request).<br>"
                        "- The caller only said “I need help” / “I have a problem” and the agent never learned the topic. "
                        "(“I need help.” / “Someone will call you back.” stays NO.)<br><br>"
                        "Identity and contact collection alone never proves capture when no reason was stated.<br><br>"
                        "Do not mark NO because the agent did not recap, skipped probing after the caller already explained, or used yes/no confirms after the reason was known.<br><br>"
                        f"Gate: reason ≥ 90% on this 490, strict &lt; 30, lenient still ≤ {reason_q['lenient']}."
                    ),
                }
            ],
        },
        {
            "num": 4,
            "title": f"Validate concern — column I is empty ({val_q['lenient']} false Yeses)",
            "severity": "warning",
            "severityLabel": f"{val_q['agree']}/{val_q['total']} · {val_q['strict']} strict / {val_q['lenient']} lenient · AI Yes 99% vs human 86%",
            "owner": "AI Engineering",
            "ownerClass": "info",
            "problem": f"""<p>Validate is the only large lenient miss. AI Yes-rate 98.6% vs human 85.5%. {val_q['lenient']} of {m['total_l']} lenient errors ({round(val_q['lenient']/m['total_l']*100) if m['total_l'] else 0}%) live here. Column I row 11 is <strong>blank</strong> — the model only has the question stem (“If no concerns were expressed always answer yes”). There is no protocol to edit. Write I11.</p>
<p>Do not paste this in the same cycle as recs 1–3 if you need a clean read on those. Second paste. Target: cut lenient ~in half without creating a new strict pile.</p>""",
            "protocols": [
                {
                    "label": "Validate concern — write I11 (it does not exist today)",
                    "current": (
                        "Column I is empty. The model only has the question stem:<br><br>"
                        "If the caller expressed a concern, did the agent use at least one specific phrase to acknowledge and "
                        "validate the concern? If no concerns were expressed always answer yes."
                    ),
                    "recommended": (
                        "Write this into column I.<br><br>"
                        "Mark YES if no concern / frustration / urgency language is in the transcript (purely transactional: name, unit, callback).<br><br>"
                        "<strong style=\"color:var(--red);\">Mark NO</strong> when the caller states a problem, delay, missed callback, "
                        "unsafe condition, billing issue, or clear frustration and the agent does not acknowledge it with a validating "
                        "phrase (“I understand,” “I am sorry you are dealing with that,” “that makes sense”). Proceeding to a callback "
                        "without acknowledgment is not validation.<br><br>"
                        "Do not auto-Yes just because the agent took a message. Gate: validate agreement ≥ 90%, lenient &lt; 35, strict still ≤ 10."
                    ),
                }
            ],
        },
        {
            "num": 5,
            "title": "I2 / I4 / I5 / I7 / I10 are already Yes-friendly — do not rewrite them",
            "severity": "success",
            "severityLabel": f"Hold {round(hold_q['agree']/hold_q['total']*100,1)}% · contact {round(contact_q['agree']/contact_q['total']*100,1)}% · FHA 100% · greeting/unit/ack are follow-through misses",
            "owner": "Austin + AI Engineering",
            "ownerClass": "info",
            "problem": f"""<p>Hold I10 (lookup ≠ hold), contact I4 (attempt is enough, N/A and anonymous are Yes), and FHA I13 are done. Greeting I2, unit I5, and ownership I7 are already loose — remaining {greet_q['strict']} / {unit_q['strict']} / {ack_q['strict']} stricts are the model not following those I-cols. Do not paste new greeting / unit / ownership text in the same cycle as recs 1–3.</p>
<p><strong>90% math:</strong> {m['total_agree']}/{m['total_comp']} today. Need {m['need_90']} more scored agreements (3969/{m['total_comp']}). Name+next-steps half-recovery is {name_q['strict']//2 + next_q['strict']//2} → {round((m['total_agree'] + name_q['strict']//2 + next_q['strict']//2)/m['total_comp']*100,1)}%. All name strict recovered is {round((m['total_agree'] + name_q['strict'])/m['total_comp']*100,1)}% by itself.</p>
<p>Re-run on this same 490 after recs 1 and 3 (I3 last-name drop + I6 revert) and I8 enforcement. Same human tab. New AI tab. Success = ≥90% scored agreement, name ≥85%, next steps ≥90%, reason ≥90%, validate not worse than {val_pct}%.</p>""",
            "protocols": [
                {
                    "label": "Unit (I5) — already asking-is-enough; do not rewrite",
                    "current": (
                        "Mark YES if the agent asked for the unit or apartment number, the caller offered it, or it was confirmed. "
                        "Asking is enough. Street address counts. Refusal or anonymous → YES. Mark NO only if never asked, never offered, never confirmed."
                    ),
                    "recommended": (
                        "<strong style=\"color:var(--red);\">Do not change I5.</strong> 40 stricts are misses against a protocol that already "
                        "says asking is enough. Same for I7 (first-person callback/note language already counts) and I2 (property+name "
                        "OR the standard greeting). Leave I4 / I10 / I13 / I14 alone."
                    ),
                },
                {
                    "label": "Large Test execution checklist — vs column I, not vs the stems",
                    "current": f"Source: AI_ QA 2026 (17).xlsx · AI Resident Fundamentals · column I. Large Test: {m['agreement_pct']}% on {len(run['both'])} calls, {m['total_dis']} disagrees ({m['total_s']}S/{m['total_l']}L). Name {name_pct}%. Next steps {next_pct}%. Reason {reason_pct}%. Validate {val_pct}%.",
                    "recommended": (
                        "One paste, two protocol edits + one enforcement: (1) I3 drop last-name, (2) I6 revert the (17) tightening so "
                        "specific stated reason + callback = Yes, (3) I8 unchanged — enforce via column J examples. Do not revert "
                        "reason to open-ended. Do not import AI 7. Do not write I11 (validate) until the next cycle. Do not rewrite "
                        "I2 / I4 / I5 / I7 / I10. Same 490 IDs. Success = 90%+, name ≥ 85%, next steps ≥ 90%, reason ≥ 90%."
                    ),
                },
            ],
        },
    ]

    root_cause = {
        "overall": f"Large Test is {len(run['both'])} July/August resident calls at {m['agreement_pct']}% scored agreement — {m['need_90']} agreements short of 90%. Live protocol is column I (Other Prompt) on AI Resident Fundamentals in AI_ QA 2026 (17).xlsx. Dominant bias is AI too strict ({strict_share}% of disagrees). Name I3 last-name bar, I8 not being followed, and I6 (17) fighting the humans are the work. Validate I11 does not exist.",
        "what_worked": f"Hold {round(hold_q['agree']/hold_q['total']*100,1)}%, FHA 100%, secure-info 99.4%, contact {round(contact_q['agree']/contact_q['total']*100,1)}%. {m['perfect']} of {len(run['both'])} calls are perfect. Mean scores are both high (human {m['mean_h']}, AI {m['mean_a']}) — this is not a broken grader, it is three questions.",
        "what_didnt": f"Name {name_pct}% ({name_q['strict']}S), next steps {next_pct}% ({next_q['strict']}S), reason {reason_pct}% ({reason_q['strict']}S, {reason_w} weighted pts). Validate {val_pct}% the other way ({val_q['lenient']}L). Greeting {greet_q['strict']}S, unit {unit_q['strict']}S, ownership {ack_q['strict']}S as a second-tier cluster.",
        "path_to_90": f"Need {m['need_90']} more scored agreements ({int(0.9*m['total_comp'])}/{m['total_comp']}). Half of name strict + half of next-steps strict = {name_q['strict']//2 + next_q['strict']//2} → {round((m['total_agree'] + name_q['strict']//2 + next_q['strict']//2)/m['total_comp']*100,1)}%. Edits: I3 drop last-name, I6 revert the (17) tightening, I8 unchanged (enforce). Re-run this same 490.",
        "strictDetail": root_strict,
        "lenientDetail": root_lenient,
    }

    return key_findings, root_cause, recs


# ── Run 2 (Large Test 2) content — the shipped protocol run ─────────────────
def build_run2_content(run, run1):
    m = run["metrics"]
    r1 = run1["metrics"]
    cfg = run["cfg"]
    qmap = {q["short"]: q for q in run["questions_data"]}
    q1map = {q["short"]: q for q in run1["questions_data"]}

    name_q, next_q, reason_q, val_q = qmap["name_usage"], qmap["next_steps"], qmap["reason_for_call"], qmap["validate_concern"]
    greet_q, unit_q, ack_q, contact_q, hold_q = qmap["greeting"], qmap["unit_number"], qmap["acknowledged"], qmap["contact_info"], qmap["hold_permission"]
    name_q1, next_q1, reason_q1, val_q1 = q1map["name_usage"], q1map["next_steps"], q1map["reason_for_call"], q1map["validate_concern"]

    def pct(q):
        return round(q["agree"] / q["total"] * 100, 1) if q["total"] else 0

    name_pct, next_pct, reason_pct, val_pct = pct(name_q), pct(next_q), pct(reason_q), pct(val_q)
    name_pct1, next_pct1, reason_pct1, val_pct1 = pct(name_q1), pct(next_q1), pct(reason_q1), pct(val_q1)
    contact_pct, ack_pct, unit_pct, greet_pct, hold_pct = pct(contact_q), pct(ack_q), pct(unit_q), pct(greet_q), pct(hold_q)

    strict_share = round(m["total_s"] / m["total_dis"] * 100) if m["total_dis"] else 0
    agree_delta = round(m["agreement_pct"] - r1["agreement_pct"], 1)
    dis_delta = m["total_dis"] - r1["total_dis"]
    strict_delta = m["total_s"] - r1["total_s"]
    lenient_delta = m["total_l"] - r1["total_l"]
    delta_delta = round(m["avg_delta"] - r1["avg_delta"], 1)
    perfect_delta = m["perfect"] - r1["perfect"]

    def q_delta(q_new, q_old):
        p_new = pct(q_new)
        p_old = pct(q_old)
        d = round(p_new - p_old, 1)
        return f"{p_new}% ({'+' if d >= 0 else ''}{d}pp, {q_new['strict']}S/{q_new['lenient']}L vs {q_old['strict']}S/{q_old['lenient']}L)"

    key_findings = f"""<p><strong>Large Test 2 clears 90% — {m['agreement_pct']}% scored agreement on the same 490 July/August resident calls (+{agree_delta}pp vs Large Test).</strong> Spreadsheet tabs <strong>Manual Grades JulyAugust Reside</strong> vs <strong>AI JulyAugust Resident Simulation 2.0</strong> (Sheet14). Disagreements dropped {r1['total_dis']} → <strong>{m['total_dis']}</strong> ({dis_delta:+d}). Strict {r1['total_s']} → <strong>{m['total_s']}</strong> ({strict_delta:+d}). Lenient {r1['total_l']} → <strong>{m['total_l']}</strong> ({lenient_delta:+d}). Avg score delta {r1['avg_delta']}% → <strong>{m['avg_delta']}%</strong> ({delta_delta:+.1f}pp). Perfect agreement {r1['perfect']} → <strong>{m['perfect']}</strong> of {len(run['both'])} ({perfect_delta:+d}). Mean AI {r1['mean_a']} → <strong>{m['mean_a']}%</strong>, mean human unchanged at {m['mean_h']}%.</p>
<p><strong>Three protocol edits landed all three targets.</strong> Reason for the call {reason_pct1}% → <strong>{reason_pct}%</strong> (I6 reverted the file (17) tightening — 7-pt slot moved from 63 strict to {reason_q['strict']}). Next steps {next_pct1}% → <strong>{next_pct}%</strong> (I8 expanded YES list — 88 strict to {next_q['strict']}). Name usage {name_pct1}% → <strong>{name_pct}%</strong> (I3 dropped the last-name requirement — 118 strict to {name_q['strict']}). Nothing on the 20-call series predicted this size of gain because the 20-call series never had reason-for-call at scale.</p>
<p><strong>Validate concern regressed because I11 was written from scratch.</strong> {val_pct1}% ({val_q1['strict']}S/{val_q1['lenient']}L) → <strong>{val_pct}%</strong> ({val_q['strict']}S/{val_q['lenient']}L). New I11 requires an explicit validating phrase — it cut {val_q1['lenient'] - val_q['lenient']} lenient but created {val_q['strict'] - val_q1['strict']} new strict Nos, so net agreement dropped {round(val_pct1 - val_pct, 1)}pp on this question. The lever exists now; next paste is a knob-turn.</p>
<p><strong>Live protocol is column I (Other Prompt) in <code>AI_ QA 2026 (20).xlsx</code>.</strong> Changed rows this cycle: I3 (name — last name dropped), I6 (reason — reverted to file 16 bar with 4 examples), I8 (next steps — expanded YES list to explicitly credit "let the office know," transfers, "already being worked," accepts declines), I11 (validate — written from empty), I12 (neutral — full attorney protocol), I14 (secure info — clarified "reading back caller's own info ≠ DQ"). Not touched: I2, I4, I5, I7, I10, I13.</p>
<p><strong>What is left.</strong> Name {name_pct}% ({name_q['strict']}S) is still the largest remaining gap — dropping last-name closed 34 stricts, not all 118. Second tier: greeting {greet_pct}% ({greet_q['strict']}S), unit {unit_pct}% ({unit_q['strict']}S), acknowledged {ack_pct}% ({ack_q['strict']}S), validate {val_pct}% ({val_q['strict']}S/{val_q['lenient']}L), next steps {next_pct}% ({next_q['strict']}S). Contact {contact_pct}% and hold {hold_pct}% are locked. FHA still 100%. <strong>→ Open Recommendations</strong> for the next-cycle knob-turns.</p>"""

    root_strict = f"""<p><strong>Name usage ({name_q['strict']} strict / {name_q['lenient']} lenient, {name_q['weight']*name_q['strict']} weighted pts):</strong> Still the worst — {name_pct}% (up from {name_pct1}%). AI Yes-rate {round(name_q['ai_yes']/name_q['total']*100,1)}% vs human {round(name_q['h_yes']/name_q['total']*100,1)}%. Dropping the last-name requirement in I3 recovered 34 stricts. The remaining {name_q['strict']} are misses where humans credit a name that appears in the transcript once (agent thanks the caller by first name, spells the name back, addresses them mid-call) but AI still says No. These are follow-through, not protocol. Options: (a) tighten I3 wording of "used it at least once" to explicitly include spell-back and mid-call address; (b) put 3–5 example call IDs in column J.</p>
<p><strong>Greeting ({greet_q['strict']}S/0L), unit ({unit_q['strict']}S/0L), acknowledged ({ack_q['strict']}S/{ack_q['lenient']}L):</strong> Second-tier strict cluster. All three have Yes-friendly column I today (I2 credits property+name OR standard greeting; I5 credits asking; I7 credits first-person callback / note language). These are the same follow-through misses that showed up in Run 1 — the (20) edits did not touch them, so the numbers barely moved (43→{greet_q['strict']}, 40→{unit_q['strict']}, 39→{ack_q['strict']}). Do not rewrite I2 / I5 / I7. Put example call IDs in column J or accept the current level.</p>
<p><strong>Validate concern strict ({val_q['strict']}S) — new pile from I11:</strong> Run 1 had 3 strict; Run 2 has {val_q['strict']}. The new I11 says "if the caller states a problem, delay, missed callback, outage, lockout, access issue, billing surprise, safety issue, or clear frustration, the agent must use at least one validating phrase." That list is broad — model is triggering on almost any problem-adjacent language and demanding an explicit apology. Loosen the trigger (require actual frustration / emotion, not just any request) or expand the validating-phrase list (e.g. "let me help you with that," "I'll take care of that," "I can help").</p>
<p><strong>Next steps ({next_q['strict']}S) and reason for the call ({reason_q['strict']}S):</strong> Both cleared 90% (91.6% and 94.1%). Remaining strict cases are edge-case truncations and callers with no stated reason where humans still credited a Yes. Not worth another protocol edit — the risk of losing the shipped gain is higher than the return.</p>"""

    contact_q1 = q1map["contact_info"]
    root_lenient = f"""<p><strong>Validate concern lenient ({val_q['lenient']}L, {val_q['weight']*val_q['lenient']} over-credit pts):</strong> Dropped from {val_q1['lenient']} to {val_q['lenient']} because I11 was written. Remaining lenient are calls where the caller stated a problem, the agent said something conversational ("sure," "no problem," "I can help"), and AI credited that as validation but humans wanted an explicit sorry/understand phrase. Tighten the validating-phrase list to only count "I'm sorry," "I understand," "I'm sorry to hear that," "that's frustrating," "that makes sense" — drop the softer helpers ("sure," "no problem," "of course").</p>
<p><strong>Reason for the call lenient ({reason_q['lenient']}L, up from {reason_q1['lenient']}):</strong> Expected side effect of reverting I6 — the (16) bar credits more Yeses, including some that humans said No. Do not tighten I6 back. {reason_q['lenient']} at 7 pts is {reason_q['weight']*reason_q['lenient']} points of over-credit; the strict recovery ({reason_q1['strict']} → {reason_q['strict']}) was worth {reason_q['weight']*(reason_q1['strict']-reason_q['strict'])} points saved — net positive.</p>
<p><strong>Next steps lenient ({next_q['lenient']}L, up from {next_q1['lenient']}):</strong> Same trade-off as reason. I8 expansion credited more Yeses; humans said No on {next_q['lenient']} of them. Net still favorable ({r1['total_l']}→{m['total_l']} on lenient overall; {r1['total_s']}→{m['total_s']} on strict).</p>
<p><strong>Name lenient ({name_q['lenient']}L, up from {name_q1['lenient']}):</strong> Dropping the last-name bar in I3 created {name_q['lenient'] - name_q1['lenient']} new false Yeses. Very small compared to the {name_q1['strict']-name_q['strict']} strict cases recovered.</p>
<p><strong>Contact info lenient ({contact_q['lenient']}L vs {contact_q1['lenient']}):</strong> Nudged up by {contact_q['lenient']-contact_q1['lenient']}. I4 was not touched. Noise, not a trend.</p>"""

    recs = [
        {
            "num": 1,
            "title": f"Ship the file (20) protocols as the new baseline — {m['agreement_pct']}% cleared 90%",
            "severity": "success",
            "severityLabel": f"+{agree_delta}pp vs Large Test · {dis_delta:+d} disagrees · {strict_delta:+d} strict · perfect {r1['perfect']}→{m['perfect']}",
            "owner": "Austin + AI Engineering",
            "ownerClass": "info",
            "problem": f"""<p>Recs 1, 3, and 4 from Large Test all shipped in file (20) and all three cleared. This is the first run of the entire resident series that hit the 90% agreement target.</p>
<ul>
<li>Reason for the call: {reason_pct1}% → <strong>{reason_pct}%</strong> (I6 reverted). 63 strict → {reason_q['strict']}. Human Yes-rate 95%, AI Yes-rate {round(reason_q['ai_yes']/reason_q['total']*100,1)}% (was 84.9%).</li>
<li>Next steps: {next_pct1}% → <strong>{next_pct}%</strong> (I8 expanded YES list). 88 strict → {next_q['strict']}. AI Yes-rate 81.0% → {round(next_q['ai_yes']/next_q['total']*100,1)}%.</li>
<li>Name usage: {name_pct1}% → <strong>{name_pct}%</strong> (I3 dropped last-name). 118 strict → {name_q['strict']}. AI Yes-rate 69.0% → {round(name_q['ai_yes']/name_q['total']*100,1)}%.</li>
</ul>
<p>Cost was {lenient_delta:+d} lenient (net trade: {abs(strict_delta) - lenient_delta} agreements recovered). Freeze this rev. Do not roll back I3 / I6 / I8. Do not add competing protocols in those cells.</p>""",
            "protocols": [
                {
                    "label": "Baseline — freeze file (20) I3 / I6 / I8; move on",
                    "current": f"AI_ QA 2026 (20).xlsx · AI Resident Fundamentals · rows 3 / 6 / 8 shipped and cleared the 90% gate.",
                    "recommended": (
                        "<strong style=\"color:var(--red);\">Freeze I3, I6, I8 as the new resident baseline.</strong> "
                        "Do not tighten reason back toward file (17). Do not add negative examples to I8 in this cycle. "
                        "Do not require a second name reuse in I3. Any protocol change in the next cycle should target "
                        "validate (rec 2), name follow-through (rec 3), or the second-tier strict cluster (rec 4)."
                    ),
                }
            ],
        },
        {
            "num": 2,
            "title": f"Validate concern — I11 over-strict on second-order concern language ({val_q['strict']}S / {val_q['lenient']}L)",
            "severity": "critical",
            "severityLabel": f"{val_pct}% (was {val_pct1}%) · {val_q['strict']}S / {val_q['lenient']}L · AI Yes {round(val_q['ai_yes']/val_q['total']*100,1)}% vs human {round(val_q['h_yes']/val_q['total']*100,1)}%",
            "owner": "AI Engineering",
            "ownerClass": "info",
            "problem": f"""<p>Writing I11 was correct — Large Test validate was 3S/67L, so lenient dominated. Run 2 flipped {val_q1['lenient']-val_q['lenient']} of those lenients to agree, but created {val_q['strict']-val_q1['strict']} new stricts. Net agreement is {val_pct}% ({val_pct1}%→{val_pct}%, {val_pct-val_pct1:+.1f}pp — a regression on this question).</p>
<p>The current I11 trigger list is broad: "problem, delay, missed callback, outage, lockout, access issue, billing surprise, safety issue, or clear frustration." Almost every resident maintenance call fits. Model is now demanding "I'm sorry" on transactional calls where the caller was calm and said "my package didn't come." Humans grade those as Yes-by-default (no expressed frustration).</p>
<p>Two knobs: (a) tighten the trigger — only require validation when the caller expressed emotion / frustration / repetition ("this is the third time," "I'm frustrated," "this is unacceptable"). (b) expand the validating phrase list — count "let me take care of that," "I can help with that," "I'll get someone on it" as validation. (a) is the higher-leverage move.</p>""",
            "protocols": [
                {
                    "label": "Validate concern — tighten I11 trigger to expressed emotion, not any problem",
                    "current": (
                        "If no concern, problem, delay, or frustration is in the transcript, mark YES.<br><br>"
                        "If the caller states a problem, delay, missed callback, outage, lockout, access issue, billing surprise, safety issue, or clear frustration, the agent must use at least one validating phrase. "
                        "Examples: \"I'm sorry,\" \"I'm sorry to hear that,\" \"I understand,\" \"that makes sense,\" \"that's frustrating,\" \"that's not good,\" \"oh no.\"<br><br>"
                        "Mark NO when that kind of concern is present and the agent only processes the call (Okay / got it / name / unit / callback / \"I'll send a message\") with no validating phrase. Taking a message is not validation.<br><br>"
                        "One validating phrase is enough. Do not require multiple apologies or \"I will personally follow up.\""
                    ),
                    "recommended": (
                        "<strong style=\"color:var(--red);\">Narrow the trigger.</strong> A first-turn transactional request "
                        "(\"my package didn't come,\" \"I locked myself out,\" \"the AC isn't working\") is not automatically a concern that requires validation.<br><br>"
                        "<strong>Mark YES if any of these is true:</strong><br>"
                        "- No concern, problem, delay, or frustration is in the transcript<br>"
                        "- The caller stated a request in a neutral tone and the agent moved to next steps without a validating phrase<br>"
                        "- The agent used a validating phrase (\"I'm sorry,\" \"I'm sorry to hear that,\" \"I understand,\" \"that makes sense,\" \"that's frustrating,\" \"oh no,\" <strong style=\"color:var(--red);\">or</strong> \"let me help you with that,\" \"I'll take care of that,\" \"I can help with that,\" \"let's get this sorted\")<br><br>"
                        "<strong>Mark NO only when both:</strong><br>"
                        "1. The caller expressed clear frustration, repetition (\"this is the third time,\" \"I've been waiting\"), safety concern, or emotional distress; AND<br>"
                        "2. The agent responded without any validating or empathetic phrase, just processing the call.<br><br>"
                        "Do not demand \"I'm sorry\" on calm first-turn maintenance requests. Do not count silence during hold as failure. "
                        f"Gate: validate ≥ 90%, strict &lt; 15, lenient &lt; 40."
                    ),
                }
            ],
        },
        {
            "num": 3,
            "title": f"Name usage — {name_pct}% is still the largest remaining gap ({name_q['strict']} strict)",
            "severity": "warning",
            "severityLabel": f"{name_pct}% (was {name_pct1}%) · {name_q['strict']}S / {name_q['lenient']}L · AI Yes {round(name_q['ai_yes']/name_q['total']*100,1)}% vs human {round(name_q['h_yes']/name_q['total']*100,1)}%",
            "owner": "AI Engineering",
            "ownerClass": "info",
            "problem": f"""<p>I3 dropping the last-name bar recovered 34 of 118 stricts. The remaining {name_q['strict']} are calls where humans see a name being used (agent thanks the caller by first name, spells the name back, addresses them mid-call) but AI still says No. AI Yes-rate is {round(name_q['ai_yes']/name_q['total']*100,1)}%, human is {round(name_q['h_yes']/name_q['total']*100,1)}% — still a {round(name_q['h_yes']/name_q['total']*100,1) - round(name_q['ai_yes']/name_q['total']*100,1):.1f}pp gap.</p>
<p>Two options: (a) further loosen I3 to explicitly credit spell-back ("So that's L-U-V-I-A? Okay, Luvia") and message-taking name confirmations as "using the name." (b) put 5–10 of these {name_q['strict']} strict call IDs in column J (Additional Prompts) as concrete examples. Option (a) is the higher-leverage change — the current I3 says "using the name includes any of these" and lists three cases, but does not include spell-back. Add it.</p>""",
            "protocols": [
                {
                    "label": "Name usage — expand I3's 'using the name' list to include spell-back and message-taking confirmation",
                    "current": (
                        "The transcript must show that the agent either asked for the caller's name or the caller volunteered it, AND the agent used that name at least once. First name is enough. Last name is not required. Both must occur for Yes.<br><br>"
                        "Using the name includes any of these:<br>"
                        "- Addressing the caller by name (\"Thanks, Mia\")<br>"
                        "- Repeating or confirming it while taking the message (\"Lisa Butler… okay, Lisa\")<br>"
                        "- A close transcript spelling of the same name (Lluvia / Yuvia, Vania / Tanya, Antoine / Antonian). Do not mark No because Deepgram spelled it differently."
                    ),
                    "recommended": (
                        "Keep the file (20) I3 verbatim. Add these to the \"Using the name includes\" list:<br><br>"
                        "- <strong style=\"color:var(--red);\">Spelling the name back to confirm</strong> (\"L-U-V-I-A? Okay, Luvia\") counts as using it.<br>"
                        "- <strong style=\"color:var(--red);\">Reading the name during a message summary or callback confirmation</strong> (\"Okay, I have you down as Kimberly, unit 208, phone 555-1234\") counts as using it.<br>"
                        "- <strong style=\"color:var(--red);\">Addressing the caller by name during hold-return or hand-off</strong> (\"Thanks for holding, Devin\") counts as using it.<br>"
                        "- <strong style=\"color:var(--red);\">Truncated names</strong> (agent says \"Kim\" when caller gave \"Kimberly\") count as using it.<br><br>"
                        f"Do not require the name to appear twice. Do not require the name in the greeting. Gate: name ≥ 85%, strict &lt; 60. Accept the {name_q['lenient']} lenient as noise."
                    ),
                }
            ],
        },
        {
            "num": 4,
            "title": f"Second-tier strict cluster — greeting / unit / acknowledged (I2 / I5 / I7 already Yes-friendly)",
            "severity": "warning",
            "severityLabel": f"Greeting {greet_pct}% ({greet_q['strict']}S) · unit {unit_pct}% ({unit_q['strict']}S) · acknowledged {ack_pct}% ({ack_q['strict']}S)",
            "owner": "AI Engineering",
            "ownerClass": "info",
            "problem": f"""<p>Greeting, unit, and acknowledged all sit in the 91–93% band. I2 (greeting), I5 (unit), and I7 (acknowledged) were not touched between Large Test and Large Test 2 — the numbers moved from 43 → {greet_q['strict']}, 40 → {unit_q['strict']}, 39 → {ack_q['strict']} on strict. All three protocols are already Yes-friendly. These are follow-through misses, not missing rules.</p>
<p>Do not rewrite I2, I5, or I7 in the same cycle as rec 2 (validate) or rec 3 (name expansion). Each protocol change costs about 3–5 disagreements of read-noise as the model recalibrates. Pick one at a time.</p>
<p>Cheapest path: put 5 example call IDs from each strict list into column J (Additional Prompts). That is enforcement without rewriting the protocol. If cell J doesn't move the number after one paste, then rewrite I5 to explicitly say "unit number confirmed also includes: caller says 'unit 208' before agent asks; agent restates 'apartment 208' during message-taking."</p>""",
            "protocols": [
                {
                    "label": "Second-tier — enforcement via column J, no I2/I5/I7 rewrites this cycle",
                    "current": f"I2 (greeting) at {greet_pct}% · I5 (unit) at {unit_pct}% · I7 (acknowledged/ownership) at {ack_pct}%. All three are Yes-friendly today. All three still have double-digit strict piles.",
                    "recommended": (
                        "For this cycle, put 5 strict call IDs per question in column J (Additional Prompts) as \"treat these as YES; the current I passed for these\":<br><br>"
                        f"- Greeting J: {', '.join(greet_q['strict_ids'][:5])}<br>"
                        f"- Unit J: {', '.join(unit_q['strict_ids'][:5])}<br>"
                        f"- Acknowledged J: {', '.join(ack_q['strict_ids'][:5])}<br><br>"
                        "Do not rewrite I2 / I5 / I7. Do not add negative examples. If column J moves the numbers >2pp per question, keep it. If not, rewrite I5 first (highest weighted at 4 pts × 37 strict = 148 pts). Do not touch I2 or I7 until validate (rec 2) and name (rec 3) are locked."
                    ),
                }
            ],
        },
        {
            "num": 5,
            "title": f"Do not touch hold / contact / FHA / secure info — {round(hold_q['agree']/hold_q['total']*100,1)}% / {contact_pct}% / 100% / 99.6%",
            "severity": "success",
            "severityLabel": f"Hold {round(hold_q['agree']/hold_q['total']*100,1)}% · contact {contact_pct}% · FHA 100% · secure info 99.6%",
            "owner": "Austin + AI Engineering",
            "ownerClass": "info",
            "problem": f"""<p>These four are locked. I10 (hold — lookup ≠ hold), I4 (contact — attempt is enough, N/A / anonymous = Yes), I13 (FHA), and I14 (secure info — reading back caller's own info ≠ DQ). Contact nudged from 96.1% to {contact_pct}% (13→{contact_q['lenient']} lenient); noise, not signal.</p>
<p>Do not rewrite any of these. Do not add strict edge cases. Any change here is more likely to break the current agreement than to improve it.</p>
<p><strong>Next-cycle target math:</strong> {m['total_agree']}/{m['total_comp']} today ({m['agreement_pct']}%). If rec 2 (validate) drops the strict pile from {val_q['strict']} back to ≤10 without losing more than {val_q['lenient']+5} lenient, that's +{val_q['strict']-10} - 5 = ~{val_q['strict']-15} net agreements → {round((m['total_agree']+val_q['strict']-15)/m['total_comp']*100,1)}%. If rec 3 (name) recovers 20 more strict without adding >5 lenient, that's +15 → another 0.3pp. Combined target: ≥92% on Large Test 3.</p>""",
            "protocols": [
                {
                    "label": "Locked — do not modify",
                    "current": (
                        f"I10 (hold): {round(hold_q['agree']/hold_q['total']*100,1)}% ({hold_q['strict']}S / {hold_q['lenient']}L)<br>"
                        f"I4 (contact): {contact_pct}% ({contact_q['strict']}S / {contact_q['lenient']}L)<br>"
                        "I13 (FHA): 100% (0/0)<br>"
                        "I14 (secure info): 99.6% (0S / 2L)"
                    ),
                    "recommended": (
                        "<strong style=\"color:var(--red);\">Freeze.</strong> Any edit to these four rows in the file (21+) cycle "
                        "increases the risk of losing the current agreement more than it improves it. Focus the next protocol paste on I11 (validate — rec 2) and the I3 expansion (name — rec 3). Do not touch I4, I10, I13, I14."
                    ),
                }
            ],
        },
    ]

    root_cause = {
        "overall": f"Large Test 2 is {m['agreement_pct']}% scored agreement on the same 490 July/August resident calls (+{agree_delta}pp vs Large Test). The three file (20) protocol edits (I3 name, I6 reason, I8 next steps) hit their targets: reason {reason_pct1}%→{reason_pct}%, next steps {next_pct1}%→{next_pct}%, name {name_pct1}%→{name_pct}%. Validate regressed {val_pct1}%→{val_pct}% because writing I11 traded {val_q1['lenient']-val_q['lenient']} lenient for {val_q['strict']-val_q1['strict']} strict. Dominant bias is still AI too strict ({strict_share}% of disagrees), but the strict pile shrank {r1['total_s']}→{m['total_s']}.",
        "what_worked": f"Reason for the call cleared 90% (+{reason_pct-reason_pct1:.1f}pp) after I6 reverted the file (17) tightening. Next steps cleared 90% (+{next_pct-next_pct1:.1f}pp) after I8's YES list was expanded. Name usage +{name_pct-name_pct1:.1f}pp after I3 dropped the last-name bar. Hold, contact, FHA, secure info all held. Perfect calls {r1['perfect']}→{m['perfect']} ({perfect_delta:+d}). Mean AI score {r1['mean_a']}→{m['mean_a']}%.",
        "what_didnt": f"Validate concern regressed {val_pct1}%→{val_pct}% — the new I11 is too broad on which caller statements trigger the requirement, so it's now demanding \"I'm sorry\" on calm first-turn transactional calls. Name still 79.8% ({name_q['strict']} strict) — needs a second I3 pass. Greeting / unit / acknowledged barely moved (43→{greet_q['strict']}, 40→{unit_q['strict']}, 39→{ack_q['strict']}) because I2 / I5 / I7 were not touched.",
        "path_to_90": f"Already cleared 90%. Next-cycle target: ≥92%. Tighten I11 trigger to expressed emotion / repetition (rec 2, ~{val_q['strict']-10} agreements). Expand I3 \"using the name\" list to include spell-back and message-taking (rec 3, ~20 agreements). Enforcement via column J for I2 / I5 / I7 (rec 4). Do not touch hold / contact / FHA / secure info (rec 5). Do not roll back I3 / I6 / I8.",
        "strictDetail": root_strict,
        "lenientDetail": root_lenient,
    }

    return key_findings, root_cause, recs


# Content builders per run
CONTENT_BUILDERS = {
    90: build_run1_content,
    91: build_run2_content,
}


# ── JS emit ─────────────────────────────────────────────────────────────────
def emit_recs(recs_list):
    chunks = []
    for r in recs_list:
        proto_js = []
        for p in r.get("protocols") or []:
            proto_js.append(
                "          {\n"
                f'            label: "{js_escape(p["label"])}",\n'
                f'            current: "{js_escape(p["current"])}",\n'
                f'            recommended: `{p["recommended"]}`\n'
                "          }"
            )
        proto_block = ",\n".join(proto_js)
        chunks.append(
            f"""      {{
        num: {r['num']},
        title: "{js_escape(r['title'])}",
        severity: "{r['severity']}",
        severityLabel: "{js_escape(r['severityLabel'])}",
        owner: "{r['owner']}",
        ownerClass: "{r['ownerClass']}",
        problem: `{r['problem']}`,
        protocols: [
{proto_block}
        ]
      }}"""
        )
    return "    recommendations: [\n" + ",\n".join(chunks) + "\n    ]"


def emit_questions(qs):
    lines = []
    for q in qs:
        s_ids = ",".join(f'"{i}"' for i in q["strict_ids"])
        l_ids = ",".join(f'"{i}"' for i in q["lenient_ids"])
        lines.append(
            f'      {{ short: "{q["short"]}", label: "{js_escape(q["label"])}", weight: {q["weight"]}, '
            f'agree: {q["agree"]}, disagree: {q["disagree"]}, total: {q["total"]}, '
            f'lenient: {q["lenient"]}, strict: {q["strict"]}, dir: "{q["dir"]}",\n'
            f"        evidence: [], strictCalls: [{s_ids}], lenientCalls: [{l_ids}] }}"
        )
    return "    questions: [\n" + ",\n".join(lines) + "\n    ]"


def emit_calls(calls):
    lines = []
    for c in calls:
        details_parts = []
        for d in c["details"]:
            w_js = f'"{d["w"]}"' if d["w"] == "DQ" else str(d["w"])
            details_parts.append(
                f'          {{ q: "{js_escape(d["q"])}", w: {w_js}, ai: "{d["ai"]}", h: "{d["h"]}", reason: "{js_escape(d["reason"])}" }}'
            )
        details_str = ",\n".join(details_parts)
        if details_str:
            details_str = "\n" + details_str + "\n        "
        lines.append(
            f'      {{ id: "{c["id"]}", human: {c["human"]}, ai: {c["ai"]}, strict: {c["strict"]}, lenient: {c["lenient"]}, weightLost: {c["weightLost"]}, disqualifier: {"true" if c["disqualifier"] else "false"},\n'
            f"        note: '{js_escape(c['note'])}',\n"
            f"        details: [{details_str}],\n"
            f"        recommendation: '{js_escape(c['recommendation'])}' }}"
        )
    return "    calls: [\n" + ",\n".join(lines) + "\n    ]"


def emit_run_js(run, other_runs):
    m = run["metrics"]
    cfg = run["cfg"]
    if cfg["run_id"] == 90:
        key_findings, root_cause, recs = build_run1_content(run)
    elif cfg["run_id"] == 91:
        # Run 2 references Run 1 for deltas
        run1 = next(r for r in other_runs if r["cfg"]["run_id"] == 90)
        key_findings, root_cause, recs = build_run2_content(run, run1)
    else:
        raise SystemExit(f"No content builder for run id {cfg['run_id']}")

    return f"""  {{
    id: {cfg['run_id']},
    label: "{cfg['label']}",
    date: "{cfg['date']}",
    largeEval: true,
    totalPoints: {TOTAL_POINTS},
    scoredQuestions: {len(SCORED_KEYS)},
    criteriaCount: {len(ALL_KEYS)},
    description: "{js_escape(cfg['description'])}",
    changes: "{js_escape(cfg['changes'])}",
    meta: {{
      agreement: {m['agreement_pct']},
      totalDisagreements: {m['total_dis']},
      avgDelta: {m['avg_delta']},
      strictErrors: {m['total_s']},
      lenientErrors: {m['total_l']},
      target: 90
    }},
    keyFindings: `{key_findings}`,
{emit_questions(run['questions_data'])},
{emit_calls(run['calls_data'])},
    recTargetDisagreements: {int(m['total_comp'] * 0.10)},
{emit_recs(recs)},
    rootCause: {{
      overall: "{js_escape(root_cause['overall'])}",
      what_worked: "{js_escape(root_cause['what_worked'])}",
      what_didnt: "{js_escape(root_cause['what_didnt'])}",
      path_to_90: "{js_escape(root_cause['path_to_90'])}",
      strictDetail: `{root_cause['strictDetail']}`,
      lenientDetail: `{root_cause['lenientDetail']}`
    }},
    matrixQuestions: [
      {{ short: "Greeting", key: "greeting" }},
      {{ short: "Name", key: "name_usage" }},
      {{ short: "Contact info", key: "contact_info" }},
      {{ short: "Unit #", key: "unit_number" }},
      {{ short: "Reason for call", key: "reason_for_call" }},
      {{ short: "Acknowledged", key: "acknowledged" }},
      {{ short: "Next steps", key: "next_steps" }},
      {{ short: "Hold perm.", key: "hold_permission" }},
      {{ short: "Validate concern", key: "validate_concern" }},
      {{ short: "FHA (DQ)", key: "fha" }},
      {{ short: "Secure info (DQ)", key: "secure_info" }}
    ]
  }}"""


def emit_answers(run):
    ad_lines = []
    for cid in run["both"]:
        cd = run["answer_data"][cid]
        parts = []
        for q in ALL_KEYS:
            h, a = cd[q]
            h_str = str(h) if h is not None else "null"
            a_str = str(a) if a is not None else "null"
            parts.append(f"{q}:[{h_str},{a_str}]")
        ad_lines.append(f'    "{cid}": {{{",".join(parts)}}}')
    return f"  {run['cfg']['run_id']}: {{\n" + ",\n".join(ad_lines) + "\n  }"


# Emit both runs
run_blocks = [emit_run_js(r, runs_built) for r in runs_built]
answer_blocks = [emit_answers(r) for r in runs_built]

RUN_START = "// ═══ LARGE TEST START ═══"
RUN_END = "// ═══ LARGE TEST END ═══"
ANS_START = "// ═══ LARGE TEST ANSWERS START ═══"
ANS_END = "// ═══ LARGE TEST ANSWERS END ═══"

run_block = f"{RUN_START}\n" + ",\n".join(run_blocks) + f"\n  {RUN_END}"
ans_block = f"{ANS_START}\n" + ",\n".join(answer_blocks) + f"\n  {ANS_END}"


def replace_between(html, start, end, new_block):
    if start in html and end in html:
        i0 = html.index(start)
        i1 = html.index(end) + len(end)
        return html[:i0] + new_block + html[i1:], True
    return html, False


print("\nPatching HTML...")
with open(HTML_PATH, "r") as f:
    html = f.read()

html, ok = replace_between(html, RUN_START, RUN_END, run_block)
if ok:
    print(f"  Replaced Large Test block(s) — {len(run_blocks)} run(s) inside")
else:
    needle = "\n];\n\nconst residentMatrixQuestions = ["
    if needle not in html:
        raise SystemExit("Could not find residentRuns closing / residentMatrixQuestions")
    html = html.replace(
        needle,
        ",\n\n" + run_block + "\n];\n\nconst residentMatrixQuestions = [",
        1,
    )
    print(f"  Inserted Large Test block(s) — {len(run_blocks)} run(s) inside")

html, ok = replace_between(html, ANS_START, ANS_END, ans_block)
if ok:
    print(f"  Replaced Large Test answer block(s) — {len(answer_blocks)} run(s) inside")
else:
    needle = "\n};\n\nconst residentRuns = ["
    if needle not in html:
        raise SystemExit("Could not find residentRunAnswerData closing")
    html = html.replace(
        needle,
        ",\n  " + ans_block + "\n};\n\nconst residentRuns = [",
        1,
    )
    print(f"  Inserted Large Test answer block(s) — {len(answer_blocks)} run(s) inside")

# Subtitle
old_sub_1 = "Run 1.0: 12 calls \\u00b7 Run 2.0–5.0 + 8.0: 20 calls \\u00b7 12 scoring criteria"
old_sub_2 = "Run 1.0: 12 \\u00b7 Run 2.0\\u20138.0: 20 \\u00b7 Large Test: 490 calls \\u00b7 overlapping 9 scored + 2 DQs"
new_sub = "Run 1.0: 12 \\u00b7 Run 2.0\\u20138.0: 20 \\u00b7 Large Test 1+2: 490 calls each \\u00b7 overlapping 9 scored + 2 DQs"
if new_sub in html:
    print("  Subtitle already mentions Large Test 1+2")
elif old_sub_2 in html:
    html = html.replace(old_sub_2, new_sub)
    print("  Updated resident leadConfig subtitle (from Large Test 1 baseline)")
elif old_sub_1 in html:
    html = html.replace(old_sub_1, new_sub)
    print("  Updated resident leadConfig subtitle (from Run 1-8 baseline)")
else:
    print("  WARNING: could not update subtitle (pattern changed)")

with open(HTML_PATH, "w") as f:
    f.write(html)

print(f"\nWrote {HTML_PATH}")
for run in runs_built:
    m = run["metrics"]
    print(
        f"  {run['cfg']['label']}: {m['agreement_pct']}% agreement, "
        f"{m['total_dis']} disagreements, {m['avg_delta']}% avg delta, "
        f"{len(run['both'])} calls"
    )
print(f"  Date stamp {date.today().isoformat()}")
