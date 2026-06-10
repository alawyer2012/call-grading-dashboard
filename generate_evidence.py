#!/usr/bin/env python3
"""
Generate transcript evidence for every disagreed question across warm and cold leads.
Reads transcripts, cross-references with answer data, extracts relevant excerpts,
and injects evidence[] arrays into index.html question objects.
"""

import openpyxl
import json
import re
import os
import glob
import html as html_mod

# ═══════════════════════════════════════════════════════════════════════════
# CONFIGURATION
# ═══════════════════════════════════════════════════════════════════════════

HTML_PATH = "/Users/alawyer/Entrata PM/Dashboard/call-grading/index.html"

WARM_XLSX = "/Users/alawyer/Downloads/20 Call Warm Lead Comparison (1).xlsx"
WARM_TRANSCRIPT_DIR = "/Users/alawyer/Downloads/warm-lead-transcripts"

COLD_XLSX = "/Users/alawyer/Downloads/20 Call Cold Lead Comparison (2).xlsx"
COLD_TRANSCRIPT_DIR = "/Users/alawyer/Downloads/call-transcripts"

# Question-to-keyword mapping for transcript excerpt extraction
QUESTION_KEYWORDS = {
    "greeting": ["thank you for calling", "my name is", "how may i", "how can i", "welcome"],
    "name_usage": ["your name", "who am i speaking", "who do i have", "first and last", "my name is"],
    "conversational": ["how", "what", "tell me", "when", "where", "why"],
    "rapport": ["great", "wonderful", "exciting", "sorry", "understand", "appreciate", "welcome",
                 "awesome", "fantastic", "love", "glad", "happy", "nice", "perfect", "congratul",
                 "no worries", "definitely", "absolutely", "enjoy"],
    "phone": ["phone", "number", "call back", "callback", "cell", "contact number", "reach you"],
    "occupants": ["how many", "occupant", "people", "roommate", "live with", "who will be"],
    "pets": ["pet", "dog", "cat", "animal", "breed", "weight"],
    "pet_breed": ["breed", "what kind", "type of pet", "how much", "weight"],
    "how_heard": ["how did you hear", "how did you find", "where did you", "apartments.com", "zillow",
                   "find us", "hear about"],
    "open_ended_qs": ["what are you looking", "tell me about", "what brings you", "what's important",
                       "what do you", "how many bedroom", "when are you", "move-in", "what size"],
    "pricing_disclaimer": ["price", "rate", "rent", "cost", "fee", "subject to change",
                            "not guaranteed", "may vary", "special", "discount", "promotion", "month free"],
    "text_email_perm": ["text", "permission", "opt in", "consent"],
    "disclaimers": ["disclaimer", "subject to", "not guarantee", "change", "equal housing",
                     "fair housing", "availability"],
    "feature_amenity": ["pool", "gym", "fitness", "laundry", "parking", "balcony", "patio",
                         "renovated", "stainless", "granite", "washer", "dryer", "amenity",
                         "amenities", "dog park", "clubhouse", "business center"],
    "tour_offer": ["tour", "visit", "come by", "stop by", "showing", "schedule", "come see",
                    "check out", "walk through"],
    "email": ["email", "@", "gmail", "yahoo", "hotmail", "outlook", "address"],
    "pricing_value": ["value", "great deal", "competitive", "affordable", "worth", "savings",
                       "investment", "compare", "market"],
    "acknowledged": ["understand", "absolutely", "of course", "definitely", "certainly",
                      "i can help", "let me help", "no problem", "no worries", "i'd be happy",
                      "happy to help", "glad to", "i can assist"],
    "inclusive_lang": ["we ", "we'", " us ", " our ", "we do", "we have", "we offer",
                        "our community", "our property", "our team"],
    "closing": ["anything else", "is there anything", "thank you for calling", "have a great",
                 "have a good", "take care", "bye", "goodbye", "next step", "follow up"],
    "tour_next_step": ["schedule", "set up", "next step", "tour", "visit", "come by",
                        "callback", "call back", "follow up", "email you"],
    "fha": ["fair housing", "fha", "discriminat", "race", "religion", "familial"],
    "secure_info": ["social security", "ssn", "credit card", "bank account", "routing"],
    "no_contact": ["contact", "reach out", "follow up", "get back"],
}

STRICT_EXPLANATIONS = {
    "greeting": "Agent greeted with property name/intro, but AI didn't detect the pattern.",
    "name_usage": "Agent used caller's name during the call, but AI missed it.",
    "conversational": "Human credited conversational gathering, but AI required more explicit flow.",
    "rapport": "Agent showed personalization/empathy, but AI didn't credit it as rapport.",
    "phone": "Phone number was gathered or confirmed, but AI didn't detect the exchange.",
    "occupants": "Occupant information was discussed or volunteered, but AI missed it.",
    "pets": "Pet information was discussed or volunteered, but AI missed it.",
    "pet_breed": "Pet breed/type was discussed, but AI missed it.",
    "how_heard": "How-heard info was discussed or volunteered, but AI missed it.",
    "open_ended_qs": "Human counted open-ended questions, but AI grammar-parsed too literally.",
    "pricing_disclaimer": "Agent stated pricing disclaimer or context, but AI required exact wording.",
    "pricing_value": "Agent conveyed pricing value, but AI required more explicit language.",
    "disclaimers": "Agent stated required disclaimers, but AI missed due to phrasing variation.",
    "feature_amenity": "Agent mentioned features/amenities, but AI required more explicit benefit language.",
    "tour_offer": "Agent offered a tour/visit, but AI required exact 'tour' wording.",
    "email": "Email was gathered, but AI didn't detect the phrasing used.",
    "acknowledged": "Agent acknowledged caller's question, but AI didn't credit it.",
    "inclusive_lang": "Agent used 'we/us/our' language, but AI missed common phrasings.",
    "closing": "Agent confirmed next steps or closed the call, but AI required formal structure.",
    "tour_next_step": "Agent set a next step, but AI required a more explicit commitment.",
    "text_email_perm": "Agent offered text/email permission, but AI didn't detect the offer.",
    "fha": "Human found no FHA violation, but AI flagged one.",
    "secure_info": "Human found no secure info issue, but AI flagged one.",
    "no_contact": "Human found no contact failure, but AI flagged one.",
}

LENIENT_EXPLANATIONS = {
    "greeting": "AI credited the greeting, but human says it was insufficient.",
    "name_usage": "AI detected name usage, but human says name wasn't properly used.",
    "conversational": "AI credited info-gathering as conversational; human required genuine flow.",
    "rapport": "AI counted tone/pleasantries as rapport; human required personalization.",
    "phone": "AI detected phone exchange, but human says it wasn't properly gathered.",
    "occupants": "AI over-credited an occupant mention that wasn't a proper confirmation.",
    "pets": "AI over-credited a pet mention that wasn't a proper confirmation.",
    "pet_breed": "AI credited breed mention, but human says it wasn't properly discussed.",
    "how_heard": "AI over-credited without proper confirmation of how-heard.",
    "open_ended_qs": "AI counted questions as open-ended; human classified them as closed-ended.",
    "pricing_disclaimer": "AI credited a basic statement; human required explicit disclaimer.",
    "pricing_value": "AI credited generic pricing mention; human required value language.",
    "disclaimers": "AI credited disclaimers; human did not consider them sufficient.",
    "feature_amenity": "AI credited a feature mention; human required benefit-selling language.",
    "tour_offer": "AI credited a vague offer; human required explicit tour invitation.",
    "email": "AI credited email collection, but human disagrees.",
    "acknowledged": "AI credited acknowledgment; human required more explicit recognition.",
    "inclusive_lang": "AI detected inclusive language; human says usage was insufficient.",
    "closing": "AI counted a generic goodbye; human required confirmed next steps.",
    "tour_next_step": "AI credited a next step; human required more concrete commitment.",
    "text_email_perm": "AI credited text/email permission; human disagrees.",
    "fha": "AI found no FHA violation; human flagged one.",
    "secure_info": "AI found no secure info issue; human flagged one.",
    "no_contact": "AI found no contact failure; human flagged one.",
}


def extract_agent_name(transcript_text):
    """Extract agent name from greeting line."""
    m = re.search(r'[Mm]y name is (\w+)', transcript_text)
    if m:
        return m.group(1)
    return "Agent"


def find_relevant_excerpt(transcript_lines, question_key, max_lines=6):
    """Find the most relevant transcript excerpt for a question disagreement."""
    keywords = QUESTION_KEYWORDS.get(question_key, [])
    if not keywords:
        return None

    best_score = 0
    best_start = 0

    for i, line in enumerate(transcript_lines):
        lower = line.lower()
        score = sum(1 for kw in keywords if kw.lower() in lower)
        if score > best_score:
            best_score = score
            best_start = i

    if best_score == 0:
        for i, line in enumerate(transcript_lines):
            if line.strip().startswith("Channel 1:") and len(line.strip()) > 20:
                best_start = i
                break

    start = max(0, best_start - 1)
    end = min(len(transcript_lines), start + max_lines)
    excerpt_lines = transcript_lines[start:end]

    return excerpt_lines


def format_excerpt_html(lines, question_key):
    """Format transcript lines as HTML with speaker labels and keyword highlights."""
    keywords = QUESTION_KEYWORDS.get(question_key, [])
    html_parts = []

    for line in lines:
        line = line.strip()
        if not line:
            continue
        escaped = html_mod.escape(line)

        if escaped.startswith("Channel 1:"):
            escaped = escaped.replace("Channel 1:", '<span class="speaker">Agent:</span>', 1)
        elif escaped.startswith("Channel 0:"):
            escaped = escaped.replace("Channel 0:", '<span class="speaker">Caller:</span>', 1)

        for kw in keywords:
            pattern = re.compile(re.escape(html_mod.escape(kw)), re.IGNORECASE)
            escaped = pattern.sub(lambda m: f'<span class="highlight">{m.group()}</span>', escaped)

        html_parts.append(escaped)

    return "<br>".join(html_parts)


def js_escape_evidence(s):
    """Escape a string for embedding in a JS template literal."""
    return s.replace("\\", "\\\\").replace("`", "\\`").replace("${", "\\${")


# ═══════════════════════════════════════════════════════════════════════════
# LOAD WARM LEAD DATA
# ═══════════════════════════════════════════════════════════════════════════

print("Loading warm lead data...")

WARM_MANUAL_COL_MAP = {
    11: "disclaimers", 12: "tour_offer", 13: "email", 14: "acknowledged",
    15: "inclusive_lang", 16: "closing", 17: "greeting", 18: "name_usage",
    19: "conversational", 20: "rapport", 21: "phone", 22: "open_ended_qs",
    23: "pricing_disclaimer", 24: "text_email_perm", 25: "fha", 26: "secure_info",
}
WARM_AI_COL_MAP = {
    2: "greeting", 3: "name_usage", 4: "conversational", 5: "rapport",
    6: "phone", 7: "open_ended_qs", 8: "pricing_disclaimer", 9: "text_email_perm",
    10: "disclaimers", 11: "tour_offer", 12: "email", 13: "acknowledged",
    14: "inclusive_lang", 15: "closing", 16: "fha", 17: "secure_info",
}
WARM_SCORED = ["greeting", "name_usage", "conversational", "rapport", "phone",
               "open_ended_qs", "pricing_disclaimer", "disclaimers", "tour_offer",
               "email", "acknowledged", "inclusive_lang", "closing"]

def parse_yn(val):
    if val is None or str(val).strip() == "":
        return None
    return 1 if str(val).strip().lower() == "yes" else 0

wb_warm = openpyxl.load_workbook(WARM_XLSX, data_only=True)
warm_human = {}
ws = wb_warm['Manual']
for row in range(2, ws.max_row + 1):
    cid = ws.cell(row, 3).value
    if cid is None: continue
    cid = str(int(cid))
    warm_human[cid] = {WARM_MANUAL_COL_MAP[c]: parse_yn(ws.cell(row, c).value) for c in WARM_MANUAL_COL_MAP}

warm_ai = {}
ws2 = wb_warm['AI']
for row in range(2, ws2.max_row + 1):
    cid = ws2.cell(row, 1).value
    if cid is None: continue
    cid = str(int(cid))
    warm_ai[cid] = {WARM_AI_COL_MAP[c]: parse_yn(ws2.cell(row, c).value) for c in WARM_AI_COL_MAP}

warm_transcripts = {}
for f in glob.glob(os.path.join(WARM_TRANSCRIPT_DIR, "*.json")):
    cid = os.path.basename(f).replace(".json", "")
    with open(f) as fh:
        data = json.load(fh)
    paras = data.get("results", {}).get("paragraphs", {})
    full = paras.get("transcript", "").strip()
    warm_transcripts[cid] = full

print(f"  Warm: {len(warm_human)} human, {len(warm_ai)} AI, {len(warm_transcripts)} transcripts")

# ═══════════════════════════════════════════════════════════════════════════
# LOAD COLD LEAD DATA
# ═══════════════════════════════════════════════════════════════════════════

print("Loading cold lead data...")

COLD_MANUAL_COL_MAP = {
    3: "open_ended_qs", 4: "pricing_disclaimer", 5: "text_email_perm",
    6: "disclaimers", 7: "feature_amenity", 8: "tour_offer", 9: "email",
    10: "pricing_value", 11: "greeting", 12: "tour_next_step",
    13: "acknowledged", 14: "inclusive_lang", 15: "name_usage",
    16: "conversational", 17: "rapport", 18: "phone", 19: "occupants",
    20: "pets", 21: "how_heard", 22: "pet_breed", 23: "fha",
    24: "secure_info", 25: "no_contact", 26: "closing",
}
COLD_AI_COL_MAP = {
    2: "fha", 3: "secure_info", 4: "no_contact", 5: "greeting",
    6: "name_usage", 7: "conversational", 8: "rapport", 9: "phone",
    10: "occupants", 11: "pets", 12: "pet_breed", 13: "how_heard",
    14: "open_ended_qs", 15: "pricing_disclaimer", 16: "text_email_perm",
    17: "disclaimers", 18: "feature_amenity", 19: "tour_offer", 20: "email",
    21: "pricing_value", 22: "tour_next_step", 23: "acknowledged",
    24: "inclusive_lang", 25: "closing",
}
COLD_SCORED = ["greeting", "name_usage", "conversational", "rapport", "phone",
               "occupants", "pets", "pet_breed", "how_heard", "open_ended_qs",
               "pricing_disclaimer", "text_email_perm", "disclaimers", "feature_amenity",
               "tour_offer", "email", "pricing_value", "tour_next_step",
               "acknowledged", "inclusive_lang", "closing"]

wb_cold = openpyxl.load_workbook(COLD_XLSX, data_only=True)
cold_human = {}
ws = wb_cold['Manual']
for row in range(2, ws.max_row + 1):
    cid = ws.cell(row, 2).value
    if cid is None: continue
    cid = str(int(cid))
    cold_human[cid] = {COLD_MANUAL_COL_MAP[c]: parse_yn(ws.cell(row, c).value) for c in COLD_MANUAL_COL_MAP}

cold_ai_runs = {}
for tab_name in ["AI", "AI 2"]:
    if tab_name not in wb_cold.sheetnames:
        continue
    ws2 = wb_cold[tab_name]
    ai_data = {}
    for row in range(2, ws2.max_row + 1):
        cid = ws2.cell(row, 1).value
        if cid is None: continue
        cid = str(int(cid))
        ai_data[cid] = {COLD_AI_COL_MAP[c]: parse_yn(ws2.cell(row, c).value) for c in COLD_AI_COL_MAP}
    cold_ai_runs[tab_name] = ai_data
    print(f"  Cold {tab_name}: {len(ai_data)} calls")

cold_transcripts = {}
for f in glob.glob(os.path.join(COLD_TRANSCRIPT_DIR, "*.txt")):
    cid = os.path.basename(f).replace(".txt", "")
    with open(f) as fh:
        cold_transcripts[cid] = fh.read().strip()

print(f"  Cold: {len(cold_human)} human, {len(cold_transcripts)} transcripts")


# ═══════════════════════════════════════════════════════════════════════════
# GENERATE EVIDENCE
# ═══════════════════════════════════════════════════════════════════════════

def generate_evidence_for_question(q_key, human_data, ai_data, transcripts, call_ids, max_evidence=5):
    """Generate evidence entries for calls where AI and human disagree on a question."""
    evidence = []

    for cid in sorted(call_ids):
        h = human_data.get(cid, {}).get(q_key)
        a = ai_data.get(cid, {}).get(q_key)
        if h is None or a is None or h == a:
            continue

        transcript_text = transcripts.get(cid, "")
        if not transcript_text:
            continue

        lines = [l for l in transcript_text.split("\n") if l.strip()]
        agent_name = extract_agent_name(transcript_text)

        excerpt_lines = find_relevant_excerpt(lines, q_key)
        if not excerpt_lines:
            continue

        excerpt_html = format_excerpt_html(excerpt_lines, q_key)

        is_strict = (h == 1 and a == 0)
        explanation = STRICT_EXPLANATIONS.get(q_key, "AI missed this behavior.") if is_strict else LENIENT_EXPLANATIONS.get(q_key, "AI over-credited this behavior.")

        evidence.append({
            "callId": cid,
            "agent": agent_name,
            "ai": "Yes" if a == 1 else "No",
            "human": "Yes" if h == 1 else "No",
            "transcript": excerpt_html,
            "explanation": explanation,
        })

        if len(evidence) >= max_evidence:
            break

    return evidence


# Generate warm lead evidence
print("\nGenerating warm lead evidence...")
warm_call_ids = sorted(warm_ai.keys())
warm_evidence = {}
for q in WARM_SCORED:
    ev = generate_evidence_for_question(q, warm_human, warm_ai, warm_transcripts, warm_call_ids)
    if ev:
        warm_evidence[q] = ev
        print(f"  {q}: {len(ev)} evidence entries")

# Generate cold lead evidence for each run
print("\nGenerating cold lead evidence...")
cold_call_ids = sorted(cold_ai_runs.get("AI", {}).keys())
cold_evidence_by_run = {}
for tab_name, ai_data in cold_ai_runs.items():
    run_id = 1 if tab_name == "AI" else 2
    cold_evidence_by_run[run_id] = {}
    for q in COLD_SCORED:
        ev = generate_evidence_for_question(q, cold_human, ai_data, cold_transcripts, cold_call_ids)
        if ev:
            cold_evidence_by_run[run_id][q] = ev
            print(f"  Run {run_id} {q}: {len(ev)} evidence entries")


# ═══════════════════════════════════════════════════════════════════════════
# INJECT INTO HTML
# ═══════════════════════════════════════════════════════════════════════════

print("\nInjecting evidence into HTML...")

with open(HTML_PATH, "r") as f:
    html = f.read()


def evidence_to_js(evidence_list):
    """Convert evidence list to JS array string."""
    if not evidence_list:
        return "[]"
    parts = []
    for ev in evidence_list:
        t = js_escape_evidence(ev["transcript"])
        e = js_escape_evidence(ev["explanation"])
        parts.append(
            f'{{callId:"{ev["callId"]}",agent:"{js_escape_evidence(ev["agent"])}",'
            f'ai:"{ev["ai"]}",human:"{ev["human"]}",'
            f'transcript:`{t}`,'
            f'explanation:`{e}`}}'
        )
    return "[" + ",".join(parts) + "]"


def inject_evidence_into_runs(html, evidence_dict, runs_var_name):
    """Find question objects within a runs array and replace their evidence arrays."""
    changed = 0

    for q_key, ev_list in evidence_dict.items():
        ev_js = evidence_to_js(ev_list)

        pattern = re.compile(
            r'(short:\s*"' + re.escape(q_key) + r'"[^}]*?evidence:\s*)\[[^\]]*\]',
            re.DOTALL
        )

        search_start = html.find(f"const {runs_var_name} = [")
        if search_start < 0:
            continue

        if runs_var_name == "warmRuns":
            search_end = html.find("const warmMatrixQuestions", search_start)
        elif runs_var_name == "coldRuns":
            search_end = html.find("const coldMatrixQuestions", search_start)
        else:
            search_end = len(html)

        if search_end < 0:
            search_end = len(html)

        region = html[search_start:search_end]
        matches = list(pattern.finditer(region))

        for match in reversed(matches):
            replacement = match.group(1) + ev_js
            abs_start = search_start + match.start()
            abs_end = search_start + match.end()
            html = html[:abs_start] + replacement + html[abs_end:]
            changed += 1

    return html, changed


# Inject warm lead evidence (all runs share same evidence since we only have Run 1)
html, warm_count = inject_evidence_into_runs(html, warm_evidence, "warmRuns")
print(f"  Warm lead: {warm_count} evidence arrays injected")

# Inject cold lead evidence
# For cold leads, runs 1 and 2 may have different disagreements, but questions are in the runs array
# We need to inject into the right run's questions
for run_id, ev_dict in cold_evidence_by_run.items():
    html, cold_count = inject_evidence_into_runs(html, ev_dict, "coldRuns")
    print(f"  Cold lead Run {run_id}: {cold_count} evidence arrays injected")

with open(HTML_PATH, "w") as f:
    f.write(html)

print(f"\n✅ Evidence injection complete. HTML saved.")
total_ev = sum(len(v) for v in warm_evidence.values()) + sum(
    len(v) for rd in cold_evidence_by_run.values() for v in rd.values()
)
print(f"   Total evidence entries: {total_ev}")
