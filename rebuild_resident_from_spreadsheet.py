#!/usr/bin/env python3
"""
Resident rebuild: Extract data from '20 Call Resident Comparison' spreadsheet,
generate residentRuns / residentAnswerData / residentRunAnswerData, and patch
index.html (including leadConfig + PENDING_TYPES).
"""

import openpyxl
import re
import sys
from datetime import date

XLSX_PATH = "/Users/alawyer/Downloads/20 Call Resident Comparison (9).xlsx"
HTML_PATH = "/Users/alawyer/Entrata PM/Dashboard/call-grading/index.html"

# Full rebuild wipes hand-written keyFindings / recs / rootCause.
# To add a run without that, patch index.html only (see git history for Run 4.0 / 5.0 / 8.0).
# AI 8 C6 is a NEW question (reason for the call) replacing open-ended. Use ai_col_overrides.
AI_TABS = [
    {
        "tab": "AI",
        "id": 1,
        "label": "Run 1.0",
        "date": "August 11, 2026",
        "description": "Baseline — resident fundamentals scorecard (10 scored questions + 2 DQs). Partial benchmark: 12 calls with matching AI + human grades.",
        "changes": "Initial resident benchmark run (12 matching call IDs)",
    },
    {
        "tab": "AI 2",
        "id": 2,
        "label": "Run 2.0",
        "date": "August 20, 2026",
        "description": "Protocol updates from Run 1.0 recs 1–4 (hold, ownership, closing, secure-info DQ). Full 20-call set.",
        "changes": "Hold: lookup language is not a hold. Ownership: credit callback/note language. Closing: next steps without requiring “anything else?”. Secure-info: caller phone readback is not a DQ.",
    },
    {
        "tab": "AI 3",
        "id": 3,
        "label": "Run 3.0",
        "date": "August 21, 2026",
        "description": "Protocol pack from Run 2.0 recs 1–6 on the same 20 calls. Open-ended leftovers cleared, then over-corrected.",
        "changes": "Open-ended two-pole rule, validate split, voicemail-forward ownership, contact N/A on policy-only / refused callback.",
    },
    {
        "tab": "AI 5",
        "id": 4,
        "label": "Run 4.0",
        "date": "August 27, 2026",
        "description": "Same 20 calls. Spreadsheet tab AI 5 published as Run 4.0 (AI 4 tab was not shipped). Open-ended recovered 1; closing went 20/20. Still short of 90% and of Run 2.0’s 9 perfects.",
        "changes": "Resident I6 rewrite still in place (not the Run 2.0 revert). Closing conference/voicemail leftover cleared. Keep-No open-ended restored on 270583790 / 265908165.",
    },
    {
        "tab": "AI 6",
        "id": 5,
        "label": "Run 5.0",
        "date": "August 31, 2026",
        "description": "Same 20 calls. Spreadsheet tab AI 6 published as Run 5.0. Open-ended 10→9/20. Closing 20→19/20. Agreement 88.5% → 87.5%, still short of 90%.",
        "changes": "Ops-Yes 269788776 landed. 271912426 open-ended recovered. 273900663 false Yes cleared. Lost Yeses on 272543647 / 274349154 / 270842045. Keep-No 265908165 broke. Closing leftover on 269829113 returned.",
    },
    {
        "tab": "AI 8",
        "id": 8,
        "label": "Run 8.0",
        "date": "August 31, 2026",
        "description": "Same 20 calls. Spreadsheet tab AI 8 published as Run 8.0 (AI 7 was not shipped). C6 is now reason-for-call (replaces open-ended, still 7 pts). Agreement 87.5% → 89.5%, one scored agreement short of 90%.",
        "changes": "New question: “Did the agent capture the reason for the call?” Reason-for-call 9/20 → 15/20. Ops-Yes 272063323 recovered. Closing 19 → 17/20. Perfects 6 → 8.",
        "ai_col_overrides": {6: "reason_for_call"},
        "human_alias": {"reason_for_call": "open_ended_qs"},
    },
]

# New Manual: Col 1=Overall Score, Col 2=Call Id, Cols 3-14=answers
MANUAL_COL_MAP = {
    3: "greeting",
    4: "name_usage",
    5: "contact_info",
    6: "unit_number",
    7: "open_ended_qs",
    8: "acknowledged",
    9: "closing",
    10: "hold_permission",
    11: "validate_concern",
    12: "neutral_language",
    13: "fha",
    14: "secure_info",
}

# AI tab: Col 1=call_id, Cols 2-13=answers (hold/validate/neutral order differs from Manual)
AI_COL_MAP = {
    2: "greeting",
    3: "name_usage",
    4: "contact_info",
    5: "unit_number",
    6: "open_ended_qs",
    7: "acknowledged",
    8: "closing",
    9: "neutral_language",
    10: "hold_permission",
    11: "validate_concern",
    12: "fha",
    13: "secure_info",
}

WEIGHTS = {
    "greeting": 4,
    "name_usage": 3,
    "contact_info": 4,
    "unit_number": 4,
    "open_ended_qs": 7,
    "reason_for_call": 7,
    "acknowledged": 5,
    "closing": 3,
    "hold_permission": 2,
    "validate_concern": 5,
    "neutral_language": 5,
}
# open-ended and reason-for-call share the same 7-pt slot; never sum both
TOTAL_POINTS = 42
DQ_KEYS = ["fha", "secure_info"]
SCORED_KEYS = [k for k in WEIGHTS if k != "reason_for_call"]
ALL_KEYS = SCORED_KEYS + DQ_KEYS

Q_LABELS = {
    "greeting": "Greeting",
    "name_usage": "Name usage",
    "contact_info": "Contact info",
    "unit_number": "Unit number",
    "open_ended_qs": "Open-ended questions",
    "reason_for_call": "Reason for the call",
    "acknowledged": "Acknowledged/ownership",
    "closing": "Closing",
    "hold_permission": "Hold permission",
    "validate_concern": "Validate concern",
    "neutral_language": "Neutral language",
    "fha": "FHA violation (DQ)",
    "secure_info": "Secure info (DQ)",
}

Q_FULL_LABELS = {
    "greeting": "Greeting (property name + intro)",
    "name_usage": "Asked for name + used it",
    "contact_info": "Contact info confirmed (if update requested)",
    "unit_number": "Unit / apartment number confirmed",
    "open_ended_qs": "Open-ended questions (2+ distinct)",
    "reason_for_call": "Reason for the call (replaces open-ended)",
    "acknowledged": "Acknowledged caller / took ownership",
    "closing": "Closing / confirm next steps",
    "hold_permission": "Asked permission before hold",
    "validate_concern": "Validated caller concern (if expressed)",
    "neutral_language": "Neutral / de-escalating language",
}

STRICT_REASONS = {
    "greeting": "Agent greeted with property name and intro. AI did not detect it.",
    "name_usage": "Agent used caller's name but AI didn't detect it.",
    "contact_info": "Agent confirmed contact info. AI did not credit it.",
    "unit_number": "Unit/apartment number was confirmed. AI missed it.",
    "open_ended_qs": "Human counted questions as open-ended. AI grammar-parsed too literally.",
    "reason_for_call": "Human credited capturing the reason for the call. AI marked No.",
    "acknowledged": "Agent acknowledged caller / took ownership. AI did not credit it.",
    "closing": "Agent confirmed next steps. AI required more formal closing structure.",
    "hold_permission": "Agent asked hold permission (or hold not needed). AI did not credit it.",
    "validate_concern": "Human credited concern validation. AI did not.",
    "neutral_language": "Agent used neutral/de-escalating language. AI missed it.",
}

LENIENT_REASONS = {
    "greeting": "AI credited greeting; human says it was insufficient.",
    "name_usage": "AI detected name usage but human says name wasn't properly used.",
    "contact_info": "AI credited contact confirmation; human disagrees.",
    "unit_number": "AI credited unit confirmation; human disagrees.",
    "open_ended_qs": "AI counted questions as open-ended; human classified them as closed-ended.",
    "reason_for_call": "AI credited capturing the reason for the call; human open-ended grade is No.",
    "acknowledged": "AI credited acknowledgment; human required more explicit ownership.",
    "closing": "AI counted a generic goodbye; human required confirmed next steps.",
    "hold_permission": "AI credited hold permission; human says it was not properly asked.",
    "validate_concern": "AI credited concern validation; human did not.",
    "neutral_language": "AI credited neutral language; human found it insufficient.",
    "fha": "AI flagged an FHA violation; human did not find one.",
    "secure_info": "AI flagged secure info disclosure; human did not find one.",
}


def parse_yes_no(val):
    if val is None or str(val).strip() == "":
        return None
    return 1 if str(val).strip().lower().startswith("y") else 0


def slot_keys(answers):
    """Use reason-for-call instead of open-ended when that 7-pt slot is present."""
    if "reason_for_call" in answers:
        return ["reason_for_call" if k == "open_ended_qs" else k for k in SCORED_KEYS]
    return list(SCORED_KEYS)


def calc_score(answers):
    scored = slot_keys(answers)
    earned = sum(WEIGHTS[q] for q in scored if answers.get(q) == 1)
    base = (earned / TOTAL_POINTS) * 100 if TOTAL_POINTS else 0
    dq_count = sum(1 for dq in DQ_KEYS if answers.get(dq) == 1)
    return round(base * max(0, 1 - 0.20 * dq_count), 2)


def js_escape(s):
    return s.replace("\\", "\\\\").replace("'", "\\'").replace('"', '\\"').replace("`", "\\`")


print("Reading spreadsheet...")
wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)

# Prefer higher human overall score when duplicate call IDs appear (keeps 100% for 272543647)
ws_manual = wb["New Manual"]
human_data = {}
for row in range(2, ws_manual.max_row + 1):
    call_id_raw = ws_manual.cell(row, 2).value
    if call_id_raw is None:
        continue
    call_id = str(int(float(call_id_raw)))
    answers = {q_key: parse_yes_no(ws_manual.cell(row, col).value) for col, q_key in MANUAL_COL_MAP.items()}
    score = ws_manual.cell(row, 1).value
    sheet_score = float(score) if score is not None else 0.0
    if sheet_score > 1.5:
        sheet_score = sheet_score / 100.0
    if call_id in human_data and human_data[call_id]["sheet_score_norm"] >= sheet_score:
        print(f"  Skipping lower-score duplicate human row for {call_id} (kept {human_data[call_id]['sheet_score_norm']:.3f}, drop {sheet_score:.3f})")
        continue
    if call_id in human_data:
        print(f"  Replacing human row for {call_id} with higher score {sheet_score:.3f}")
    human_data[call_id] = {
        "answers": answers,
        "sheet_score": score,
        "sheet_score_norm": sheet_score,
    }

print(f"Extracted {len(human_data)} unique calls from New Manual tab")

ai_runs = {}
for run_info in AI_TABS:
    tab_name = run_info["tab"]
    if tab_name not in wb.sheetnames:
        print(f"⚠️  Tab '{tab_name}' not found — skipping")
        continue
    ws_ai = wb[tab_name]
    ai_data = {}
    for row in range(2, ws_ai.max_row + 1):
        call_id_raw = ws_ai.cell(row, 1).value
        if call_id_raw is None:
            continue
        call_id = str(int(float(call_id_raw)))
        col_map = dict(AI_COL_MAP)
        col_map.update(run_info.get("ai_col_overrides") or {})
        answers = {q_key: parse_yes_no(ws_ai.cell(row, col).value) for col, q_key in col_map.items()}
        if call_id in ai_data:
            print(f"  Skipping duplicate AI row for {call_id}")
            continue
        ai_data[call_id] = {"answers": answers}
    ai_runs[tab_name] = ai_data
    print(f"Extracted {len(ai_data)} unique calls from '{tab_name}' tab")

# Run 1.0 is a 12-call slice; later runs use every ID that has both human + that tab's AI grades.
def matching_ids(tab_name):
    return sorted(set(human_data.keys()) & set(ai_runs[tab_name].keys()))


benchmark_call_ids = matching_ids(AI_TABS[0]["tab"])
print(f"\nRun 1.0 matching set: {len(benchmark_call_ids)} calls")
print("  " + ", ".join(benchmark_call_ids))
human_only = sorted(set(human_data.keys()) - set(benchmark_call_ids))
if human_only:
    print(f"  Human IDs not in Run 1.0 AI tab (used on later runs): {', '.join(human_only)}")

print("\n" + "=" * 70)
print("SCORE VALIDATION — New Manual (benchmark calls)")
print("=" * 70)
score_mismatches = 0
for call_id in benchmark_call_ids:
    computed = calc_score(human_data[call_id]["answers"])
    sheet_score = round(human_data[call_id]["sheet_score_norm"] * 100, 2)
    delta = abs(computed - sheet_score)
    mark = "✅" if delta <= 2.0 else "❌"
    if delta > 2.0:
        score_mismatches += 1
    print(f"{mark} {call_id}: computed={computed}% sheet={sheet_score}% (Δ{delta:.2f})")
print(f"\nScore validation: {score_mismatches} mismatches out of {len(benchmark_call_ids)} (tolerance ±2pts)")


def build_run_data(human_data, ai_data, benchmark_ids, run_info=None):
    alias = (run_info or {}).get("human_alias") or {}
    sample_ai = next(iter(ai_data.values()), {}).get("answers", {}) if ai_data else {}
    run_scored = slot_keys(sample_ai)
    run_keys = run_scored + DQ_KEYS
    answer_data = {}
    for call_id in benchmark_ids:
        h = dict(human_data[call_id]["answers"])
        a = dict(ai_data.get(call_id, {}).get("answers", {}))
        for new_key, old_key in alias.items():
            if new_key not in h and old_key in h:
                h[new_key] = h[old_key]
        answer_data[call_id] = {q_key: [h.get(q_key), a.get(q_key)] for q_key in run_keys}

    calls_data = []
    total_agree_scored = 0
    total_disagree_scored = 0
    total_strict_scored = 0
    total_lenient_scored = 0
    total_comparisons_scored = 0
    score_deltas = []

    for call_id in sorted(answer_data.keys()):
        cd = answer_data[call_id]
        h_score = calc_score({q: cd[q][0] for q in run_keys})
        ai_answers = {q: cd[q][1] for q in run_keys}
        has_ai = any(v is not None for q, v in ai_answers.items() if q in run_scored)
        a_score = calc_score(ai_answers) if has_ai else None

        strict, lenient = 0, 0
        details = []

        for q in run_keys:
            h_val, a_val = cd[q]
            if a_val is None or h_val is None:
                continue
            is_dq = q in DQ_KEYS
            if not is_dq:
                total_comparisons_scored += 1
            if h_val == a_val:
                if not is_dq:
                    total_agree_scored += 1
            else:
                if not is_dq:
                    total_disagree_scored += 1
                w = WEIGHTS.get(q, "DQ")
                if h_val == 1 and a_val == 0:
                    strict += 1
                    if not is_dq:
                        total_strict_scored += 1
                    details.append({
                        "q": Q_LABELS[q], "w": "DQ" if is_dq else w, "ai": "No", "h": "Yes",
                        "reason": STRICT_REASONS.get(q, "AI missed this behavior."), "key": q,
                    })
                else:
                    lenient += 1
                    if not is_dq:
                        total_lenient_scored += 1
                    details.append({
                        "q": Q_LABELS[q], "w": "DQ" if is_dq else w, "ai": "Yes", "h": "No",
                        "reason": LENIENT_REASONS.get(q, "AI over-credited this behavior."), "key": q,
                    })

        weight_lost = sum(d["w"] for d in details if d["w"] != "DQ")
        has_dq_disagree = any(d["w"] == "DQ" for d in details)
        if a_score is not None:
            score_deltas.append(abs(h_score - a_score))

        total_disagree_call = strict + lenient
        if total_disagree_call == 0:
            note = "Perfect agreement — AI and human matched on every question."
            rec = "No action needed — perfect agreement."
        else:
            if strict > 0 and lenient > 0:
                note = f"{total_disagree_call} disagreements: {strict} strict, {lenient} lenient."
            elif strict > 0:
                note = f"{strict} disagreement{'s' if strict > 1 else ''} — all AI too strict."
            else:
                note = f"{lenient} disagreement{'s' if lenient > 1 else ''} — all AI too lenient."
            strict_qs = sorted([d for d in details if d["ai"] == "No" and d["w"] != "DQ"], key=lambda x: -x["w"])[:3]
            lenient_qs = sorted([d for d in details if d["ai"] == "Yes" and d["w"] != "DQ"], key=lambda x: -x["w"])[:3]
            rec_parts = []
            if strict_qs:
                rec_parts.append("Strict fixes: " + ", ".join(f"{d['q']} ({d['w']}pts)" for d in strict_qs))
            if lenient_qs:
                rec_parts.append("Lenient fixes: " + ", ".join(f"{d['q']} ({d['w']}pts)" for d in lenient_qs))
            rec = ". ".join(rec_parts) + "."

        calls_data.append({
            "id": call_id, "human": h_score, "ai": a_score if a_score is not None else 0,
            "strict": strict, "lenient": lenient, "weightLost": weight_lost,
            "disqualifier": has_dq_disagree, "note": note, "details": details, "recommendation": rec,
            "incomplete_ai": not has_ai,
        })

    questions_data = []
    for q_key in run_keys:
        agree = disagree = strict = lenient = total = 0
        for call_id, cd in answer_data.items():
            h, a = cd[q_key]
            if a is None or h is None:
                continue
            total += 1
            if h == a:
                agree += 1
            elif h == 1 and a == 0:
                disagree += 1
                strict += 1
            else:
                disagree += 1
                lenient += 1
        if disagree == 0:
            continue
        if strict > lenient:
            direction = "strict"
        elif lenient > strict:
            direction = "lenient"
        else:
            direction = "mixed"
        questions_data.append({
            "short": q_key,
            "label": Q_FULL_LABELS.get(q_key, Q_LABELS[q_key]),
            "weight": WEIGHTS.get(q_key, 0),
            "agree": agree, "disagree": disagree, "total": total,
            "strict": strict, "lenient": lenient, "dir": direction,
        })
    questions_data.sort(key=lambda q: -q["disagree"])

    agreement_pct = round((total_agree_scored / total_comparisons_scored) * 100, 1) if total_comparisons_scored else 0
    avg_delta = round(sum(score_deltas) / len(score_deltas), 1) if score_deltas else 0

    return {
        "answer_data": answer_data,
        "calls": calls_data,
        "questions": questions_data,
        "meta": {
            "agreement": agreement_pct,
            "totalDisagreements": total_disagree_scored,
            "avgDelta": avg_delta,
            "strictErrors": total_strict_scored,
            "lenientErrors": total_lenient_scored,
            "target": 90,
        },
        "total_comparisons": total_comparisons_scored,
    }


all_runs = []
for run_info in AI_TABS:
    tab_name = run_info["tab"]
    if tab_name not in ai_runs:
        continue
    print(f"\n{'=' * 70}")
    print(f"BUILDING {run_info['label']} (tab: {tab_name})")
    print(f"{'=' * 70}")
    run_ids = matching_ids(tab_name)
    print(f"  Matching IDs: {len(run_ids)}")
    rd = build_run_data(human_data, ai_runs[tab_name], run_ids, run_info)
    rd["info"] = run_info
    all_runs.append(rd)
    m = rd["meta"]
    print(f"  Agreement: {m['agreement']}% ({m['agreement'] * rd['total_comparisons'] / 100:.0f}/{rd['total_comparisons']})")
    print(f"  Total disagreements: {m['totalDisagreements']}")
    print(f"  Strict errors: {m['strictErrors']}")
    print(f"  Lenient errors: {m['lenientErrors']}")
    print(f"  Avg score delta: {m['avgDelta']}%")


def build_run_js(run_data):
    info = run_data["info"]
    meta = run_data["meta"]
    calls_data = run_data["calls"]
    questions_data = run_data["questions"]
    n_calls = len(calls_data)

    meta_js = f"""    meta: {{
      agreement: {meta['agreement']},
      totalDisagreements: {meta['totalDisagreements']},
      avgDelta: {meta['avgDelta']},
      strictErrors: {meta['strictErrors']},
      lenientErrors: {meta['lenientErrors']},
      target: 90
    }}"""

    total_comparisons = run_data["total_comparisons"]
    total_disagree = meta["totalDisagreements"]
    strict_top = sorted([q for q in questions_data if q["strict"] > q["lenient"]], key=lambda x: -x["disagree"])[:5]
    lenient_top = sorted([q for q in questions_data if q["lenient"] > q["strict"]], key=lambda x: -x["disagree"])[:3]

    strict_summary = ", ".join(
        f"<em>{q['label']}</em> ({q['agree']}/{q['total']} = {round(q['agree']/q['total']*100)}%)"
        for q in strict_top[:3]
    ) or "none"
    lenient_summary = ", ".join(
        f"<em>{q['label']}</em> ({q['agree']}/{q['total']} = {round(q['agree']/q['total']*100)}%)"
        for q in lenient_top[:3]
    ) or "none"

    bias_pct = round(meta["strictErrors"] / total_disagree * 100) if total_disagree else 0
    dominant = "too strict" if meta["strictErrors"] > meta["lenientErrors"] else "too lenient"
    key_findings = (
        f'<p>The AI disagrees with humans on <strong>{round(100 - meta["agreement"], 1)}% of question-level answers</strong> '
        f'across these {n_calls} resident calls ({total_disagree} of {total_comparisons} scored comparisons). '
        f'The dominant pattern is the AI being <strong>{dominant}</strong> '
        f'— {bias_pct}% of disagreements ({meta["strictErrors"]} of {total_disagree}).</p>'
        f'<p>Worst strict questions: {strict_summary}. '
        f'Worst lenient questions: {lenient_summary}.</p>'
        f'<p>Average absolute score delta is <strong>{meta["avgDelta"]} percentage points</strong> (weighted, 42-pt card).</p>'
        f'<p><em>Note: Partial benchmark — {n_calls} matching AI+human call IDs. Evidence/transcripts deferred.</em></p>'
    )
    key_findings_js = f"    keyFindings: `{key_findings}`"

    # Seed recommendations from top disagreement drivers
    recs = []
    for i, q in enumerate(strict_top[:3], 1):
        recs.append({
            "num": i,
            "title": f"Tighten {q['label']} detection — {q['strict']} strict errors ({q['agree']}/{q['total']} agree)",
            "severity": "critical" if q["disagree"] >= 5 else "warning",
            "severityLabel": f"{q['strict']} strict / {q['lenient']} lenient",
            "owner": "AI Engineering",
            "ownerClass": "info",
            "problem": f"<p><strong>{q['label']}</strong> is a top strict driver "
                       f"({q['agree']}/{q['total']} agreement). AI under-credits behavior humans marked Yes.</p>",
            "action": "<p>Review protocol text for this question against the resident fundamentals sheet and loosen "
                      "detection thresholds where humans consistently credit the behavior.</p>",
            "protocols": [{
                "label": q["label"],
                "current": "(from AI Resident Fundamentals)",
                "recommended": "Loosen detection so humans' Yes cases are credited. Replace this stub after analysis.",
            }],
        })
    for j, q in enumerate(lenient_top[:2], len(recs) + 1):
        recs.append({
            "num": j,
            "title": f"Reduce false positives on {q['label']} — {q['lenient']} lenient errors",
            "severity": "warning",
            "severityLabel": f"{q['lenient']} lenient",
            "owner": "AI Engineering",
            "ownerClass": "info",
            "problem": f"<p><strong>{q['label']}</strong> shows AI over-crediting vs human "
                       f"({q['agree']}/{q['total']} agreement).</p>",
            "action": "<p>Add negative examples to the protocol so AI does not credit weak or N/A cases.</p>",
        })

    rec_lines = []
    for r in recs:
        rec_lines.append(
            f"""      {{
        num: {r['num']},
        title: "{js_escape(r['title'])}",
        severity: "{r['severity']}",
        severityLabel: "{js_escape(r['severityLabel'])}",
        owner: "{r['owner']}",
        ownerClass: "{r['ownerClass']}",
        problem: `{r['problem']}`,
        action: `{r['action']}`
      }}"""
        )
    recommendations_js = "    recommendations: [\n" + ",\n".join(rec_lines) + "\n    ]"

    q_lines = []
    for q in questions_data:
        q_lines.append(
            f'      {{ short: "{q["short"]}", label: "{js_escape(q["label"])}", weight: {q["weight"]}, '
            f'agree: {q["agree"]}, disagree: {q["disagree"]}, total: {q["total"]}, '
            f'lenient: {q["lenient"]}, strict: {q["strict"]}, dir: "{q["dir"]}",\n'
            f'        evidence: []}}'
        )
    questions_js = "    questions: [\n" + ",\n".join(q_lines) + "\n    ]"

    calls_lines = []
    for c in calls_data:
        details_parts = []
        for d in c["details"]:
            w_js = f'"{d["w"]}"' if d["w"] == "DQ" else str(d["w"])
            details_parts.append(
                f'          {{ q: "{js_escape(d["q"])}", w: {w_js}, ai: "{d["ai"]}", h: "{d["h"]}", reason: "{js_escape(d["reason"])}" }}'
            )
        details_str = ",\n".join(details_parts)
        if details_str:
            details_str = "\n" + details_str + "\n        "
        calls_lines.append(
            f'      {{ id: "{c["id"]}", human: {c["human"]}, ai: {c["ai"]}, strict: {c["strict"]}, lenient: {c["lenient"]}, weightLost: {c["weightLost"]}, disqualifier: {"true" if c["disqualifier"] else "false"},\n'
            f'        note: \'{js_escape(c["note"])}\',\n'
            f'        details: [{details_str}],\n'
            f'        recommendation: \'{js_escape(c["recommendation"])}\' }}'
        )
    calls_js = "    calls: [\n" + ",\n".join(calls_lines) + "\n    ]"

    return f"""  {{
    id: {info['id']},
    label: "{info['label']}",
    date: "{info['date']}",
    description: "{js_escape(info['description'])}",
    changes: "{js_escape(info['changes'])}",
{meta_js},
{key_findings_js},
{questions_js},
{calls_js},
{recommendations_js},
    rootCause: {{
      summary: 'Resident Run 1.0 baseline on a 42-point weighted scorecard (10 scored questions + FHA/secure-info DQs). Partial set of {n_calls} matching calls.',
      strictDetail: '<p>Strict errors concentrate on hold permission, closing, and ownership/acknowledgment — AI under-credits behaviors humans marked Yes.</p>',
      lenientDetail: '<p>Lenient errors are fewer and spread across secure-info DQ, validate-concern, and neutral language.</p>'
    }}
  }}"""


run_answer_data_parts = []
for rd in all_runs:
    run_id = rd["info"]["id"]
    ad = rd["answer_data"]
    ad_lines = []
    for call_id in sorted(ad.keys()):
        cd = ad[call_id]
        parts = []
        for q in cd.keys():
            h, a = cd[q]
            h_str = str(h) if h is not None else "null"
            a_str = str(a) if a is not None else "null"
            parts.append(f"{q}:[{h_str},{a_str}]")
        ad_lines.append(f'    "{call_id}": {{{",".join(parts)}}}')
    run_answer_data_parts.append(f"  {run_id}: {{\n" + ",\n".join(ad_lines) + "\n  }")

resident_run_answer_data_js = "const residentRunAnswerData = {\n" + ",\n".join(run_answer_data_parts) + "\n};"

ad = all_runs[0]["answer_data"]
ad_lines = []
for call_id in sorted(ad.keys()):
    cd = ad[call_id]
    parts = []
    for q in ALL_KEYS:
        h, a = cd[q]
        h_str = str(h) if h is not None else "null"
        a_str = str(a) if a is not None else "null"
        parts.append(f"{q}:[{h_str},{a_str}]")
    ad_lines.append(f'  "{call_id}": {{{",".join(parts)}}}')
resident_answer_data_js = "const residentAnswerData = {\n" + ",\n".join(ad_lines) + "\n};"

runs_js_parts = [build_run_js(rd) for rd in all_runs]
resident_runs_js = "const residentRuns = [\n" + ",\n".join(runs_js_parts) + "\n];"

resident_matrix_js = """const residentMatrixQuestions = [
  { short: "Greeting", key: "greeting" },
  { short: "Name", key: "name_usage" },
  { short: "Contact info", key: "contact_info" },
  { short: "Unit #", key: "unit_number" },
  { short: "Open-ended Qs", key: "open_ended_qs" },
  { short: "Acknowledged", key: "acknowledged" },
  { short: "Closing", key: "closing" },
  { short: "Hold perm.", key: "hold_permission" },
  { short: "Validate concern", key: "validate_concern" },
  { short: "Neutral lang", key: "neutral_language" },
  { short: "FHA (DQ)", key: "fha" },
  { short: "Secure info (DQ)", key: "secure_info" },
];"""

n_calls = max(len(rd["calls"]) for rd in all_runs) if all_runs else len(benchmark_call_ids)
resident_lead_config = f"""  resident: {{
    label: 'Residents',
    callCount: {n_calls},
    criteriaCount: 12,
    totalPoints: {TOTAL_POINTS},
    scoredQuestions: {len(SCORED_KEYS)},
    runs: residentRuns,
    answerData: residentAnswerData,
    runAnswerData: residentRunAnswerData,
    matrixQuestions: residentMatrixQuestions,
    subtitle: (r) => r ? `${{(r.calls && r.calls.length) ? r.calls.length : {n_calls}}} resident calls \\u00b7 12 scoring criteria \\u00b7 ${{r.label}} \\u2014 ${{r.date}}` : `Run 1.0: 12 calls \\u00b7 Run 2.0–5.0 + 8.0: 20 calls \\u00b7 12 scoring criteria`,
    formulaDesc: '{TOTAL_POINTS} total weighted points across {len(SCORED_KEYS)} regular questions (weights from AI Resident Fundamentals). Score = (earned points / {TOTAL_POINTS}) \\u00d7 100%.',
    dqDesc: 'Each disqualifier triggered (FHA violation, secure info disclosure) applies a 20% reduction.',
  }}"""

# ── Patch the HTML ──
print(f"\nPatching HTML...")
with open(HTML_PATH, "r") as f:
    html = f.read()


def replace_const_object(html_text, const_name, new_js):
    needle = f"const {const_name} = {{"
    if needle not in html_text:
        return html_text, False
    start = html_text.index(needle)
    # find matching closing };
    depth = 0
    i = start + len(needle) - 1
    while i < len(html_text):
        ch = html_text[i]
        if ch == "{":
            depth += 1
        elif ch == "}":
            depth -= 1
            if depth == 0:
                end = i + 1
                if end < len(html_text) and html_text[end] == ";":
                    end += 1
                return html_text[:start] + new_js + html_text[end:], True
        i += 1
    return html_text, False


def replace_const_array(html_text, const_name, new_js):
    needle = f"const {const_name} = ["
    if needle not in html_text:
        return html_text, False
    start = html_text.index(needle)
    depth = 0
    i = start + len(needle) - 1
    while i < len(html_text):
        ch = html_text[i]
        if ch == "[":
            depth += 1
        elif ch == "]":
            depth -= 1
            if depth == 0:
                end = i + 1
                if end < len(html_text) and html_text[end] == ";":
                    end += 1
                return html_text[:start] + new_js + html_text[end:], True
        i += 1
    return html_text, False


# Insert / replace resident data blocks after warm matrix
marker = "const warmMatrixQuestions = ["
if "const residentAnswerData = {" in html:
    html, ok = replace_const_object(html, "residentAnswerData", resident_answer_data_js)
    print("✅ Replaced residentAnswerData" if ok else "❌ Failed residentAnswerData")
else:
    # insert after warmMatrixQuestions array
    start = html.index(marker)
    depth = 0
    i = start + len(marker) - 1
    while i < len(html):
        if html[i] == "[":
            depth += 1
        elif html[i] == "]":
            depth -= 1
            if depth == 0:
                end = i + 1
                if end < len(html) and html[end] == ";":
                    end += 1
                block = (
                    "\n\n// ═══ RESIDENT DATA ═══\n"
                    + resident_answer_data_js
                    + "\n\n"
                    + resident_run_answer_data_js
                    + "\n\n"
                    + resident_runs_js
                    + "\n\n"
                    + resident_matrix_js
                    + "\n"
                )
                html = html[:end] + block + html[end:]
                print("✅ Inserted resident data blocks")
                break
        i += 1

if "const residentRunAnswerData = {" in html and "Inserted resident" not in open(HTML_PATH).read():
    # If we replaced answer data only, also refresh the rest
    html, ok = replace_const_object(html, "residentRunAnswerData", resident_run_answer_data_js)
    print("✅ Replaced residentRunAnswerData" if ok else "… residentRunAnswerData insert path")
    html, ok = replace_const_array(html, "residentRuns", resident_runs_js)
    print("✅ Replaced residentRuns" if ok else "… residentRuns insert path")
    html, ok = replace_const_array(html, "residentMatrixQuestions", resident_matrix_js)
    print("✅ Replaced residentMatrixQuestions" if ok else "… residentMatrix insert path")

# Refresh if blocks already existed from a prior run of this script
if "// ═══ RESIDENT DATA ═══" in html:
    html, _ = replace_const_object(html, "residentAnswerData", resident_answer_data_js)
    html, _ = replace_const_object(html, "residentRunAnswerData", resident_run_answer_data_js)
    html, _ = replace_const_array(html, "residentRuns", resident_runs_js)
    html, _ = replace_const_array(html, "residentMatrixQuestions", resident_matrix_js)
    print("✅ Refreshed resident data blocks")

# Patch leadConfig: add resident entry
if "resident: {" in html and "runs: residentRuns" in html:
    # replace existing resident leadConfig entry
    m = re.search(r"  resident: \{[\s\S]*?\n  \}", html)
    if m and "runs: residentRuns" in m.group(0):
        html = html[: m.start()] + resident_lead_config + html[m.end() :]
        print("✅ Replaced resident leadConfig")
    else:
        # insert before closing of leadConfig
        cfg_end = html.index("\n};", html.index("const leadConfig = {"))
        # find warm block end and insert after
        warm_m = re.search(r"  warm: \{[\s\S]*?\n  \}", html)
        if warm_m:
            insert_at = warm_m.end()
            html = html[:insert_at] + ",\n" + resident_lead_config + html[insert_at:]
            print("✅ Inserted resident leadConfig after warm")
else:
    warm_m = re.search(r"  warm: \{[\s\S]*?\n  \}", html)
    if not warm_m:
        print("❌ Could not find warm leadConfig to insert after")
        sys.exit(1)
    insert_at = warm_m.end()
    html = html[:insert_at] + ",\n" + resident_lead_config + html[insert_at:]
    print("✅ Inserted resident leadConfig after warm")

# Remove resident from PENDING_TYPES
old_pending = """const PENDING_TYPES = {
  workorder: { label: 'Work Orders', icon: '🔧' },
  resident: { label: 'Residents', icon: '🏠' },
  other: { label: 'Other', icon: '📋' },
};"""
new_pending = """const PENDING_TYPES = {
  workorder: { label: 'Work Orders', icon: '🔧' },
  other: { label: 'Other', icon: '📋' },
};"""
if old_pending in html:
    html = html.replace(old_pending, new_pending)
    print("✅ Removed resident from PENDING_TYPES")
elif "resident: { label: 'Residents'" in html:
    html = re.sub(
        r"\n\s*resident: \{ label: 'Residents', icon: '🏠' \},?",
        "",
        html,
        count=1,
    )
    print("✅ Removed resident PENDING_TYPES line")
else:
    print("… PENDING_TYPES already without resident (or unexpected format)")

with open(HTML_PATH, "w") as f:
    f.write(html)

print(f"\n✅ Wrote updated HTML to {HTML_PATH}")
print(f"   Resident runs: {len(all_runs)}")
for rd in all_runs:
    m = rd["meta"]
    print(
        f"   {rd['info']['label']}: {m['agreement']}% agreement, "
        f"{m['totalDisagreements']} disagreements, {m['avgDelta']}% avg delta, "
        f"{len(rd['calls'])} calls"
    )
